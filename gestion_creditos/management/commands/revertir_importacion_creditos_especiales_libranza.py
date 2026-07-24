import csv
import json
import unicodedata
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from gestion_creditos.management.commands.importar_creditos_especiales_libranza import (
    HEADER_MAP,
    LEGACY_IMPORT_OBSERVATION,
    MONTHS_ES,
)
from gestion_creditos.models import Credito, Empresa


def _normalize_header(value):
    text = unicodedata.normalize('NFKD', str(value or '')).encode('ascii', 'ignore').decode('ascii')
    text = text.strip().lower()
    return ' '.join(text.split())


class Command(BaseCommand):
    help = (
        'Revierte una importacion de creditos especiales legacy de libranza '
        'de forma controlada, usando el archivo fuente para identificar coincidencias.'
    )

    def add_arguments(self, parser):
        parser.add_argument('--archivo', required=True, help='Ruta local al archivo CSV o XLSX usado para importar.')
        parser.add_argument('--empresa-nombre', required=True, help='Nombre exacto de la empresa destino.')
        parser.add_argument('--apply', action='store_true', help='Aplica la eliminacion. Sin esta bandera solo hace dry-run.')
        parser.add_argument('--report-file', help='Ruta opcional para guardar reporte JSON.')

    def handle(self, *args, **options):
        empresa = self._resolver_empresa(options['empresa_nombre'])
        rows = list(self._iter_rows(options['archivo']))
        if not rows:
            raise CommandError('El archivo no contiene filas.')

        matches = []
        for index, row in enumerate(rows, start=2):
            normalized = self._normalize_row(row)
            credito = self._find_credit(empresa, normalized)
            matches.append({
                'fila': index,
                'documento': normalized['documento'],
                'nombre': normalized['nombre'],
                'monto_inicial': str(normalized['monto_inicial']) if normalized['monto_inicial'] is not None else '',
                'fecha_primer_pago': str(normalized['fecha_primer_pago']) if normalized['fecha_primer_pago'] else '',
                'encontrado': bool(credito),
                'numero_credito': credito.numero_credito if credito else '',
                'estado': credito.estado if credito else '',
                'historial_pagos': credito.historial_pagos.count() if credito else 0,
                'wompi_intentos': credito.wompi_intentos.count() if credito else 0,
                'can_delete': bool(credito and credito.historial_pagos.count() == 0 and credito.wompi_intentos.count() == 0),
                'motivos_bloqueo': self._blockers(credito),
            })

        report = {
            'mode': 'apply' if options['apply'] else 'dry-run',
            'empresa': empresa.nombre,
            'total_filas': len(matches),
            'encontrados': sum(1 for row in matches if row['encontrado']),
            'eliminables': sum(1 for row in matches if row['can_delete']),
            'filas': matches,
        }

        if not options['apply']:
            self.stdout.write(self.style.WARNING('Dry-run: no se eliminaran creditos.'))
            self._print_summary(matches)
            self._write_report(options.get('report_file'), report)
            return

        deleted = []
        with transaction.atomic():
            for match in matches:
                if not match['can_delete']:
                    continue
                credito = Credito.objects.get(numero_credito=match['numero_credito'])
                deleted.append(credito.numero_credito)
                credito.delete()

        report['deleted'] = deleted
        self._print_summary(matches, deleted=deleted)
        self._write_report(options.get('report_file'), report)

    def _print_summary(self, matches, deleted=None):
        for match in matches:
            status = 'OK' if match['can_delete'] else ('ENCONTRADO' if match['encontrado'] else 'NO_MATCH')
            self.stdout.write(
                f"Fila {match['fila']}: {status} | documento={match['documento']} | "
                f"nombre={match['nombre']} | credito={match['numero_credito'] or '-'}"
            )
            if match['motivos_bloqueo']:
                self.stdout.write(f"  Bloqueos: {', '.join(match['motivos_bloqueo'])}")
        if deleted:
            self.stdout.write(self.style.SUCCESS(f'Se eliminaron {len(deleted)} creditos importados legacy.'))

    def _write_report(self, report_file, payload):
        if not report_file:
            return
        with open(report_file, 'w', encoding='utf-8') as fh:
            json.dump(payload, fh, ensure_ascii=False, indent=2, default=str)
        self.stdout.write(self.style.SUCCESS(f'Reporte guardado en {report_file}'))

    def _resolver_empresa(self, empresa_nombre):
        empresa = Empresa.objects.filter(nombre__iexact=empresa_nombre.strip()).first()
        if not empresa:
            raise CommandError(f'No existe una empresa con nombre exacto "{empresa_nombre}".')
        return empresa

    def _iter_rows(self, path):
        extension = path.rsplit('.', 1)[-1].lower() if '.' in path else ''
        if extension == 'csv':
            with open(path, 'r', encoding='utf-8-sig', newline='') as fh:
                reader = csv.DictReader(fh)
                for row in reader:
                    yield self._map_headers(row)
            return
        if extension == 'xlsx':
            from openpyxl import load_workbook

            workbook = load_workbook(path, read_only=True, data_only=True)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                return
            headers = [_normalize_header(value) for value in rows[0]]
            for values in rows[1:]:
                raw = {headers[idx]: values[idx] for idx in range(min(len(headers), len(values)))}
                if any(value not in (None, '') for value in raw.values()):
                    yield self._map_headers(raw)
            return
        raise CommandError('Usa un archivo CSV o XLSX.')

    def _map_headers(self, row):
        normalized = {}
        for key, value in row.items():
            mapped = HEADER_MAP.get(_normalize_header(key))
            if mapped:
                normalized[mapped] = value
        return normalized

    def _normalize_row(self, row):
        documento = ''.join(ch for ch in str(row.get('documento') or '') if ch.isdigit())
        return {
            'documento': documento,
            'nombre': str(row.get('nombre') or '').strip(),
            'monto_inicial': self._parse_money(row.get('monto_inicial')),
            'fecha_primer_pago': self._parse_date(row.get('fecha_primer_pago')),
            'nro_cuotas': self._parse_int(row.get('nro_cuotas')),
        }

    def _find_credit(self, empresa, row):
        if not row['documento'] or row['monto_inicial'] is None or row['fecha_primer_pago'] is None or row['nro_cuotas'] is None:
            return None
        return (
            Credito.objects.filter(
                linea=Credito.LineaCredito.LIBRANZA,
                tipo_regla_credito=Credito.TipoReglaCredito.ESPECIAL,
                observacion_regla_especial=LEGACY_IMPORT_OBSERVATION,
                detalle_libranza__empresa=empresa,
                detalle_libranza__cedula=row['documento'],
                monto_aprobado=row['monto_inicial'],
                plazo=row['nro_cuotas'],
                fecha_primera_cuota_forzada=row['fecha_primer_pago'],
            )
            .select_related('detalle_libranza')
            .first()
        )

    def _blockers(self, credito):
        if not credito:
            return []
        blockers = []
        if credito.historial_pagos.exists():
            blockers.append('tiene_historial_pagos')
        if credito.wompi_intentos.exists():
            blockers.append('tiene_wompi_intentos')
        return blockers

    def _parse_date(self, raw):
        if isinstance(raw, datetime):
            return raw.date()
        if isinstance(raw, date):
            return raw
        value = str(raw or '').strip()
        if not value:
            return None
        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
        parts = value.replace('.', '-').replace('/', '-').split('-')
        if len(parts) == 3:
            day, month_raw, year = parts
            month = MONTHS_ES.get(month_raw[:3].lower())
            if month:
                year_num = int(year)
                if year_num < 100:
                    year_num += 2000
                try:
                    return datetime(year_num, month, int(day)).date()
                except ValueError:
                    return None
        return None

    def _parse_money(self, raw):
        value = str(raw or '').strip()
        if not value:
            return None
        normalized = value.replace('$', '').replace(' ', '').replace('.', '').replace(',', '.')
        try:
            return Decimal(normalized)
        except (InvalidOperation, ValueError):
            return None

    def _parse_int(self, raw):
        value = ''.join(ch for ch in str(raw or '') if ch.isdigit())
        if not value:
            return None
        return int(value)
