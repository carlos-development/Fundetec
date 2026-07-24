import csv
import io
import json
import unicodedata
from datetime import datetime
from decimal import Decimal, InvalidOperation

from dateutil.relativedelta import relativedelta
from django.core.management.base import BaseCommand, CommandError


MONTHS_ES = {
    'ene': 1,
    'feb': 2,
    'mar': 3,
    'abr': 4,
    'may': 5,
    'jun': 6,
    'jul': 7,
    'ago': 8,
    'sep': 9,
    'oct': 10,
    'nov': 11,
    'dic': 12,
}


def _normalize_header(value):
    text = unicodedata.normalize('NFKD', str(value or '')).encode('ascii', 'ignore').decode('ascii')
    text = text.strip().lower()
    return ' '.join(text.split())


HEADER_MAP = {
    'cedula': 'documento',
    'nombre': 'nombre',
    'fecha desembolso': 'fecha_desembolso',
    'fecha primer pago': 'fecha_primer_pago',
    'cuota': 'cuota',
    'nro cuotas': 'nro_cuotas',
    'cuotas pagadas': 'cuotas_pagadas',
    'cuotas pendientes': 'cuotas_pendientes',
}


class Command(BaseCommand):
    help = (
        'Previsualiza e identifica faltantes de una base legacy de creditos '
        'especiales de libranza antes de crear registros reales.'
    )

    def add_arguments(self, parser):
        parser.add_argument('--archivo', required=True, help='Ruta local a archivo CSV o XLSX.')
        parser.add_argument('--report-file', help='Ruta opcional para guardar reporte JSON.')

    def handle(self, *args, **options):
        rows = list(self._iter_rows(options['archivo']))
        if not rows:
            raise CommandError('El archivo no contiene filas.')

        normalized = [self._normalize_row(row, index + 2) for index, row in enumerate(rows)]
        summary = {
            'total_filas': len(normalized),
            'filas_validas_previsualizacion': sum(1 for row in normalized if not row['errores']),
            'filas_con_error': sum(1 for row in normalized if row['errores']),
            'filas': normalized,
            'faltantes_operativos_globales': [
                'empresa o convenio del empleado',
                'monto aprobado real / capital inicial',
                'tasa mensual real del credito',
                'porcentaje de comision real',
                'correo, telefono y direccion si se quiere alta completa del detalle',
            ],
        }

        for row in normalized:
            estado = 'OK' if not row['errores'] else 'ERROR'
            self.stdout.write(
                f"Fila {row['fila']}: {estado} | documento={row['documento']} | "
                f"nombre={row['nombre']} | proximo_pago={row['fecha_proximo_pago_inferida'] or '-'}"
            )
            if row['faltantes']:
                self.stdout.write(f"  Faltantes: {', '.join(row['faltantes'])}")
            if row['errores']:
                self.stdout.write(f"  Errores: {', '.join(row['errores'])}")

        if options.get('report_file'):
            with open(options['report_file'], 'w', encoding='utf-8') as fh:
                json.dump(summary, fh, ensure_ascii=False, indent=2, default=str)
            self.stdout.write(self.style.SUCCESS(f"Reporte guardado en {options['report_file']}"))

    def _iter_rows(self, path):
        extension = path.rsplit('.', 1)[-1].lower() if '.' in path else ''
        if extension == 'csv':
            with open(path, 'r', encoding='utf-8-sig', newline='') as fh:
                reader = csv.DictReader(fh)
                for row in reader:
                    yield self._map_headers(row)
            return
        if extension == 'xlsx':
            from openpyxl import load_workbook

            workbook = load_workbook(path, read_only=True, data_only=True)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                return
            headers = [self._normalized_key(value) for value in rows[0]]
            for values in rows[1:]:
                raw = {headers[idx]: values[idx] for idx in range(min(len(headers), len(values)))}
                yield self._map_headers(raw)
            return
        raise CommandError('Usa un archivo CSV o XLSX.')

    def _map_headers(self, row):
        normalized = {}
        for key, value in row.items():
            mapped = HEADER_MAP.get(self._normalized_key(key))
            if mapped:
                normalized[mapped] = value
        return normalized

    def _normalized_key(self, value):
        return _normalize_header(value)

    def _normalize_row(self, row, fila):
        errores = []
        documento = ''.join(ch for ch in str(row.get('documento') or '') if ch.isdigit())
        nombre = str(row.get('nombre') or '').strip()

        fecha_desembolso = self._parse_date(row.get('fecha_desembolso'), 'fecha_desembolso', errores)
        fecha_primer_pago = self._parse_date(row.get('fecha_primer_pago'), 'fecha_primer_pago', errores)
        cuota = self._parse_money(row.get('cuota'), 'cuota', errores)
        nro_cuotas = self._parse_int(row.get('nro_cuotas'), 'nro_cuotas', errores)
        cuotas_pagadas = self._parse_int(row.get('cuotas_pagadas'), 'cuotas_pagadas', errores)
        cuotas_pendientes = self._parse_int(row.get('cuotas_pendientes'), 'cuotas_pendientes', errores)

        if not documento:
            errores.append('Documento obligatorio.')
        if not nombre:
            errores.append('Nombre obligatorio.')
        if nro_cuotas is not None and cuotas_pagadas is not None and cuotas_pendientes is not None:
            if nro_cuotas != cuotas_pagadas + cuotas_pendientes:
                errores.append('La suma de cuotas pagadas y pendientes no coincide con el total.')

        fecha_proximo_pago = None
        if fecha_primer_pago and cuotas_pagadas is not None:
            fecha_proximo_pago = fecha_primer_pago + relativedelta(months=cuotas_pagadas)

        faltantes = [
            'empresa/convenio',
            'monto_aprobado_real',
            'tasa_mensual',
            'porcentaje_comision',
        ]

        return {
            'fila': fila,
            'documento': documento,
            'nombre': nombre,
            'fecha_desembolso': fecha_desembolso,
            'fecha_primer_pago': fecha_primer_pago,
            'fecha_proximo_pago_inferida': fecha_proximo_pago,
            'cuota': cuota,
            'nro_cuotas': nro_cuotas,
            'cuotas_pagadas': cuotas_pagadas,
            'cuotas_pendientes': cuotas_pendientes,
            'faltantes': faltantes,
            'errores': errores,
        }

    def _parse_date(self, raw, field_name, errores):
        value = str(raw or '').strip()
        if not value:
            errores.append(f'{field_name} es obligatorio.')
            return None
        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
        parts = value.replace('.', '-').replace('/', '-').split('-')
        if len(parts) == 3:
            day, month_raw, year = parts
            month = MONTHS_ES.get(month_raw[:3].lower())
            if month:
                year_num = int(year)
                if year_num < 100:
                    year_num += 2000
                try:
                    return datetime(year_num, month, int(day)).date()
                except ValueError:
                    pass
        errores.append(f'{field_name} no tiene un formato valido.')
        return None

    def _parse_money(self, raw, field_name, errores):
        value = str(raw or '').strip()
        if not value:
            errores.append(f'{field_name} es obligatorio.')
            return None
        normalized = value.replace('$', '').replace(' ', '').replace('.', '').replace(',', '.')
        try:
            return Decimal(normalized)
        except (InvalidOperation, ValueError):
            errores.append(f'{field_name} no es numerico.')
            return None

    def _parse_int(self, raw, field_name, errores):
        value = ''.join(ch for ch in str(raw or '') if ch.isdigit())
        if not value:
            errores.append(f'{field_name} es obligatorio.')
            return None
        return int(value)
