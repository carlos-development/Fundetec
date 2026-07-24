import csv
import io
import json
import unicodedata
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation

from django.contrib.auth import get_user_model
from django.core.files.base import ContentFile
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone
from django.utils.text import slugify

from gestion_creditos import credit_services
from gestion_creditos.management.commands.crear_credito_especial_libranza import (
    MINIMAL_PDF_BYTES,
    MINIMAL_PNG_BYTES,
)
from gestion_creditos.models import (
    Credito,
    CreditoLibranza,
    CuotaAmortizacion,
    Empresa,
    HistorialEstado,
    VinculoLaboralEmpresa,
)
from gestion_creditos.services.name_normalization import normalize_name_upper


MONTHS_ES = {
    'ene': 1,
    'feb': 2,
    'mar': 3,
    'abr': 4,
    'may': 5,
    'jun': 6,
    'jul': 7,
    'ago': 8,
    'sep': 9,
    'oct': 10,
    'nov': 11,
    'dic': 12,
}


LEGACY_IMPORT_OBSERVATION = 'Credito especial legacy importado por backend desde base controlada de Fertobra.'


def _normalize_header(value):
    text = unicodedata.normalize('NFKD', str(value or '')).encode('ascii', 'ignore').decode('ascii')
    text = text.strip().lower()
    return ' '.join(text.split())


def _repair_mojibake(value):
    text = str(value or '').strip()
    if not text:
        return ''
    suspicious_markers = ('Ã', 'Â', 'â', '€', '™')
    if any(marker in text for marker in suspicious_markers):
        try:
            repaired = text.encode('latin-1').decode('utf-8')
            return repaired.strip()
        except (UnicodeEncodeError, UnicodeDecodeError):
            return text
    return text


HEADER_MAP = {
    'pagare': 'pagare',
    'empresa': 'empresa',
    'cedula': 'documento',
    'nombre': 'nombre',
    'correo': 'correo',
    'celular': 'telefono',
    'monto inicial': 'monto_inicial',
    'fecha desembolso': 'fecha_desembolso',
    'fecha primer pago': 'fecha_primer_pago',
    'cuota': 'cuota',
    'nro cuotas': 'nro_cuotas',
    'cuotas pagadas': 'cuotas_pagadas',
    'cuotas pendientes': 'cuotas_pendientes',
    'tasa': 'tasa',
    'comision real': 'comision_real',
}


REQUIRED_FIELDS = {
    'empresa',
    'documento',
    'nombre',
    'monto_inicial',
    'fecha_desembolso',
    'fecha_primer_pago',
    'cuota',
    'nro_cuotas',
    'cuotas_pagadas',
    'cuotas_pendientes',
    'tasa',
    'comision_real',
}


class Command(BaseCommand):
    help = (
        'Importa una base legacy de creditos especiales de libranza desde CSV/XLSX '
        'con validaciones de duplicado, vinculo laboral y modo dry-run/apply.'
    )

    def add_arguments(self, parser):
        parser.add_argument('--archivo', required=True, help='Ruta local al archivo CSV o XLSX.')
        parser.add_argument(
            '--empresa-nombre',
            required=True,
            help='Nombre exacto de la empresa destino. Se validara contra el archivo.',
        )
        parser.add_argument('--apply', action='store_true', help='Aplica la importacion. Sin esta bandera solo hace dry-run.')
        parser.add_argument('--report-file', help='Ruta opcional para guardar un reporte JSON.')

    def handle(self, *args, **options):
        empresa = self._resolver_empresa(options['empresa_nombre'])
        rows = list(self._iter_rows(options['archivo']))
        if not rows:
            raise CommandError('El archivo no contiene filas para importar.')

        normalized_rows = [self._normalize_row(raw, empresa, index + 2) for index, raw in enumerate(rows)]

        report = {
            'generated_at': timezone.now().isoformat(),
            'mode': 'apply' if options['apply'] else 'dry-run',
            'empresa': empresa.nombre,
            'total_filas': len(normalized_rows),
            'filas_validas': sum(1 for row in normalized_rows if row['can_import']),
            'filas_con_bloqueo': sum(1 for row in normalized_rows if not row['can_import']),
            'filas': normalized_rows,
        }

        if not options['apply']:
            self.stdout.write(self.style.WARNING('Dry-run: no se importaran creditos.'))
            self._print_summary(normalized_rows)
            self._write_report(options.get('report_file'), report)
            return

        created = []
        with transaction.atomic():
            for row in normalized_rows:
                if not row['can_import']:
                    continue
                credito = self._import_row(row, empresa)
                created.append({
                    'numero_credito': credito.numero_credito,
                    'documento': row['documento'],
                    'nombre': row['nombre'],
                })

        report['created'] = created
        self._print_summary(normalized_rows, created=created)
        self._write_report(options.get('report_file'), report)

    def _print_summary(self, rows, created=None):
        for row in rows:
            status = 'OK' if row['can_import'] else 'ERROR'
            existing = f" | existente={row['existing_credit_number']}" if row['existing_credit_number'] else ''
            self.stdout.write(
                f"Fila {row['fila']}: {status} | documento={row['documento']} | "
                f"nombre={row['nombre']} | empresa={row['empresa_archivo']}{existing}"
            )
            if row['warnings']:
                self.stdout.write(f"  Avisos: {', '.join(row['warnings'])}")
            if row['errors']:
                self.stdout.write(f"  Errores: {', '.join(row['errors'])}")
        if created:
            self.stdout.write(self.style.SUCCESS(f'Se crearon {len(created)} creditos especiales legacy.'))

    def _write_report(self, report_file, payload):
        if not report_file:
            return
        with open(report_file, 'w', encoding='utf-8') as fh:
            json.dump(payload, fh, ensure_ascii=False, indent=2, default=str)
        self.stdout.write(self.style.SUCCESS(f'Reporte guardado en {report_file}'))

    def _iter_rows(self, path):
        extension = path.rsplit('.', 1)[-1].lower() if '.' in path else ''
        if extension == 'csv':
            with open(path, 'r', encoding='utf-8-sig', newline='') as fh:
                reader = csv.DictReader(fh)
                for row in reader:
                    yield self._map_headers(row)
            return
        if extension == 'xlsx':
            from openpyxl import load_workbook

            workbook = load_workbook(path, read_only=True, data_only=True)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                return
            headers = [self._normalized_key(value) for value in rows[0]]
            for values in rows[1:]:
                raw = {headers[idx]: values[idx] for idx in range(min(len(headers), len(values)))}
                if any(value not in (None, '') for value in raw.values()):
                    yield self._map_headers(raw)
            return
        raise CommandError('Usa un archivo CSV o XLSX.')

    def _map_headers(self, row):
        normalized = {}
        for key, value in row.items():
            mapped = HEADER_MAP.get(self._normalized_key(key))
            if mapped:
                normalized[mapped] = value
        return normalized

    def _normalized_key(self, value):
        return _normalize_header(value)

    def _resolver_empresa(self, empresa_nombre):
        empresa = Empresa.objects.filter(nombre__iexact=empresa_nombre.strip()).first()
        if not empresa:
            raise CommandError(f'No existe una empresa con nombre exacto "{empresa_nombre}".')
        return empresa

    def _normalize_row(self, row, empresa, fila):
        errors = []
        warnings = []

        missing_fields = [field for field in REQUIRED_FIELDS if row.get(field) in (None, '')]
        for field in missing_fields:
            errors.append(f'Campo obligatorio ausente: {field}.')

        empresa_archivo = _repair_mojibake(row.get('empresa'))
        if empresa_archivo and _normalize_header(empresa_archivo) != _normalize_header(empresa.nombre):
            errors.append(f'La empresa del archivo ("{empresa_archivo}") no coincide con "{empresa.nombre}".')

        documento = ''.join(ch for ch in str(row.get('documento') or '') if ch.isdigit())
        nombre = _repair_mojibake(row.get('nombre'))
        correo = _repair_mojibake(row.get('correo')).lower()
        telefono = ''.join(ch for ch in str(row.get('telefono') or '') if ch.isdigit())

        monto_inicial = self._parse_money(row.get('monto_inicial'), 'monto_inicial', errors)
        cuota = self._parse_money(row.get('cuota'), 'cuota', errors)
        tasa = self._parse_percent(row.get('tasa'), 'tasa', errors)
        comision_real = self._parse_percent(row.get('comision_real'), 'comision_real', errors)
        fecha_desembolso = self._parse_date(row.get('fecha_desembolso'), 'fecha_desembolso', errors)
        fecha_primer_pago = self._parse_date(row.get('fecha_primer_pago'), 'fecha_primer_pago', errors)
        nro_cuotas = self._parse_int(row.get('nro_cuotas'), 'nro_cuotas', errors)
        cuotas_pagadas = self._parse_int(row.get('cuotas_pagadas'), 'cuotas_pagadas', errors)
        cuotas_pendientes = self._parse_int(row.get('cuotas_pendientes'), 'cuotas_pendientes', errors)

        if not documento:
            errors.append('Documento obligatorio.')
        if not nombre:
            errors.append('Nombre obligatorio.')
        if telefono and len(telefono) > 10:
            warnings.append('Celular tiene mas de 10 digitos; se guardara normalizado.')
        if correo and '@' not in correo:
            errors.append('Correo no tiene formato valido.')
        if nro_cuotas is not None and cuotas_pagadas is not None and cuotas_pendientes is not None:
            if nro_cuotas != cuotas_pagadas + cuotas_pendientes:
                errors.append('La suma de cuotas pagadas y pendientes no coincide con el total.')
        if cuotas_pagadas is not None and cuotas_pagadas < 0:
            errors.append('Cuotas pagadas no puede ser negativo.')
        if fecha_primer_pago and fecha_desembolso and fecha_primer_pago < fecha_desembolso:
            warnings.append('La fecha del primer pago es anterior al desembolso; revisa si el dato historico es correcto.')

        existing_credit = None
        if documento and monto_inicial is not None and nro_cuotas is not None and fecha_primer_pago:
            existing_credit = (
                Credito.objects.filter(
                    linea=Credito.LineaCredito.LIBRANZA,
                    detalle_libranza__empresa=empresa,
                    detalle_libranza__cedula=documento,
                    monto_aprobado=monto_inicial,
                    plazo=nro_cuotas,
                    fecha_primera_cuota_forzada=fecha_primer_pago,
                )
                .order_by('numero_credito')
                .first()
            )
            if existing_credit:
                errors.append(f'Ya existe un credito equivalente: {existing_credit.numero_credito}.')

        return {
            'fila': fila,
            'empresa_archivo': empresa_archivo,
            'pagare': str(row.get('pagare') or '').strip(),
            'documento': documento,
            'nombre': nombre,
            'correo': correo,
            'telefono': telefono,
            'monto_inicial': monto_inicial,
            'fecha_desembolso': fecha_desembolso,
            'fecha_primer_pago': fecha_primer_pago,
            'cuota': cuota,
            'nro_cuotas': nro_cuotas,
            'cuotas_pagadas': cuotas_pagadas,
            'cuotas_pendientes': cuotas_pendientes,
            'tasa': tasa,
            'comision_real': comision_real,
            'errors': errors,
            'warnings': warnings,
            'can_import': not errors,
            'existing_credit_number': existing_credit.numero_credito if existing_credit else '',
        }

    def _import_row(self, row, empresa):
        user = self._resolve_or_create_user(row, empresa)
        self._ensure_vinculo(user, empresa, row)

        monto_inicial = row['monto_inicial']
        comision = (monto_inicial * row['comision_real'] / Decimal('100')).quantize(Decimal('0.01'))
        iva_comision = (comision * Decimal('0.19')).quantize(Decimal('0.01'))
        fecha_desembolso_dt = datetime.combine(row['fecha_desembolso'], time(hour=12, minute=0))
        estado = Credito.EstadoCredito.PAGADO if row['cuotas_pendientes'] == 0 else Credito.EstadoCredito.ACTIVO

        credito = Credito.objects.create(
            usuario=user,
            linea=Credito.LineaCredito.LIBRANZA,
            estado=estado,
            documento_enviado=True,
            monto_solicitado=monto_inicial,
            plazo_solicitado=row['nro_cuotas'],
            monto_aprobado=monto_inicial,
            plazo=row['nro_cuotas'],
            tasa_interes=row['tasa'],
            comision=comision,
            iva_comision=iva_comision,
            total_a_pagar=(row['cuota'] * row['nro_cuotas']).quantize(Decimal('0.01')),
            valor_cuota=row['cuota'],
            fecha_desembolso=fecha_desembolso_dt,
            tipo_regla_credito=Credito.TipoReglaCredito.ESPECIAL,
            fecha_primera_cuota_forzada=row['fecha_primer_pago'],
            plazo_forzado=row['nro_cuotas'],
            tasa_forzada=row['tasa'],
            observacion_regla_especial=LEGACY_IMPORT_OBSERVATION,
        )

        detalle = CreditoLibranza(
            credito=credito,
            nombres=row['nombre'],
            apellidos='',
            cedula=row['documento'],
            direccion='NO REGISTRADA',
            telefono=row['telefono'][:20],
            correo_electronico=row['correo'],
            empresa=empresa,
            ingresos_mensuales=row['monto_inicial'],
        )
        self._attach_placeholder(detalle, 'cedula_frontal', f'{credito.numero_credito}_cedula_frontal.png', MINIMAL_PNG_BYTES)
        self._attach_placeholder(detalle, 'cedula_trasera', f'{credito.numero_credito}_cedula_trasera.png', MINIMAL_PNG_BYTES)
        self._attach_placeholder(
            detalle,
            'certificado_bancario',
            f'{credito.numero_credito}_certificado_bancario.pdf',
            MINIMAL_PDF_BYTES,
        )
        detalle.certificado_bancario_metadata = {
            'estado': 'legacy_backend',
            'mensaje': 'Documento placeholder generado por importador de credito especial legacy.',
        }
        detalle.certificado_bancario_estado_extraccion = 'pendiente'
        detalle.save()

        credit_services.activar_credito(credito)
        self._sync_paid_installments(credito, row['cuotas_pagadas'])

        HistorialEstado.objects.create(
            credito=credito,
            estado_anterior=None,
            estado_nuevo=credito.estado,
            motivo='Importacion backend de credito especial legacy.',
            usuario_modificacion=None,
        )
        return credito

    def _resolve_or_create_user(self, row, empresa):
        User = get_user_model()
        correo = row['correo']
        documento = row['documento']

        user = None
        if correo:
            user = User.objects.filter(email__iexact=correo).first()

        if not user:
            detalle_existente = (
                CreditoLibranza.objects.filter(cedula=documento, empresa=empresa)
                .select_related('credito__usuario')
                .first()
            )
            if detalle_existente:
                user = detalle_existente.credito.usuario

        if not user:
            vinculo_existente = VinculoLaboralEmpresa.objects.filter(
                empresa=empresa,
                documento_empleado=documento,
            ).select_related('usuario').first()
            if vinculo_existente:
                user = vinculo_existente.usuario

        if user:
            updates = []
            if correo and user.email != correo:
                user.email = correo
                updates.append('email')
            normalized_name = normalize_name_upper(row['nombre'])
            if normalized_name and not user.first_name:
                user.first_name = normalized_name[:150]
                updates.append('first_name')
            if updates:
                user.save(update_fields=updates)
            return user

        username_seed = slugify(f"{empresa.slug}-{documento}") or documento
        username = username_seed
        suffix = 2
        while User.objects.filter(username=username).exists():
            username = f'{username_seed}-{suffix}'
            suffix += 1

        user = User.objects.create(
            username=username,
            email=correo,
            first_name=normalize_name_upper(row['nombre'])[:150],
            last_name='',
        )
        user.set_unusable_password()
        user.save(update_fields=['password'])
        return user

    def _ensure_vinculo(self, user, empresa, row):
        vinculo, _ = VinculoLaboralEmpresa.objects.get_or_create(
            usuario=user,
            empresa=empresa,
            documento_empleado=row['documento'],
            defaults={
                'tipo_documento': 'CC',
                'nombre_empleado': normalize_name_upper(row['nombre']),
                'correo_empleado': row['correo'],
                'telefono_empleado': row['telefono'][:20],
                'estado_vinculo': VinculoLaboralEmpresa.EstadoVinculo.ACTIVO,
                'fecha_alta_aprobado': row['fecha_desembolso'],
                'validado_por_pagador': True,
                'observaciones': 'Vinculo creado por importacion backend de creditos especiales legacy.',
            },
        )
        update_fields = []
        normalized_name = normalize_name_upper(row['nombre'])
        if vinculo.nombre_empleado != normalized_name:
            vinculo.nombre_empleado = normalized_name
            update_fields.append('nombre_empleado')
        if row['correo'] and vinculo.correo_empleado != row['correo']:
            vinculo.correo_empleado = row['correo']
            update_fields.append('correo_empleado')
        if row['telefono'] and vinculo.telefono_empleado != row['telefono'][:20]:
            vinculo.telefono_empleado = row['telefono'][:20]
            update_fields.append('telefono_empleado')
        if vinculo.fecha_alta_aprobado != row['fecha_desembolso']:
            vinculo.fecha_alta_aprobado = row['fecha_desembolso']
            update_fields.append('fecha_alta_aprobado')
        if not vinculo.validado_por_pagador:
            vinculo.validado_por_pagador = True
            update_fields.append('validado_por_pagador')
        if vinculo.estado_vinculo != VinculoLaboralEmpresa.EstadoVinculo.ACTIVO:
            vinculo.estado_vinculo = VinculoLaboralEmpresa.EstadoVinculo.ACTIVO
            update_fields.append('estado_vinculo')
        if update_fields:
            vinculo.save(update_fields=update_fields)
        return vinculo

    def _sync_paid_installments(self, credito, cuotas_pagadas):
        cuotas = list(credito.tabla_amortizacion.order_by('numero_cuota'))
        total_cuotas = len(cuotas)
        cuotas_pagadas = min(cuotas_pagadas, total_cuotas)

        for index, cuota in enumerate(cuotas, start=1):
            if index <= cuotas_pagadas:
                cuota.pagada = True
                cuota.monto_pagado = cuota.valor_cuota
                cuota.fecha_pago = datetime.combine(cuota.fecha_vencimiento, time(hour=12, minute=0))
            else:
                cuota.pagada = False
                cuota.monto_pagado = None
                cuota.fecha_pago = None
            cuota.save(update_fields=['pagada', 'monto_pagado', 'fecha_pago'])

        credit_services.recalcular_credito_desde_tabla_amortizacion(credito, persist=True)

    def _attach_placeholder(self, instance, field_name, filename, payload):
        getattr(instance, field_name).save(filename, ContentFile(payload), save=False)

    def _parse_date(self, raw, field_name, errors):
        if isinstance(raw, datetime):
            return raw.date()
        if isinstance(raw, date):
            return raw
        value = str(raw or '').strip()
        if not value:
            errors.append(f'{field_name} es obligatorio.')
            return None
        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
        parts = value.replace('.', '-').replace('/', '-').split('-')
        if len(parts) == 3:
            day, month_raw, year = parts
            month = MONTHS_ES.get(month_raw[:3].lower())
            if month:
                year_num = int(year)
                if year_num < 100:
                    year_num += 2000
                try:
                    return datetime(year_num, month, int(day)).date()
                except ValueError:
                    pass
        errors.append(f'{field_name} no tiene un formato valido.')
        return None

    def _parse_money(self, raw, field_name, errors):
        value = str(raw or '').strip()
        if not value:
            errors.append(f'{field_name} es obligatorio.')
            return None
        normalized = value.replace('$', '').replace(' ', '').replace('.', '').replace(',', '.')
        try:
            return Decimal(normalized)
        except (InvalidOperation, ValueError):
            errors.append(f'{field_name} no es numerico.')
            return None

    def _parse_percent(self, raw, field_name, errors):
        value = str(raw or '').strip().replace('%', '')
        if not value:
            errors.append(f'{field_name} es obligatorio.')
            return None
        normalized = value.replace(' ', '').replace('.', '').replace(',', '.')
        try:
            return Decimal(normalized)
        except (InvalidOperation, ValueError):
            errors.append(f'{field_name} no es numerico.')
            return None

    def _parse_int(self, raw, field_name, errors):
        value = ''.join(ch for ch in str(raw or '') if ch.isdigit())
        if not value:
            errors.append(f'{field_name} es obligatorio.')
            return None
        return int(value)
