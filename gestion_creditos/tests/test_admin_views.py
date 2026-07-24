from datetime import timedelta
from decimal import Decimal
import io

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from gestion_creditos.models import Credito, CreditoAdelantoNomina, Empresa, VinculoLaboralEmpresa


User = get_user_model()


class AdminViewsSmokeTest(TestCase):
    def setUp(self):
        self.staff = User.objects.create_user(
            username='staff-admin',
            email='staff-admin@aprobado.test',
            password='123456',
            is_staff=True,
        )
        self.client.login(username='staff-admin', password='123456')

    def test_paginas_admin_principales_responden(self):
        for url_name in [
            'gestion:dashboard',
            'gestion:solicitudes',
            'gestion:adelantos_nomina',
            'gestion:creditos_activos',
            'gestion:cartera_mora',
        ]:
            response = self.client.get(reverse(url_name))
            self.assertEqual(response.status_code, 200, url_name)

    def test_detalle_credito_adelanto_renderiza_capacidad_descuento(self):
        empresa = Empresa.objects.create(
            nombre='Empresa Admin Smoke',
            tipo_empresa=Empresa.TipoEmpresa.MIXTA,
        )
        usuario = User.objects.create_user(
            username='empleado-smoke',
            email='empleado-smoke@aprobado.test',
            password='123456',
            first_name='Empleado',
            last_name='Smoke',
        )
        vinculo = VinculoLaboralEmpresa.objects.create(
            usuario=usuario,
            empresa=empresa,
            documento_empleado='123456789',
            nombre_empleado='EMPLEADO SMOKE',
            estado_vinculo=VinculoLaboralEmpresa.EstadoVinculo.ACTIVO,
            fecha_alta_aprobado=timezone.localdate() - timedelta(days=90),
            salario_base_mensual=Decimal('2400000.00'),
            auxilio_transporte_mensual=Decimal('162000.00'),
            descuentos_fijos_mensuales=Decimal('350000.00'),
        )
        credito = Credito.objects.create(
            usuario=usuario,
            numero_credito='CR-ADMIN-SMOKE-0001',
            linea=Credito.LineaCredito.ADELANTO_NOMINA,
            estado=Credito.EstadoCredito.SOLICITUD,
            monto_solicitado=Decimal('500000.00'),
            plazo_solicitado=1,
        )
        CreditoAdelantoNomina.objects.create(
            credito=credito,
            vinculo_laboral=vinculo,
            monto_solicitado=Decimal('500000.00'),
            monto_maximo_calculado=Decimal('600000.00'),
            dias_adelanto=5,
            salario_base_usado=Decimal('2400000.00'),
        )

        response = self.client.get(reverse('gestion:credito_detalle', args=[credito.id]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Capacidad de descuento')

    def test_dashboard_export_descarga_excel_funcional(self):
        response = self.client.get(reverse('gestion:dashboard_export'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        workbook = load_workbook(io.BytesIO(response.content))
        self.assertIn('Resumen ejecutivo', workbook.sheetnames)
        self.assertIn('Recaudo contable', workbook.sheetnames)
        self.assertIn('Detalle contable', workbook.sheetnames)
        self.assertIn('Detalle operativo', workbook.sheetnames)

    def test_dashboard_renderiza_indicadores_contables(self):
        response = self.client.get(reverse('gestion:dashboard'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Total recaudado')
        self.assertContains(response, 'Capital recuperado')
        self.assertContains(response, 'Top empresas')
        self.assertContains(response, 'Presencia nacional')
        self.assertContains(response, 'data-company-chart-mode="monto"', html=False)
        self.assertContains(response, 'data-company-chart-mode="creditos"', html=False)
        self.assertContains(response, 'Aún no hay información geográfica suficiente para construir el mapa de presencia.')
