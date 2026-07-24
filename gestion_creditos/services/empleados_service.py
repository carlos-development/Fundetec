import csv
import io
from datetime import datetime
from decimal import Decimal

from django.contrib.auth import get_user_model
from django.utils.text import slugify
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font

from gestion_creditos.models import Empresa, VinculoLaboralEmpresa
from gestion_creditos.services.name_normalization import build_full_name_upper, normalize_name_upper


EMPLOYEE_UPLOAD_HEADERS = [
    'tipo_documento',
    'documento',
    'nombres',
    'apellidos',
    'correo',
    'celular',
    'salario_base',
    'auxilio_transporte',
    'descuentos_fijos',
    'fecha_alta_aprobado',
    'estado_vinculo',
    'convenio_activo',
]


def _to_decimal(value, default='0'):
    try:
        return Decimal(str(value or default).replace(',', '').strip())
    except Exception:
        return Decimal(str(default))


def _parse_date(value):
    raw = str(value or '').strip()
    if not raw:
        raise ValueError('La fecha es obligatoria.')
    for fmt in ('%Y-%m-%d', '%d/%m/%Y'):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    raise ValueError('La fecha debe venir como DD/MM/AAAA o YYYY-MM-DD. Ejemplo: 30/03/2026.')


def _normalize_phone(value):
    return ''.join(ch for ch in str(value or '') if ch.isdigit())


def _username_from_email(email, documento):
    base = slugify((email or '').split('@')[0]) or documento or 'empleado'
    return base[:150]


def _iter_rows(uploaded_file):
    extension = (uploaded_file.name.rsplit('.', 1)[-1] if '.' in uploaded_file.name else '').lower()
    if extension == 'csv':
        content = uploaded_file.read().decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(content))
        for row in reader:
            yield row
        return

    from openpyxl import load_workbook

    workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return
    headers = [str(value or '').strip() for value in rows[0]]
    for values in rows[1:]:
        yield {headers[idx]: values[idx] for idx in range(min(len(headers), len(values)))}


def plantilla_empleados_csv():
    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=EMPLOYEE_UPLOAD_HEADERS)
    writer.writeheader()
    writer.writerow(
        {
            'tipo_documento': 'CC',
            'documento': '1234567890',
            'nombres': 'Daniel',
            'apellidos': 'Angel',
            'correo': 'daniel@empresa.com',
            'celular': '3001234567',
            'salario_base': '2500000',
            'auxilio_transporte': '162000',
            'descuentos_fijos': '300000',
            'fecha_alta_aprobado': '2026-02-01',
            'estado_vinculo': 'ACTIVO',
            'convenio_activo': 'SI',
        }
    )
    return buffer.getvalue()


def plantilla_empleados_xlsx():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Empleados'
    sheet.append(EMPLOYEE_UPLOAD_HEADERS)
    sheet.append(
        [
            'CC',
            '1234567890',
            'Daniel',
            'Angel',
            'daniel@empresa.com',
            '3001234567',
            '2500000',
            '162000',
            '300000',
            '2026-02-01',
            'ACTIVO',
            'SI',
        ]
    )

    for cell in sheet[1]:
        cell.font = Font(bold=True)

    widths = {
        'A': 18,
        'B': 18,
        'C': 22,
        'D': 24,
        'E': 30,
        'F': 18,
        'G': 18,
        'H': 20,
        'I': 20,
        'J': 20,
        'K': 18,
        'L': 16,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    sheet.freeze_panes = 'A2'
    sheet['J1'].comment = Comment(
        'Usa DD/MM/AAAA o YYYY-MM-DD. Ejemplo: 30/03/2026.',
        'Aprobado',
    )

    instrucciones = workbook.create_sheet(title='Instrucciones')
    instrucciones['A1'] = 'Cómo diligenciar la plantilla'
    instrucciones['A1'].font = Font(bold=True, size=13)
    instrucciones['A3'] = '1. No cambies los nombres de las columnas.'
    instrucciones['A4'] = '2. fecha_alta_aprobado debe venir como DD/MM/AAAA o YYYY-MM-DD.'
    instrucciones['A5'] = '3. salario_base, auxilio_transporte y descuentos_fijos deben ir sin símbolos adicionales.'
    instrucciones['A6'] = '4. convenio_activo usa SI o NO.'
    instrucciones['A7'] = '5. estado_vinculo usa ACTIVO, RETIRADO o SUSPENDIDO según corresponda.'
    instrucciones.column_dimensions['A'].width = 100

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


def procesar_carga_empleados(archivo, empresa: Empresa, actor=None):
    User = get_user_model()
    resultados = []

    for index, row in enumerate(_iter_rows(archivo), start=2):
        try:
            documento = ''.join(ch for ch in str(row.get('documento') or '') if ch.isdigit())
            nombres = normalize_name_upper(row.get('nombres'))
            apellidos = normalize_name_upper(row.get('apellidos'))
            correo = str(row.get('correo') or '').strip().lower()
            if not documento or not correo or not nombres:
                raise ValueError('Documento, correo y nombres son obligatorios.')

            user = User.objects.filter(email__iexact=correo).first()
            if not user:
                username = _username_from_email(correo, documento)
                base_username = username
                suffix = 2
                while User.objects.filter(username=username).exists():
                    username = f'{base_username}-{suffix}'
                    suffix += 1
                user = User.objects.create(
                    username=username,
                    email=correo,
                    first_name=nombres[:150],
                    last_name=apellidos[:150],
                    is_active=True,
                )
                user.set_unusable_password()
                user.save(update_fields=['password'])
            else:
                update_fields = []
                if nombres and user.first_name != nombres[:150]:
                    user.first_name = nombres[:150]
                    update_fields.append('first_name')
                if apellidos and user.last_name != apellidos[:150]:
                    user.last_name = apellidos[:150]
                    update_fields.append('last_name')
                if update_fields:
                    user.save(update_fields=update_fields)

            vinculo, created = VinculoLaboralEmpresa.objects.update_or_create(
                usuario=user,
                empresa=empresa,
                defaults={
                    'tipo_documento': str(row.get('tipo_documento') or 'CC').strip().upper(),
                    'documento_empleado': documento,
                    'nombre_empleado': build_full_name_upper(nombres, apellidos),
                    'correo_empleado': correo,
                    'telefono_empleado': _normalize_phone(row.get('celular')),
                    'estado_vinculo': str(row.get('estado_vinculo') or VinculoLaboralEmpresa.EstadoVinculo.ACTIVO).strip().upper(),
                    'fecha_alta_aprobado': _parse_date(row.get('fecha_alta_aprobado')),
                    'salario_base_mensual': _to_decimal(row.get('salario_base')),
                    'auxilio_transporte_mensual': _to_decimal(row.get('auxilio_transporte')),
                    'descuentos_fijos_mensuales': _to_decimal(row.get('descuentos_fijos')),
                    'validado_por_pagador': True,
                    'observaciones': 'Carga estructurada por pagador.',
                    'cargado_por': actor,
                }
            )

            convenio_flag = str(row.get('convenio_activo') or '').strip().lower()
            if convenio_flag in {'si', 'sí', 'true', '1'}:
                empresa_updates = []
                if not empresa.convenio_activo:
                    empresa.convenio_activo = True
                    empresa_updates.append('convenio_activo')
                if empresa.tipo_empresa == Empresa.TipoEmpresa.MARKETPLACE_EXTERNA:
                    empresa.tipo_empresa = Empresa.TipoEmpresa.MIXTA
                    empresa_updates.append('tipo_empresa')
                if empresa_updates:
                    empresa.save(update_fields=empresa_updates)

            resultados.append({
                'fila': index,
                'estado': 'creado' if created else 'actualizado',
                'correo': correo,
                'documento': documento,
            })
        except Exception as exc:
            resultados.append({
                'fila': index,
                'estado': 'error',
                'error': str(exc),
                'correo': str(row.get('correo') or '').strip().lower(),
                'documento': ''.join(ch for ch in str(row.get('documento') or '') if ch.isdigit()),
            })

    return resultados


def reconciliar_usuarios_empleados_legacy(empresa=None):
    resultados = []
    empresas = Empresa.objects.filter(pk=empresa.pk) if empresa else Empresa.objects.all()

    for vinculo in VinculoLaboralEmpresa.objects.select_related('usuario', 'empresa').filter(empresa__in=empresas):
        user = vinculo.usuario
        if not user.email and vinculo.correo_empleado:
            user.email = vinculo.correo_empleado.lower()
            user.save(update_fields=['email'])
        if vinculo.correo_empleado and user.email != vinculo.correo_empleado.lower():
            user.email = vinculo.correo_empleado.lower()
            user.save(update_fields=['email'])
        if vinculo.nombre_empleado:
            parts = vinculo.nombre_empleado.split()
            update_fields = []
            first_name = normalize_name_upper(parts[0][:150]) if parts else ''
            last_name = normalize_name_upper(' '.join(parts[1:])[:150]) if len(parts) > 1 else ''
            if user.first_name != first_name:
                user.first_name = first_name
                update_fields.append('first_name')
            if user.last_name != last_name:
                user.last_name = last_name
                update_fields.append('last_name')
            if update_fields:
                user.save(update_fields=update_fields)
        resultados.append({
            'empresa': vinculo.empresa.nombre,
            'usuario': user.email or user.username,
            'documento': vinculo.documento_empleado,
            'estado': 'revisado',
        })

    return resultados
