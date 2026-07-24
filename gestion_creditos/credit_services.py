from decimal import Decimal, ConversionSyntax, InvalidOperation
import logging
import uuid
import hashlib
from django.conf import settings
from django.shortcuts import get_object_or_404
from django.contrib.auth.models import User
from django.core.files.base import ContentFile

try:
    from openai import OpenAI
except ModuleNotFoundError:
    OpenAI = None
from configuraciones.models import ConfiguracionPeso
from .models import Credito, HistorialEstado, CuentaAhorro, MovimientoAhorro, ConfiguracionTasaInteres, HistorialPago, CuotaAmortizacion, Empresa, LotePagoEmpresa
from .services.accounting import registrar_detalle_contable_abono_capital, registrar_detalle_contable_pago
from .services.tasa_service import obtener_tasa_credito
from .services import dashboard_metrics
from .services.libranza_rules import (
    calcular_primera_fecha_pago_libranza,
    obtener_fecha_primera_cuota_credito,
    obtener_dia_ancla_vencimiento,
    obtener_plazo_credito_aplicado,
    obtener_tasa_credito_aplicada,
    sumar_meses_con_dia_ancla,
)
from .services.credit_lifecycle import saldar_credito_formalmente as _saldar_credito_formalmente
from .services.dashboard_metrics import (
    calcular_total_en_mora as _dashboard_calcular_total_en_mora,
    get_admin_dashboard_context as _dashboard_get_admin_dashboard_context,
)
from django.db.models import Sum, Count, Case, When, F, DecimalField, Q, Avg, Value, ExpressionWrapper, Value, ExpressionWrapper
from django.db.models.functions import TruncMonth, Coalesce
from django.utils import timezone
from datetime import timedelta, datetime
from dateutil.relativedelta import relativedelta
import json
import csv
import io
from django.db import transaction
from django.contrib import messages
from django.urls import NoReverseMatch, reverse

logger = logging.getLogger(__name__)

@transaction.atomic
def gestionar_cambio_estado_credito(credito, nuevo_estado, motivo, usuario_modificacion=None, comprobante=None):
    """
    Centraliza todos los cambios de estado de un crédito, registrando el historial.
    También envía notificaciones por email al cliente.
    """
    from .email_service import enviar_notificacion_cambio_estado

    estado_anterior = credito.estado

    if estado_anterior == nuevo_estado:
        return

    credito.estado = nuevo_estado
    credito.save()

    if nuevo_estado == Credito.EstadoCredito.ACTIVO and estado_anterior != Credito.EstadoCredito.ACTIVO:
        activar_credito(credito)

    HistorialEstado.objects.create(
        credito=credito,
        estado_anterior=estado_anterior,
        estado_nuevo=nuevo_estado,
        motivo=motivo,
        comprobante_pago=comprobante,
        usuario_modificacion=usuario_modificacion
    )
    logger.info(f"Crédito {credito.id} cambió de {estado_anterior} a {nuevo_estado}. Motivo: {motivo}")

    # Enviar notificación por email al cliente
    try:
        enviar_notificacion_cambio_estado(credito, nuevo_estado, motivo)
        logger.info(f"Notificación de email enviada para crédito {credito.id} - Estado: {nuevo_estado}")
    except Exception as e:
        logger.error(f"Error al enviar notificación de email para crédito {credito.id}: {e}")

@transaction.atomic
def preparar_documento_para_firma(credito, usuario_modificacion):
    """
    Prepara el credito para el proceso de firma.
    """
    gestionar_cambio_estado_credito(
        credito=credito,
        nuevo_estado=Credito.EstadoCredito.PENDIENTE_FIRMA,
        motivo="Credito aprobado, pendiente de firma del pagare.",
        usuario_modificacion=usuario_modificacion
    )

    from gestion_creditos.models import Pagare
    from gestion_creditos.services.pagare_service import generar_pagare_pdf
    from gestion_creditos.services.pagare_url import generar_url_publica_temporal
    from gestion_creditos.services.zapsign_client import enviar_pagare_a_zapsign, ZapSignAPIError

    try:
        pagare = getattr(credito, 'pagare', None)
        if pagare and pagare.estado in [Pagare.EstadoPagare.SENT, Pagare.EstadoPagare.SIGNED]:
            credito.documento_enviado = True
            credito.save(update_fields=['documento_enviado'])
            logger.info(
                "Pagare %s ya en estado %s, no se reenvia. credito=%s",
                pagare.numero_pagare,
                pagare.estado,
                credito.id
            )
            return

        pagare = generar_pagare_pdf(credito, usuario_modificacion, forzar_regeneracion=True)
        if pagare.estado == Pagare.EstadoPagare.CREATED:
            pagare_id = pagare.id

            def _enviar_pagare():
                try:
                    pagare_db = Pagare.objects.get(id=pagare_id)
                    if not pagare_db.archivo_pdf or not pagare_db.archivo_pdf.storage.exists(pagare_db.archivo_pdf.name):
                        logger.error(
                            f"PDF del pagare no disponible para credito {pagare_db.credito_id}. "
                            f"Archivo: {pagare_db.archivo_pdf.name if pagare_db.archivo_pdf else 'N/A'}"
                        )
                        logger.error(
                            "Pagare %s no enviado: PDF no disponible. credito=%s",
                            pagare_db.numero_pagare,
                            pagare_db.credito_id
                        )
                        return

                    url_pdf_publica = generar_url_publica_temporal(pagare_db)
                    pagare_enviado = enviar_pagare_a_zapsign(pagare_db, url_pdf_publica)
                    logger.info(
                        "Pagare %s enviado a ZapSign. credito=%s token=%s",
                        pagare_enviado.numero_pagare,
                        pagare_enviado.credito_id,
                        pagare_enviado.zapsign_doc_token
                    )

                    credito_db = pagare_enviado.credito
                    credito_db.documento_enviado = pagare_enviado.estado in [
                        Pagare.EstadoPagare.SENT,
                        Pagare.EstadoPagare.SIGNED
                    ]
                    credito_db.save(update_fields=['documento_enviado'])
                except ZapSignAPIError as e:
                    logger.error(f"Error al enviar el pagare a ZapSign para credito {credito.id}: {e}")
                    logger.error(
                        "Pagare no enviado: ZapSignAPIError. credito=%s",
                        credito.id
                    )
                except Exception as e:
                    logger.error(f"Error inesperado al enviar pagare a ZapSign para credito {credito.id}: {e}")
                    logger.error(
                        "Pagare no enviado: error inesperado. credito=%s",
                        credito.id
                    )

            # Asegurar que el pagaré y su PDF ya estén comprometidos en la BD antes de enviarlo a ZapSign.
            transaction.on_commit(_enviar_pagare)

        if pagare.estado != Pagare.EstadoPagare.CREATED:
            logger.warning(
                "Pagare %s en estado %s, no se envia a ZapSign. credito=%s",
                pagare.numero_pagare,
                pagare.estado,
                credito.id
            )

        credito.documento_enviado = pagare.estado in [Pagare.EstadoPagare.SENT, Pagare.EstadoPagare.SIGNED]
        credito.save(update_fields=['documento_enviado'])

        logger.info(f"El credito {credito.id} ha sido preparado para la firma.")

    except ZapSignAPIError as e:
        logger.error(f"Error al enviar el pagare a ZapSign para credito {credito.id}: {e}")
    except Exception as e:
        logger.error(f"Error inesperado al preparar el pagare para firma en credito {credito.id}: {e}")

def iniciar_proceso_desembolso(credito):
    """
    Inicia el proceso de desembolso.
    """
    logger.info(f"Iniciando proceso de desembolso para el crédito {credito.id}.")

    gestionar_cambio_estado_credito(
        credito=credito,
        nuevo_estado=Credito.EstadoCredito.PENDIENTE_TRANSFERENCIA,
        motivo="El pagaré ha sido firmado. El crédito está pendiente de transferencia.",
        usuario_modificacion=None
    )

    logger.info(f"Crédito {credito.id} pendiente de transferencia por el equipo de finanzas.")



@transaction.atomic
def actualizar_saldo_tras_pago(credito, monto_pagado, pago=None):
    """
    Actualiza el saldo del crédito después de recibir un pago.

    Lógica de actualización:
    1. El pago primero cubre los intereses del período
    2. El remanente del pago abona al capital
    3. Actualiza dos campos:
       - saldo_pendiente: Capital financiado total pendiente (monto + comisión + IVA)
       - capital_pendiente: Solo el monto aprobado pendiente (para mostrar al usuario)
    4. Calcula proporcionalmente cuánto del capital_pendiente se ha pagado

    Ejemplo:
    - Monto aprobado: $500,000
    - Capital financiado: $559,500 (incluye comisión + IVA)
    - Al pagar 1 cuota de $284,750:
      * saldo_pendiente: $559,500 - $284,750 = $274,750
      * capital_pendiente: $500,000 * (274,750/559,500) = $245,448 (proporción)

    Args:
        credito: Instancia del modelo Credito
        monto_pagado: Monto del pago realizado (Decimal o convertible)

    Returns:
        None (actualiza el crédito directamente)
    """
    credito_id = credito.id
    credito = Credito.objects.select_for_update().get(id=credito_id)

    # ✅ Validar y completar datos críticos si faltan
    if credito.monto_aprobado is None:
        logger.warning(
            f"No se puede actualizar saldo para crédito {credito.numero_credito}: falta monto_aprobado"
        )
        return

    updated_fields = []

    if credito.tasa_interes is None:
        credito.tasa_interes = obtener_tasa_credito_aplicada(credito, obtener_tasa_credito(credito.linea))
        updated_fields.append('tasa_interes')

    if credito.comision is None:
        credito.comision = credito.monto_aprobado * Decimal('0.10')
        updated_fields.append('comision')

    if credito.iva_comision is None:
        credito.iva_comision = (credito.comision or Decimal('0.00')) * Decimal('0.19')
        updated_fields.append('iva_comision')

    capital_financiado_inicial = credito.monto_aprobado + (credito.comision or Decimal('0.00')) + (credito.iva_comision or Decimal('0.00'))

    if credito.valor_cuota is None or credito.total_a_pagar is None:
        if credito.plazo:
            tasa_mensual_inicial = (credito.tasa_interes or Decimal('0.00')) / Decimal(100)
            if credito.valor_cuota is None:
                if tasa_mensual_inicial > 0:
                    factor = (tasa_mensual_inicial * (1 + tasa_mensual_inicial) ** credito.plazo) / (
                        ((1 + tasa_mensual_inicial) ** credito.plazo) - 1
                    )
                    credito.valor_cuota = capital_financiado_inicial * factor
                else:
                    credito.valor_cuota = capital_financiado_inicial / credito.plazo
                updated_fields.append('valor_cuota')
            if credito.total_a_pagar is None and credito.valor_cuota is not None:
                credito.total_a_pagar = credito.valor_cuota * credito.plazo
                updated_fields.append('total_a_pagar')

    if credito.saldo_pendiente is None:
        credito.saldo_pendiente = capital_financiado_inicial
        updated_fields.append('saldo_pendiente')

    if credito.capital_pendiente is None:
        credito.capital_pendiente = credito.monto_aprobado
        updated_fields.append('capital_pendiente')

    if credito.fecha_proximo_pago is None:
        hoy = timezone.now().date()
        credito.fecha_proximo_pago = obtener_fecha_primera_cuota_credito(credito, hoy)
        updated_fields.append('fecha_proximo_pago')

    if updated_fields:
        credito.save(update_fields=updated_fields)

    if (
        not credito.tabla_amortizacion.exists()
        and credito.plazo
        and credito.valor_cuota
        and credito.fecha_proximo_pago
    ):
        saldo_capital_restante = capital_financiado_inicial
        fecha_cuota = credito.fecha_proximo_pago
        dia_ancla = obtener_dia_ancla_vencimiento(credito, fecha_cuota)
        tasa_mensual_tabla = (credito.tasa_interes or Decimal('0.00')) / Decimal(100)
        cuotas = []
        for i in range(1, credito.plazo + 1):
            interes_a_pagar = saldo_capital_restante * tasa_mensual_tabla
            capital_a_pagar = credito.valor_cuota - interes_a_pagar
            if i == credito.plazo:
                capital_a_pagar = saldo_capital_restante
                interes_a_pagar = credito.valor_cuota - capital_a_pagar
                if interes_a_pagar < 0:
                    interes_a_pagar = Decimal('0.00')
                    capital_a_pagar = credito.valor_cuota

            saldo_capital_restante -= capital_a_pagar
            if saldo_capital_restante < 0:
                saldo_capital_restante = Decimal('0.00')

            cuotas.append(
                CuotaAmortizacion(
                    credito=credito,
                    numero_cuota=i,
                    fecha_vencimiento=fecha_cuota,
                    capital_a_pagar=capital_a_pagar,
                    interes_a_pagar=interes_a_pagar,
                    valor_cuota=credito.valor_cuota,
                    saldo_capital_pendiente=saldo_capital_restante
                )
            )
            if credito.linea == Credito.LineaCredito.LIBRANZA:
                fecha_cuota = sumar_meses_con_dia_ancla(fecha_cuota, 1, dia_ancla)
            else:
                fecha_cuota += relativedelta(months=1)

        if cuotas:
            CuotaAmortizacion.objects.bulk_create(cuotas, ignore_conflicts=True)
    monto_pagado = Decimal(monto_pagado)
    tasa_mensual = credito.tasa_interes / Decimal(100)

    # Saldo antes del pago (capital financiado total pendiente)
    saldo_antes_pago = credito.saldo_pendiente

    # 1. Calcular el interés generado sobre el saldo pendiente
    interes_del_periodo = saldo_antes_pago * tasa_mensual

    # 2. Determinar abono a interés y capital
    abono_a_interes = min(monto_pagado, interes_del_periodo)
    abono_a_capital = monto_pagado - abono_a_interes

    # 3. ✅ Actualizar saldo_pendiente (capital financiado total)
    credito.saldo_pendiente -= abono_a_capital

    # 4. ✅ Actualizar capital_pendiente PROPORCIONALMENTE
    # Calcular qué porcentaje del capital financiado total se ha pagado
    # y aplicarlo al monto_aprobado original
    if credito.capital_pendiente is not None and credito.total_a_pagar:
        # Calcular capital financiado inicial (si no está guardado, calcularlo)
        capital_financiado_inicial = credito.monto_aprobado + (credito.comision or 0) + (credito.iva_comision or 0)

        if capital_financiado_inicial > 0:
            # Proporción del saldo pendiente respecto al capital financiado inicial
            proporcion_pendiente = credito.saldo_pendiente / capital_financiado_inicial

            # Aplicar esa proporción al monto aprobado original
            credito.capital_pendiente = credito.monto_aprobado * proporcion_pendiente

            # Redondear a 2 decimales para evitar problemas de precisión
            credito.capital_pendiente = credito.capital_pendiente.quantize(Decimal('0.01'))

    # 5. Aplicar el pago a las cuotas pendientes (permite abonos parciales)
    _aplicar_pago_a_cuotas(credito, monto_pagado, pago=pago)

    # 6. Validar si el crédito está completamente pagado
    if credito.saldo_pendiente <= Decimal('0.01'):
        credito.saldo_pendiente = Decimal('0.00')
        if credito.capital_pendiente is not None:
            credito.capital_pendiente = Decimal('0.00')

        # Marcar como pagado si no lo está ya
        if credito.estado != Credito.EstadoCredito.PAGADO:
            gestionar_cambio_estado_credito(
                credito=credito,
                nuevo_estado=Credito.EstadoCredito.PAGADO,
                motivo="Crédito saldado automáticamente por pago."
            )
    else:
        # 7. Avanzar fecha de próximo pago si pagó cuotas completas
        if credito.valor_cuota and credito.valor_cuota > 0 and credito.fecha_proximo_pago:
            cuotas_pagadas = int(monto_pagado // credito.valor_cuota)
            if cuotas_pagadas > 0:
                if credito.linea == Credito.LineaCredito.LIBRANZA:
                    dia_ancla = obtener_dia_ancla_vencimiento(credito)
                    credito.fecha_proximo_pago = sumar_meses_con_dia_ancla(
                        credito.fecha_proximo_pago,
                        cuotas_pagadas,
                        dia_ancla,
                    )
                else:
                    credito.fecha_proximo_pago += relativedelta(months=cuotas_pagadas)

        # 8. Si estaba en mora y se puso al día, volver a ACTIVO
        hoy = timezone.now().date()
        if credito.estado == Credito.EstadoCredito.EN_MORA and credito.fecha_proximo_pago and credito.fecha_proximo_pago > hoy:
            gestionar_cambio_estado_credito(
                credito=credito,
                nuevo_estado=Credito.EstadoCredito.ACTIVO,
                motivo="Crédito actualizado a ACTIVO por pago."
            )

    # Asegurar que no queden saldos negativos
    if credito.saldo_pendiente < 0:
        credito.saldo_pendiente = Decimal('0.00')
    if credito.capital_pendiente and credito.capital_pendiente < 0:
        credito.capital_pendiente = Decimal('0.00')

    # ✅ Guardar cambios en el crédito
    credito.save()

    capital_pendiente_log = credito.capital_pendiente if credito.capital_pendiente is not None else Decimal('0.00')
    logger.info(
        f"Pago procesado para crédito {credito.numero_credito}: "
        f"Monto: ${monto_pagado:,.2f}, Interés: ${abono_a_interes:,.2f}, "
        f"Capital: ${abono_a_capital:,.2f}, Nuevo saldo: ${credito.saldo_pendiente:,.2f}, "
        f"Capital pendiente: ${capital_pendiente_log:,.2f}"
    )

    # Enviar confirmación de pago por email
    try:
        from .email_service import enviar_confirmacion_pago
        enviar_confirmacion_pago(credito, monto_pagado, credito.saldo_pendiente)
        logger.info(f"Confirmación de pago enviada por email para crédito {credito.numero_credito}")
    except Exception as e:
        logger.error(f"Error al enviar confirmación de pago por email para crédito {credito.numero_credito}: {e}")


def _aplicar_pago_a_cuotas(credito, monto_pagado, pago=None):
    """
    Aplica un pago a las cuotas pendientes, permitiendo abonos parciales.

    Regla:
    - El abono se aplica desde la cuota más próxima.
    - Si el abono cubre la cuota completa, se marca como pagada.
    - Si el abono es parcial, se actualiza monto_pagado y se deja pendiente.
    """
    monto_restante = Decimal(monto_pagado)
    cuotas_pendientes = credito.tabla_amortizacion.filter(pagada=False).order_by('numero_cuota')
    aplicaciones_contables = []

    for cuota in cuotas_pendientes:
        ya_pagado = cuota.monto_pagado or Decimal('0.00')
        restante_cuota = cuota.valor_cuota - ya_pagado

        if restante_cuota <= Decimal('0.00'):
            continue

        monto_aplicado = min(monto_restante, restante_cuota)
        if monto_aplicado <= Decimal('0.00'):
            continue

        if monto_restante >= restante_cuota:
            cuota.monto_pagado = cuota.valor_cuota
            cuota.pagada = True
            cuota.fecha_pago = timezone.now()
            cuota.save(update_fields=['monto_pagado', 'pagada', 'fecha_pago'])
            monto_restante -= restante_cuota
        else:
            cuota.monto_pagado = ya_pagado + monto_restante
            cuota.save(update_fields=['monto_pagado'])
            monto_restante = Decimal('0.00')

        aplicaciones_contables.append({
            'credito': credito,
            'cuota': cuota,
            'monto_aplicado': monto_aplicado,
        })

        if monto_restante <= Decimal('0.00'):
            break

    if pago is not None and aplicaciones_contables:
        registrar_detalle_contable_pago(pago=pago, aplicaciones=aplicaciones_contables)


def evaluar_motivacion_credito(texto: str) -> int:
    """
    Evalúa la justificación de un crédito usando la API de OpenAI (GPT-3.5-turbo).

    Asigna un puntaje de 1 a 5 basado en la calidad y coherencia de la
    justificación proporcionada por el solicitante.

    Args:
        texto (str): La justificación del solicitante para el crédito.

    Returns:
        int: Un puntaje entre 1 y 5. Devuelve 3 si el texto es muy corto o
             si ocurre un error en la API.
    """
    if not texto or len(texto) < 10:
        return 3
    if OpenAI is None or not settings.OPENAI_API_KEY:
        return 3
    try:
        client = OpenAI(api_key=settings.OPENAI_API_KEY)
        prompt = f'''Evalúa esta justificación para un crédito y asigna un puntaje del 1 al 5:
        - 1: Muy pobre
        - 2: Pobre
        - 3: Aceptable
        - 4: Bueno
        - 5: Excelente

        Justificación: "{texto}"

        Responde SOLO con el número del puntaje (1-5).'''
        response = client.chat.completions.create(
            model="gpt-3.5-turbo",
            messages=[
                {"role": "system", "content": "Eres un analista financiero experto."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.3,
            max_tokens=2
        )
        respuesta = response.choices[0].message.content.strip()
        puntaje = int(respuesta) if respuesta.isdigit() else 3
        return max(1, min(5, puntaje))
    except Exception as e:
        logger.error(f"Error al evaluar con OpenAI: {e}")
        return 3 # Retorna un puntaje neutral en caso de error


def obtener_puntaje_interno(parametros: dict) -> int:
    """
    Calcula un puntaje interno basado en un conjunto de parámetros y sus
    respectivos pesos definidos en el modelo `ConfiguracionPeso`.

    Args:
        parametros (dict): Un diccionario donde las claves son los nombres de los
                           parámetros y los valores son los niveles seleccionados.

    Returns:
        int: La suma de las estimaciones (puntajes) para los parámetros dados.
    """
    suma_estimaciones = 0
    for parametro, nivel in parametros.items():
        if nivel:
            try:
                configuracion = ConfiguracionPeso.objects.get(parametro=parametro, nivel=nivel)
                suma_estimaciones += configuracion.estimacion
            except ConfiguracionPeso.DoesNotExist:
                logger.warning(f"No se encontró configuración para {parametro} con nivel {nivel}")
    return suma_estimaciones


def filtrar_creditos(request, creditos_base):
    """
    ✅ CORRECTO: Esta función ya está bien porque busca en campos que aún existen en los detalles
    """
    queryset = creditos_base

    # Filtro de búsqueda
    search_text = request.GET.get('search', '').strip()
    if search_text:
        queryset = queryset.filter(
            Q(usuario__username__icontains=search_text) |
            Q(usuario__email__icontains=search_text) |
            Q(numero_credito__icontains=search_text) |
            Q(detalle_emprendimiento__nombre__icontains=search_text) |
            Q(detalle_emprendimiento__numero_cedula__icontains=search_text) |
            Q(detalle_libranza__nombres__icontains=search_text) |
            Q(detalle_libranza__apellidos__icontains=search_text) |
            Q(detalle_libranza__cedula__icontains=search_text) |
            Q(detalle_adelanto_nomina__vinculo_laboral__nombre_empleado__icontains=search_text) |
            Q(detalle_adelanto_nomina__vinculo_laboral__documento_empleado__icontains=search_text)
        )

    empresa_filter = request.GET.get('empresa', '').strip()
    if empresa_filter:
        queryset = queryset.filter(
            Q(
                linea=Credito.LineaCredito.LIBRANZA,
                detalle_libranza__empresa__nombre__iexact=empresa_filter,
            ) |
            Q(
                linea=Credito.LineaCredito.ADELANTO_NOMINA,
                detalle_adelanto_nomina__vinculo_laboral__empresa__nombre__iexact=empresa_filter,
            )
        )

    # Filtro de línea
    linea_filter = request.GET.get('linea', '')
    if linea_filter:
        queryset = queryset.filter(linea=linea_filter)

    # Filtro de estado
    estado_filter = request.GET.get('estado', '')
    if estado_filter:
        queryset = queryset.filter(estado=estado_filter)

    return queryset.distinct()



def calcular_total_en_mora(creditos=None):
    return dashboard_metrics.calcular_total_en_mora(creditos)



def get_admin_dashboard_context(user, request=None):
    return _dashboard_get_admin_dashboard_context(user, request=request)

def activar_credito(credito):
    from gestion_creditos.services.credit_activation import activar_credito as _activar_credito

    return _activar_credito(credito)


def _reverse_url_safe(view_name, *, urlconf=None, fallback=None, kwargs=None):
    try:
        if kwargs:
            return reverse(view_name, kwargs=kwargs, urlconf=urlconf)
        return reverse(view_name, urlconf=urlconf)
    except NoReverseMatch:
        return fallback


def get_billetera_context(user, request=None):
    """
    Prepara el contexto de datos para la vista de la billetera digital.
    """
    cuenta, created = CuentaAhorro.objects.get_or_create(
        usuario=user,
        defaults={
            'tipo_usuario': CuentaAhorro.TipoUsuario.NATURAL,
            'saldo_disponible': Decimal('0.00'),
            'saldo_objetivo': Decimal('1000000.00')
        }
    )
    
    movimientos_recientes = MovimientoAhorro.objects.filter(
        cuenta=cuenta,
        estado__in=['APROBADO', 'PROCESADO']
    ).order_by('-fecha_creacion')[:10]
    
    total_depositado = MovimientoAhorro.objects.filter(
        cuenta=cuenta,
        tipo__in=['DEPOSITO_ONLINE', 'DEPOSITO_OFFLINE'],
        estado__in=['APROBADO', 'PROCESADO']
    ).aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
    
    dias_ahorrando = (timezone.now().date() - cuenta.fecha_apertura.date()).days if cuenta.fecha_apertura else 0
    
    fecha_hace_un_mes = timezone.now() - timedelta(days=30)
    fecha_hace_dos_meses = timezone.now() - timedelta(days=60)
    
    depositos_ultimo_mes = MovimientoAhorro.objects.filter(
        cuenta=cuenta,
        tipo__in=['DEPOSITO_ONLINE', 'DEPOSITO_OFFLINE'],
        estado__in=['APROBADO', 'PROCESADO'],
        fecha_creacion__gte=fecha_hace_un_mes
    ).aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
    
    depositos_mes_anterior = MovimientoAhorro.objects.filter(
        cuenta=cuenta,
        tipo__in=['DEPOSITO_ONLINE', 'DEPOSITO_OFFLINE'],
        estado__in=['APROBADO', 'PROCESADO'],
        fecha_creacion__gte=fecha_hace_dos_meses,
        fecha_creacion__lt=fecha_hace_un_mes
    ).aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
    
    crecimiento_porcentaje = ((depositos_ultimo_mes - depositos_mes_anterior) / depositos_mes_anterior) * 100 if depositos_mes_anterior > 0 else (100 if depositos_ultimo_mes > 0 else 0)
    
    progreso_porcentaje = min((cuenta.saldo_disponible / cuenta.saldo_objetivo) * 100, 100) if cuenta.saldo_objetivo > 0 else 0
    
    tasa_actual = ConfiguracionTasaInteres.objects.filter(activa=True).order_by('-fecha_vigencia').first()
    
    interes_estimado = (cuenta.saldo_disponible * tasa_actual.tasa_anual_efectiva) / 100 if tasa_actual and cuenta.saldo_disponible > 0 else Decimal('0.00')
    
    #? --- Preparación de datos para el gráfico ---
    from dateutil.relativedelta import relativedelta
    chart_labels = []
    chart_values = []

    for i in range(9, -1, -1):
        month_date = timezone.now().replace(day=1) - relativedelta(months=i)
        fecha_inicio = month_date
        fecha_fin = fecha_inicio + relativedelta(months=1)

        total_mes = MovimientoAhorro.objects.filter(
            cuenta=cuenta,
            tipo__in=['DEPOSITO_ONLINE', 'DEPOSITO_OFFLINE', 'AJUSTE_ADMIN'],
            estado__in=['APROBADO', 'PROCESADO'],
            fecha_creacion__gte=fecha_inicio,
            fecha_creacion__lt=fecha_fin
        ).aggregate(total=Sum('monto'))['total'] or 0
        
        chart_labels.append(month_date.strftime('%b'))
        chart_values.append(float(total_mes))

    chart_data = {
        'labels': chart_labels,
        'data': chart_values
    }

    # Determinar tipo de usuario (empleado/libranza vs emprendedor)
    from gestion_creditos.models import Credito
    es_empleado = user.groups.filter(name='Empleados').exists()
    tiene_credito_libranza = Credito.objects.filter(
        usuario=user,
        linea='LIBRANZA'
    ).exists()
    es_libranza = es_empleado or tiene_credito_libranza
    urlconf = getattr(request, 'urlconf', None) if request is not None else None
    access_profile = getattr(user, 'product_access_profile', None)
    has_investor_account = hasattr(user, 'investor_account')
    is_pagador = hasattr(user, 'perfil_pagador')

    home_url = _reverse_url_safe('home', urlconf=urlconf, fallback='/')
    billetera_home_url = home_url
    dashboard_label = 'Mi panel'
    dashboard_url = home_url

    if has_investor_account:
        billetera_home_url = _reverse_url_safe('inversionista:dashboard', urlconf=urlconf, fallback=home_url)
        dashboard_url = billetera_home_url
        dashboard_label = 'Mi portafolio'
    elif is_pagador:
        billetera_home_url = _reverse_url_safe('pagador:dashboard', urlconf=urlconf, fallback=home_url)
        dashboard_url = _reverse_url_safe('pagador:dashboard', urlconf=urlconf, fallback=billetera_home_url)
        dashboard_label = 'Panel pagador'
    elif es_libranza:
        billetera_home_url = _reverse_url_safe('libranza:landing', urlconf=urlconf, fallback=home_url)
        dashboard_url = _reverse_url_safe('libranza:mi_credito', urlconf=urlconf, fallback=billetera_home_url)
        dashboard_label = 'Mi crédito'
    elif access_profile and access_profile.flow == 'EMPRENDIMIENTO':
        billetera_home_url = _reverse_url_safe('home', urlconf=urlconf, fallback=home_url)
        dashboard_url = _reverse_url_safe('emprendimiento:mi_credito', urlconf=urlconf, fallback=billetera_home_url)
        dashboard_label = 'Mi crédito'

    logout_url = (
        _reverse_url_safe('inversionista:logout', urlconf=urlconf)
        or _reverse_url_safe('pagador:logout', urlconf=urlconf)
        or _reverse_url_safe('libranza:logout', urlconf=urlconf)
        or _reverse_url_safe('emprendimiento:logout', urlconf=urlconf)
        or _reverse_url_safe('account_logout', urlconf=urlconf)
        or '/'
    )

    return {
        'cuenta': cuenta,
        'saldo_disponible': cuenta.saldo_disponible,
        'saldo_objetivo': cuenta.saldo_objetivo,
        'progreso_porcentaje': round(progreso_porcentaje, 1),
        'crecimiento_porcentaje': round(crecimiento_porcentaje, 1),
        'dias_ahorrando': dias_ahorrando,
        'emprendimientos_financiados': cuenta.emprendimientos_financiados,
        'familias_beneficiadas': cuenta.familias_beneficiadas,
        'interes_estimado': interes_estimado,
        'tasa_actual': tasa_actual,
        'movimientos_recientes': movimientos_recientes,
        'chart_data': json.dumps(chart_data),
        'total_depositado': total_depositado,
        'es_empleado': es_empleado,
        'es_libranza': es_libranza,
        'billetera_home_url': billetera_home_url,
        'billetera_dashboard_url': dashboard_url,
        'billetera_dashboard_label': dashboard_label,
        'billetera_logout_url': logout_url,
    }

def _leer_csv_pagos(csv_file):
    """
    Lee un CSV de pagos masivos soportando BOM, linea sep= y delimitadores comunes.
    """
    raw = csv_file.read()
    if isinstance(raw, str):
        text = raw
    else:
        text = raw.decode('utf-8-sig')

    lines = text.splitlines()
    if lines and lines[0].strip().lower().startswith('sep='):
        lines = lines[1:]

    cleaned = "\n".join(lines)
    sample = cleaned[:4096]

    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=[',', ';', '\t'])
    except csv.Error:
        header = lines[0] if lines else ''
        delim = ',' if header.count(',') >= header.count(';') else ';'

        class SimpleDialect(csv.Dialect):
            delimiter = delim
            quotechar = '"'
            doublequote = True
            skipinitialspace = True
            lineterminator = '\n'
            quoting = csv.QUOTE_MINIMAL

        dialect = SimpleDialect

    return csv.DictReader(io.StringIO(cleaned), dialect=dialect)





def _leer_archivo_pagos(uploaded_file):
    if hasattr(uploaded_file, 'seek'):
        uploaded_file.seek(0)
    raw = uploaded_file.read()
    if hasattr(uploaded_file, 'seek'):
        uploaded_file.seek(0)
    return raw, getattr(uploaded_file, 'name', 'pagos.csv')


def _iter_rows_pagos_archivo(file_bytes, filename):
    extension = (filename.rsplit('.', 1)[-1] if '.' in filename else '').lower()
    if extension == 'xlsx':
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return
        headers = [str(value or '').strip() for value in rows[0]]
        for row_number, values in enumerate(rows[1:], start=2):
            row = {headers[idx]: values[idx] for idx in range(min(len(headers), len(values)))}
            yield row_number, row
        return

    reader = _leer_csv_pagos(io.BytesIO(file_bytes))
    for row_number, row in enumerate(reader, start=2):
        yield row_number, row


def _normalize_pago_header(header):
    raw = str(header or '').strip().lower()
    if not raw:
        return ''
    normalized = (
        raw.replace('\n', ' ')
        .replace('\r', ' ')
        .replace('(obligatorio)', '')
        .replace('(opcional)', '')
        .replace('(dd/mm/aaaa)', '')
        .replace('[dd/mm/aaaa]', '')
        .replace(':', ' ')
    )
    normalized = ' '.join(normalized.split())
    aliases = {
        'numero credito': 'numero_credito',
        'numero de credito': 'numero_credito',
        'credito': 'numero_credito',
        'cedula empleado': 'cedula',
        'documento': 'cedula',
        'monto': 'monto_a_pagar',
        'monto a pagar': 'monto_a_pagar',
        'fecha pago': 'fecha_pago',
        'fecha aplicacion': 'fecha_pago',
        'referencia': 'referencia_pago',
        'referencia pago': 'referencia_pago',
        'observacion': 'nota',
        'observaciones': 'nota',
        'nota interna': 'nota',
    }
    if normalized in aliases:
        return aliases[normalized]
    return normalized.replace(' ', '_')


def _normalizar_fila_pago(row):
    normalized = {}
    for key, value in row.items():
        if not key:
            continue
        header = _normalize_pago_header(key)
        if not header:
            continue
        normalized[header] = value.strip() if isinstance(value, str) else value
    return normalized


def _parsear_monto_pago(value):
    raw = str(value or '').strip()
    if not raw:
        raise ValueError('El monto es obligatorio.')
    normalized = raw.replace('$', '').replace(' ', '')
    if ',' in normalized and '.' in normalized:
        if normalized.rfind(',') > normalized.rfind('.'):
            normalized = normalized.replace('.', '').replace(',', '.')
        else:
            normalized = normalized.replace(',', '')
    elif ',' in normalized:
        normalized = normalized.replace(',', '.')
    try:
        monto = Decimal(normalized)
    except (ValueError, TypeError, ConversionSyntax, InvalidOperation) as exc:
        raise ValueError(f"Monto '{raw}' no es valido.") from exc
    if monto <= 0:
        raise ValueError('El monto debe ser mayor a cero.')
    return monto


def _parsear_fecha_aplicacion(value):
    if not value:
        return timezone.now()
    if isinstance(value, datetime):
        dt = value
    elif hasattr(value, 'year') and hasattr(value, 'month') and hasattr(value, 'day'):
        dt = datetime.combine(value, datetime.min.time())
    else:
        raw = str(value).strip()
        parsed = None
        for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y', '%d-%b-%y', '%d-%m-%Y'):
            try:
                parsed = datetime.strptime(raw, fmt)
                break
            except ValueError:
                continue
        if parsed is None:
            raise ValueError(
                f"Fecha '{value}' no es válida. La fecha debe venir como DD/MM/AAAA o YYYY-MM-DD. Ejemplo: 30/03/2026."
            )
        dt = parsed
    if timezone.is_naive(dt):
        return timezone.make_aware(dt, timezone.get_current_timezone())
    return timezone.localtime(dt)


def _resolver_credito_pago_empresa(empresa, normalized_row):
    numero_credito = str(normalized_row.get('numero_credito') or '').strip().upper()
    cedula_raw = str(normalized_row.get('cedula') or normalized_row.get('documento') or '').strip()
    cedula = ''.join(ch for ch in cedula_raw if ch.isdigit())

    base_queryset = Credito.objects.filter(
        estado__in=[Credito.EstadoCredito.ACTIVO, Credito.EstadoCredito.EN_MORA]
    ).select_related('detalle_libranza', 'detalle_adelanto_nomina__vinculo_laboral__empresa')

    if numero_credito:
        credito = base_queryset.filter(numero_credito=numero_credito).first()
        if not credito:
            raise ValueError(f"No se encontro el credito {numero_credito}.")
        if credito.empresa_relacionada != empresa:
            raise ValueError(f"El credito {numero_credito} no pertenece a {empresa.nombre}.")
        return credito

    if not cedula:
        raise ValueError('Debes enviar numero_credito o cedula.')

    candidatos = base_queryset.filter(
        Q(linea=Credito.LineaCredito.LIBRANZA, detalle_libranza__empresa=empresa, detalle_libranza__cedula=cedula) |
        Q(linea=Credito.LineaCredito.ADELANTO_NOMINA, detalle_adelanto_nomina__vinculo_laboral__empresa=empresa, detalle_adelanto_nomina__vinculo_laboral__documento_empleado=cedula)
    )
    total = candidatos.count()
    if total == 0:
        raise ValueError(f"No se encontro un credito activo para la cedula {cedula}.")
    if total > 1:
        raise ValueError(f"La cedula {cedula} tiene mas de un credito activo. Usa numero_credito en el archivo.")
    return candidatos.first()


@transaction.atomic
def registrar_pago_credito(
    *,
    credito,
    monto,
    referencia_pago,
    metodo_pago=HistorialPago.MetodoPago.NO_DEFINIDO,
    origen_registro=HistorialPago.OrigenRegistro.LEGACY,
    estado=HistorialPago.EstadoPago.EXITOSO,
    usuario=None,
    empresa=None,
    comprobante=None,
    fecha_aplicacion=None,
    notas='',
    wompi_intento=None,
    lote_pago=None,
):
    pago, created = HistorialPago.objects.get_or_create(
        referencia_pago=referencia_pago,
        defaults={
            'credito': credito,
            'monto': monto,
            'estado': estado,
            'metodo_pago': metodo_pago,
            'origen_registro': origen_registro,
            'empresa_origen': empresa,
            'registrado_por': usuario,
            'comprobante': comprobante,
            'fecha_aplicacion': fecha_aplicacion or timezone.now(),
            'notas': notas or '',
            'wompi_intento': wompi_intento,
            'lote_pago': lote_pago,
        }
    )
    if not created:
        return pago, False

    if estado == HistorialPago.EstadoPago.EXITOSO:
        actualizar_saldo_tras_pago(credito, monto, pago=pago)
        recalcular_credito_desde_tabla_amortizacion(credito, persist=True)
    return pago, True


def validar_archivo_pagos_masivos(uploaded_file, empresa):
    file_bytes, filename = _leer_archivo_pagos(uploaded_file)
    return validar_archivo_pagos_masivos_bytes(file_bytes, filename, empresa)


def validar_archivo_pagos_masivos_bytes(file_bytes, filename, empresa):
    pagos_validos = []
    errores = []
    checksum = None

    try:
        checksum = hashlib.sha256(file_bytes).hexdigest()
        lote_existente = LotePagoEmpresa.objects.filter(empresa=empresa, checksum=checksum, pagos_aplicados__gt=0).first()
        if lote_existente:
            errores.append(
                f"Ya existe una carga de pagos procesada con este archivo para {empresa.nombre} (#{lote_existente.id}). "
                "No se aplicará de nuevo para evitar duplicados."
            )
            return pagos_validos, errores, {'checksum': checksum, 'filename': filename, 'lote_existente': lote_existente}

        for row_number, row in _iter_rows_pagos_archivo(file_bytes, filename):
            normalized = _normalizar_fila_pago(row)
            monto_value = normalized.get('monto_a_pagar') or normalized.get('monto')
            try:
                credito = _resolver_credito_pago_empresa(empresa, normalized)
                monto_a_pagar = _parsear_monto_pago(monto_value)
                fecha_aplicacion = _parsear_fecha_aplicacion(normalized.get('fecha_pago') or normalized.get('fecha_aplicacion'))
            except ValueError as exc:
                errores.append(f"Fila {row_number}: {exc}")
                continue

            referencia = str(normalized.get('referencia_pago') or '').strip().upper()
            if not referencia:
                referencia = f"PAG-{checksum[:6].upper()}-{row_number}"

            if HistorialPago.objects.filter(referencia_pago=referencia).exists():
                errores.append(f"Fila {row_number}: La referencia {referencia} ya existe. Ajusta el archivo para evitar doble aplicación.")
                continue

            pagos_validos.append({
                'credito_id': credito.id,
                'numero_credito': credito.numero_credito,
                'cedula': credito.cliente_documento,
                'nombre': credito.nombre_cliente,
                'monto': monto_a_pagar,
                'referencia_pago': referencia,
                'fecha_aplicacion': fecha_aplicacion,
                'nota': str(normalized.get('nota') or normalized.get('observacion') or '').strip(),
                'fila': row_number,
            })
    except Exception as e:
        logger.error(f"Error al leer archivo de pagos: {e}")
        errores.append(f"Error al procesar el archivo: {e}")

    return pagos_validos, errores, {'checksum': checksum, 'filename': filename}


def crear_borrador_pagos_masivos_archivo(uploaded_file, empresa, usuario=None):
    pagos_validos = []
    errores = []
    lote = None

    try:
        file_bytes, filename = _leer_archivo_pagos(uploaded_file)
        pagos_validos, errores, metadata = validar_archivo_pagos_masivos_bytes(file_bytes, filename, empresa)
        if errores:
            return pagos_validos, errores, lote

        lote = LotePagoEmpresa(
            empresa=empresa,
            nombre_original=metadata['filename'],
            checksum=metadata['checksum'] or '',
            creado_por=usuario,
            total_registros=len(pagos_validos),
            estado=LotePagoEmpresa.EstadoLote.CARGADO,
        )
        lote.archivo.save(metadata['filename'], ContentFile(file_bytes), save=False)
        lote.save()
    except Exception as e:
        logger.error(f"Error al crear borrador del lote de pagos: {e}")
        errores.append(f"Error al preparar la carga de pagos: {e}")

    return pagos_validos, errores, lote


def procesar_lote_pago_empresa(lote, usuario=None, comprobante=None, notas=''):
    pagos_exitosos = 0
    errores = []
    monto_total_aplicado = Decimal('0.00')

    if lote.estado != LotePagoEmpresa.EstadoLote.CARGADO:
        return pagos_exitosos, [f'La carga de pagos #{lote.id} ya fue procesada o no está disponible para confirmación.']

    try:
        with lote.archivo.open('rb') as archivo_lote:
            file_bytes = archivo_lote.read()
        pagos_validos, errores_validacion, metadata = validar_archivo_pagos_masivos_bytes(
            file_bytes,
            lote.nombre_original,
            lote.empresa,
        )
        if errores_validacion:
            return pagos_exitosos, errores_validacion

        with transaction.atomic():
            if comprobante:
                lote.comprobante = comprobante
            if notas:
                lote.notas = notas
            lote.total_registros = len(pagos_validos)
            lote.save(update_fields=['comprobante', 'notas', 'total_registros'])

            for pago_data in pagos_validos:
                credito = Credito.objects.get(id=pago_data['credito_id'])
                registrar_pago_credito(
                    credito=credito,
                    monto=pago_data['monto'],
                    referencia_pago=pago_data['referencia_pago'],
                    metodo_pago=HistorialPago.MetodoPago.TRANSFERENCIA_DIRECTA,
                    origen_registro=HistorialPago.OrigenRegistro.CARGA_MASIVA_EMPRESA,
                    usuario=usuario,
                    empresa=lote.empresa,
                    fecha_aplicacion=pago_data['fecha_aplicacion'],
                    notas=pago_data['nota'] or notas or 'Pago aplicado desde la carga de pagos de la empresa.',
                    lote_pago=lote,
                )
                pagos_exitosos += 1
                monto_total_aplicado += pago_data['monto']

            lote.checksum = metadata['checksum'] or lote.checksum
            lote.pagos_aplicados = pagos_exitosos
            lote.errores_count = 0
            lote.estado = LotePagoEmpresa.EstadoLote.PROCESADO
            lote.save(update_fields=['checksum', 'pagos_aplicados', 'errores_count', 'estado'])
    except Exception as e:
        logger.error(f"Error al procesar la carga de pagos {lote.id}: {e}")
        errores.append(f"Error inesperado al procesar la carga de pagos: {e}")
        lote.errores_count = len(errores)
        lote.estado = LotePagoEmpresa.EstadoLote.PROCESADO_CON_ERRORES
        lote.save(update_fields=['errores_count', 'estado'])

    return pagos_exitosos, errores


def procesar_pagos_masivos_archivo(uploaded_file, empresa, usuario=None, comprobante=None, notas=''):
    pagos_exitosos = 0
    errores = []
    lote = None

    try:
        _, errores_borrador, lote = crear_borrador_pagos_masivos_archivo(uploaded_file, empresa, usuario=usuario)
        if errores_borrador:
            return pagos_exitosos, errores_borrador, lote

        pagos_exitosos, errores = procesar_lote_pago_empresa(
            lote,
            usuario=usuario,
            comprobante=comprobante,
            notas=notas,
        )
    except Exception as e:
        logger.error(f"Error al procesar pagos masivos: {e}")
        errores.append(f"Error inesperado al procesar el archivo: {e}")
        if lote and lote.pk:
            lote.errores_count = len(errores)
            lote.estado = LotePagoEmpresa.EstadoLote.PROCESADO_CON_ERRORES
            lote.save(update_fields=['errores_count', 'estado'])

    return pagos_exitosos, errores, lote


def validar_csv_pagos_masivos(csv_file, empresa):
    pagos_validos, errores, _ = validar_archivo_pagos_masivos(csv_file, empresa)
    return pagos_validos, errores


def procesar_pagos_masivos_csv(csv_file, empresa):
    pagos_exitosos, errores, _ = procesar_pagos_masivos_archivo(csv_file, empresa)
    return pagos_exitosos, errores

def marcar_creditos_en_mora():
    """
    Busca créditos activos cuya fecha de pago ha vencido y los marca como EN_MORA.
    Utiliza el servicio centralizado de cambio de estado.
    Retorna el número de créditos actualizados.
    """
    hoy = timezone.now().date()
    #? Se buscan los créditos que tienen una fecha de próximo pago vencida en cualquiera de sus detalles
    creditos_vencidos = Credito.objects.filter(
        estado=Credito.EstadoCredito.ACTIVO,
        fecha_proximo_pago__lt=hoy
    ).distinct()

    if not getattr(settings, 'LIBRANZA_AUTO_MARK_MORA_ENABLED', True):
        creditos_vencidos = creditos_vencidos.exclude(linea=Credito.LineaCredito.LIBRANZA)

    creditos_actualizados = 0
    for credito in creditos_vencidos:
        try:
            gestionar_cambio_estado_credito(
                credito=credito,
                nuevo_estado=Credito.EstadoCredito.EN_MORA,
                motivo='El crédito ha entrado en mora por vencimiento de la fecha de pago.',
                usuario_modificacion=None  #? Es un proceso automático
            )
            creditos_actualizados += 1
        except Exception as e:
            logger.error(f"Error al marcar en mora el crédito {credito.numero_credito}: {e}")
            
    return creditos_actualizados

@transaction.atomic
def gestionar_consignacion_billetera(movimiento_id: int, es_aprobado: bool, usuario_admin, nota: str):
    """
    Aprueba o rechaza una consignación de billetera y actualiza el saldo si es necesario.
    """
    movimiento = get_object_or_404(
        MovimientoAhorro, 
        id=movimiento_id,
        estado=MovimientoAhorro.EstadoMovimiento.PENDIENTE
    )

    if es_aprobado:
        movimiento.estado = MovimientoAhorro.EstadoMovimiento.APROBADO
        movimiento.nota_admin = nota or 'Consignación aprobada'
        
        #? Actualizar saldo de la cuenta
        cuenta = movimiento.cuenta
        cuenta.saldo_disponible += movimiento.monto
        cuenta.save()
    else:
        movimiento.estado = MovimientoAhorro.EstadoMovimiento.RECHAZADO
        movimiento.nota_admin = nota or 'Sin motivo especificado'

    movimiento.fecha_procesamiento = timezone.now()
    movimiento.procesado_por = usuario_admin
    movimiento.save()
    
    return movimiento

@transaction.atomic
def crear_ajuste_manual_billetera(admin_user, user_email, monto, nota, comprobante):
    """
    Crea un ajuste manual en la billetera de un usuario.
    """
    try:
        usuario = User.objects.get(email=user_email)
    except User.DoesNotExist:
        raise ValueError(f'No existe un usuario con el email {user_email}')

    cuenta, created = CuentaAhorro.objects.get_or_create(
        usuario=usuario,
        defaults={
            'tipo_usuario': CuentaAhorro.TipoUsuario.NATURAL,
            'saldo_disponible': Decimal('0.00')
        }
    )

    movimiento = MovimientoAhorro.objects.create(
        cuenta=cuenta,
        tipo=MovimientoAhorro.TipoMovimiento.AJUSTE_ADMIN,
        monto=monto,
        estado=MovimientoAhorro.EstadoMovimiento.APROBADO,
        comprobante=comprobante,
        descripcion='Abono manual realizado por administrador',
        nota_admin=nota,
        referencia=f"ADMIN-{uuid.uuid4().hex[:12].upper()}",
        fecha_procesamiento=timezone.now(),
        procesado_por=admin_user
    )

    cuenta.saldo_disponible += movimiento.monto
    cuenta.save()

    return movimiento


#? ===================================================================
#? SERVICIOS DE ABONOS AL CRÉDITO Y REESTRUCTURACIÓN
#? ===================================================================

def calcular_cuotas_restantes(credito):
    """
    Calcula el número de cuotas restantes del crédito basándose en la tabla de amortización.

    Returns:
        int: Número de cuotas pendientes de pago
    """
    cuotas_pendientes = credito.tabla_amortizacion.filter(pagada=False).count()
    return cuotas_pendientes


def obtener_resumen_pagos_credito(credito, historial_pagos=None):
    from gestion_creditos.services.credit_recalculation import (
        obtener_resumen_pagos_credito as _obtener_resumen_pagos_credito,
    )

    return _obtener_resumen_pagos_credito(credito, historial_pagos=historial_pagos)


def recalcular_credito_desde_tabla_amortizacion(credito, persist=False):
    from gestion_creditos.services.credit_recalculation import (
        recalcular_credito_desde_tabla_amortizacion as _recalcular_credito_desde_tabla_amortizacion,
    )

    return _recalcular_credito_desde_tabla_amortizacion(credito, persist=persist)


def recalcular_credito_especial_sin_iva_comision(credito, *, persist=False):
    """
    Recalcula un credito especial de libranza eliminando el IVA de la comision.

    Se usa para ajustes puntuales y seguros sobre creditos especiales ya
    existentes. Requiere que el credito no tenga pagos registrados ni cuotas
    historicamente marcadas como pagadas, porque regenera por completo el plan
    de pagos.
    """
    if credito.linea != Credito.LineaCredito.LIBRANZA:
        raise ValueError('Solo aplica a creditos de libranza.')
    if credito.tipo_regla_credito != Credito.TipoReglaCredito.ESPECIAL:
        raise ValueError('Solo aplica a creditos especiales.')
    if not credito.monto_aprobado:
        raise ValueError('El credito no tiene monto_aprobado.')

    plazo_aplicado = obtener_plazo_credito_aplicado(credito)
    if not plazo_aplicado:
        raise ValueError('El credito no tiene plazo aplicable.')

    tasa_default = credito.tasa_interes or obtener_tasa_credito(credito.linea)
    tasa_aplicada = obtener_tasa_credito_aplicada(credito, tasa_default)
    if tasa_aplicada is None:
        raise ValueError('El credito no tiene tasa aplicable.')

    historial_pagos_count = credito.historial_pagos.filter(
        estado=HistorialPago.EstadoPago.EXITOSO
    ).count()
    cuotas_pagadas_count = credito.tabla_amortizacion.filter(pagada=True).count()
    wompi_intentos_count = getattr(credito, 'wompi_intentos', None)
    wompi_intentos_count = wompi_intentos_count.count() if wompi_intentos_count is not None else 0

    if historial_pagos_count:
        raise ValueError('El credito ya tiene pagos registrados.')
    if cuotas_pagadas_count:
        raise ValueError('El credito ya tiene cuotas marcadas como pagadas.')
    if wompi_intentos_count:
        raise ValueError('El credito ya tiene intentos de pago asociados.')

    comision = credito.comision or (credito.monto_aprobado * Decimal('0.10'))
    iva_actual = credito.iva_comision or Decimal('0.00')
    iva_nuevo = Decimal('0.00')
    capital_financiado = credito.monto_aprobado + comision + iva_nuevo

    primera_cuota = credito.tabla_amortizacion.order_by('numero_cuota').values_list(
        'fecha_vencimiento',
        flat=True,
    ).first()
    if primera_cuota:
        fecha_primera_cuota = primera_cuota
    elif credito.fecha_primera_cuota_forzada:
        fecha_primera_cuota = credito.fecha_primera_cuota_forzada
    elif credito.fecha_proximo_pago:
        fecha_primera_cuota = credito.fecha_proximo_pago
    else:
        fecha_primera_cuota = obtener_fecha_primera_cuota_credito(
            credito,
            credito.fecha_desembolso.date() if credito.fecha_desembolso else timezone.localdate(),
        )

    tasa_mensual = tasa_aplicada / Decimal('100')
    if tasa_mensual > 0:
        factor = (tasa_mensual * (Decimal('1.00') + tasa_mensual) ** plazo_aplicado) / (
            ((Decimal('1.00') + tasa_mensual) ** plazo_aplicado) - Decimal('1.00')
        )
        valor_cuota = (capital_financiado * factor).quantize(Decimal('0.01'))
    else:
        valor_cuota = (capital_financiado / plazo_aplicado).quantize(Decimal('0.01'))

    total_a_pagar = (valor_cuota * plazo_aplicado).quantize(Decimal('0.01'))

    fecha_cuota = fecha_primera_cuota
    dia_ancla = obtener_dia_ancla_vencimiento(credito, fecha_cuota)
    saldo_restante = capital_financiado
    cuotas_data = []

    for numero in range(1, plazo_aplicado + 1):
        interes_a_pagar = (saldo_restante * tasa_mensual).quantize(Decimal('0.01'))
        capital_a_pagar = (valor_cuota - interes_a_pagar).quantize(Decimal('0.01'))

        if numero == plazo_aplicado:
            capital_a_pagar = saldo_restante.quantize(Decimal('0.01'))
            interes_a_pagar = (valor_cuota - capital_a_pagar).quantize(Decimal('0.01'))
            if interes_a_pagar < 0:
                interes_a_pagar = Decimal('0.00')
                capital_a_pagar = valor_cuota

        saldo_restante = (saldo_restante - capital_a_pagar).quantize(Decimal('0.01'))
        if saldo_restante < 0:
            saldo_restante = Decimal('0.00')

        cuotas_data.append({
            'numero_cuota': numero,
            'fecha_vencimiento': fecha_cuota,
            'capital_a_pagar': capital_a_pagar,
            'interes_a_pagar': interes_a_pagar,
            'valor_cuota': valor_cuota,
            'saldo_capital_pendiente': saldo_restante,
        })

        fecha_cuota = sumar_meses_con_dia_ancla(fecha_cuota, 1, dia_ancla)

    result = {
        'numero_credito': credito.numero_credito,
        'linea': credito.linea,
        'tipo_regla_credito': credito.tipo_regla_credito,
        'plazo_aplicado': plazo_aplicado,
        'tasa_aplicada': tasa_aplicada,
        'comision_actual': comision,
        'iva_actual': iva_actual,
        'iva_nuevo': iva_nuevo,
        'capital_financiado_nuevo': capital_financiado,
        'valor_cuota_nuevo': valor_cuota,
        'total_a_pagar_nuevo': total_a_pagar,
        'fecha_primera_cuota': fecha_primera_cuota,
        'fecha_proximo_pago_nueva': fecha_primera_cuota,
        'saldo_pendiente_nuevo': total_a_pagar,
        'capital_pendiente_nuevo': credito.monto_aprobado,
        'cuotas_generadas': cuotas_data,
    }

    if persist:
        credito.iva_comision = iva_nuevo
        credito.valor_cuota = valor_cuota
        credito.total_a_pagar = total_a_pagar
        credito.saldo_pendiente = total_a_pagar
        credito.capital_pendiente = credito.monto_aprobado
        credito.fecha_proximo_pago = fecha_primera_cuota

        nota_ajuste = (
            f'Ajuste backend sin IVA sobre comision aplicado el '
            f'{timezone.localtime().strftime("%Y-%m-%d %H:%M")}.'
        )
        observacion_actual = (credito.observacion_regla_especial or '').strip()
        if nota_ajuste not in observacion_actual:
            credito.observacion_regla_especial = (
                f'{observacion_actual}\n{nota_ajuste}'.strip()
                if observacion_actual
                else nota_ajuste
            )

        credito.save(update_fields=[
            'iva_comision',
            'valor_cuota',
            'total_a_pagar',
            'saldo_pendiente',
            'capital_pendiente',
            'fecha_proximo_pago',
            'observacion_regla_especial',
        ])

        credito.tabla_amortizacion.all().delete()
        CuotaAmortizacion.objects.bulk_create([
            CuotaAmortizacion(credito=credito, **cuota_data)
            for cuota_data in cuotas_data
        ])

    return result


def generar_plan_pagos_actual(credito):
    """
    Genera un JSON con el plan de pagos actual del crédito.

    Args:
        credito: Instancia del modelo Credito

    Returns:
        dict: Plan de pagos con cuotas restantes
    """
    cuotas_pendientes = credito.tabla_amortizacion.filter(pagada=False).order_by('numero_cuota')

    plan = {
        'cuotas': [],
        'total_capital': Decimal('0.00'),
        'total_intereses': Decimal('0.00'),
        'total_pagar': Decimal('0.00'),
        'num_cuotas': cuotas_pendientes.count()
    }

    for cuota in cuotas_pendientes:
        plan['cuotas'].append({
            'numero': cuota.numero_cuota,
            'fecha_vencimiento': cuota.fecha_vencimiento.isoformat(),
            'capital': float(cuota.capital_a_pagar),
            'interes': float(cuota.interes_a_pagar),
            'cuota': float(cuota.valor_cuota),
            'saldo_pendiente': float(cuota.saldo_capital_pendiente)
        })
        plan['total_capital'] += cuota.capital_a_pagar
        plan['total_intereses'] += cuota.interes_a_pagar
        plan['total_pagar'] += cuota.valor_cuota

    # Convertir Decimals a float para JSON
    plan['total_capital'] = float(plan['total_capital'])
    plan['total_intereses'] = float(plan['total_intereses'])
    plan['total_pagar'] = float(plan['total_pagar'])

    return plan


def calcular_plan_con_abono(credito, monto_abono, tipo_abono='NORMAL'):
    """
    Calcula el nuevo plan de pagos después de aplicar un abono.

    Args:
        credito: Instancia del modelo Credito
        monto_abono (Decimal): Monto del abono
        tipo_abono (str): 'NORMAL', 'CAPITAL', o 'MAYOR'

    Returns:
        dict: Nuevo plan de pagos después del abono
    """
    from .models import ReestructuracionCredito

    # Obtener cuotas pendientes
    cuotas_pendientes = list(credito.tabla_amortizacion.filter(pagada=False).order_by('numero_cuota'))

    if not cuotas_pendientes:
        return {
            'cuotas': [],
            'total_capital': 0,
            'total_intereses': 0,
            'total_pagar': 0,
            'num_cuotas': 0
        }

    tasa_mensual = credito.tasa_interes / Decimal('100')  # Convertir porcentaje a decimal
    monto_restante = monto_abono

    if tipo_abono == 'CAPITAL':
        # Abono directo a capital - reduce el saldo pero mantiene el mismo plazo
        nuevo_capital_pendiente = max(Decimal('0.00'), credito.capital_pendiente - monto_abono)

        # Recalcular cuotas con el nuevo capital
        if nuevo_capital_pendiente > 0:
            cuotas_restantes = len(cuotas_pendientes)
            nueva_cuota = calcular_cuota_fija(nuevo_capital_pendiente, tasa_mensual, cuotas_restantes)
        else:
            nueva_cuota = Decimal('0.00')
            cuotas_restantes = 0

        # Generar nuevo plan
        plan = {
            'cuotas': [],
            'total_capital': Decimal('0.00'),
            'total_intereses': Decimal('0.00'),
            'total_pagar': Decimal('0.00'),
            'num_cuotas': cuotas_restantes
        }

        saldo = nuevo_capital_pendiente
        fecha_base = cuotas_pendientes[0].fecha_vencimiento

        for i in range(cuotas_restantes):
            interes = saldo * tasa_mensual
            capital = nueva_cuota - interes
            saldo -= capital

            plan['cuotas'].append({
                'numero': cuotas_pendientes[0].numero_cuota + i,
                'fecha_vencimiento': (fecha_base + relativedelta(months=i)).isoformat(),
                'capital': float(capital),
                'interes': float(interes),
                'cuota': float(nueva_cuota),
                'saldo_pendiente': float(max(Decimal('0.00'), saldo))
            })
            plan['total_capital'] += capital
            plan['total_intereses'] += interes
            plan['total_pagar'] += nueva_cuota

    else:  # NORMAL o MAYOR
        # Abono que paga cuotas completas desde la más próxima
        plan = {
            'cuotas': [],
            'total_capital': Decimal('0.00'),
            'total_intereses': Decimal('0.00'),
            'total_pagar': Decimal('0.00'),
            'num_cuotas': 0
        }

        for i, cuota in enumerate(cuotas_pendientes):
            if monto_restante >= cuota.valor_cuota:
                # El abono cubre esta cuota completa - la omitimos del nuevo plan
                monto_restante -= cuota.valor_cuota
            else:
                # El abono no cubre esta cuota - agregamos todas las cuotas restantes
                for cuota_restante in cuotas_pendientes[i:]:
                    plan['cuotas'].append({
                        'numero': cuota_restante.numero_cuota,
                        'fecha_vencimiento': cuota_restante.fecha_vencimiento.isoformat(),
                        'capital': float(cuota_restante.capital_a_pagar),
                        'interes': float(cuota_restante.interes_a_pagar),
                        'cuota': float(cuota_restante.valor_cuota),
                        'saldo_pendiente': float(cuota_restante.saldo_capital_pendiente)
                    })
                    plan['total_capital'] += cuota_restante.capital_a_pagar
                    plan['total_intereses'] += cuota_restante.interes_a_pagar
                    plan['total_pagar'] += cuota_restante.valor_cuota
                break

        plan['num_cuotas'] = len(plan['cuotas'])

    # Convertir Decimals a float para JSON
    plan['total_capital'] = float(plan['total_capital'])
    plan['total_intereses'] = float(plan['total_intereses'])
    plan['total_pagar'] = float(plan['total_pagar'])

    return plan


def calcular_cuota_fija(capital, tasa_mensual, num_cuotas):
    """
    Calcula el valor de la cuota fija usando la fórmula de amortización francesa.

    Args:
        capital (Decimal): Capital a financiar
        tasa_mensual (Decimal): Tasa de interés mensual (en decimal, ej: 0.02 para 2%)
        num_cuotas (int): Número de cuotas

    Returns:
        Decimal: Valor de la cuota mensual
    """
    if num_cuotas == 0 or capital == 0:
        return Decimal('0.00')

    if tasa_mensual == 0:
        return capital / num_cuotas

    # Fórmula: C = P * (i * (1 + i)^n) / ((1 + i)^n - 1)
    factor = (1 + tasa_mensual) ** num_cuotas
    cuota = capital * (tasa_mensual * factor) / (factor - 1)

    return cuota.quantize(Decimal('0.01'))


def calcular_ahorro_intereses(credito, monto_abono, tipo_abono='NORMAL'):
    """
    Calcula el ahorro en intereses al hacer un abono.

    Args:
        credito: Instancia del modelo Credito
        monto_abono (Decimal): Monto del abono
        tipo_abono (str): 'NORMAL', 'CAPITAL', o 'MAYOR'

    Returns:
        Decimal: Ahorro en intereses
    """
    plan_actual = generar_plan_pagos_actual(credito)
    plan_nuevo = calcular_plan_con_abono(credito, monto_abono, tipo_abono)

    ahorro = Decimal(str(plan_actual['total_intereses'])) - Decimal(str(plan_nuevo['total_intereses']))

    return max(Decimal('0.00'), ahorro)


def analizar_abono_credito(credito, monto_abono, tipo_abono='NORMAL'):
    """
    Analiza un abono al crédito y determina si requiere reestructuración.

    Args:
        credito: Instancia del modelo Credito
        monto_abono (Decimal): Monto que el cliente quiere abonar
        tipo_abono (str): 'NORMAL', 'CAPITAL', o 'MAYOR'

    Returns:
        dict: Información sobre el abono y si requiere reestructuración
    """
    from .models import ReestructuracionCredito

    cuota_normal = credito.valor_cuota or Decimal('0.00')

    # Determinar si requiere reestructuración
    requiere_reestructuracion = (
        tipo_abono == 'CAPITAL' or
        monto_abono > (cuota_normal * 2)
    )

    # Obtener planes
    plan_actual = generar_plan_pagos_actual(credito)
    plan_nuevo = calcular_plan_con_abono(credito, monto_abono, tipo_abono)

    # Calcular ahorro
    ahorro = calcular_ahorro_intereses(credito, monto_abono, tipo_abono)

    # Calcular nuevo plazo
    nuevo_plazo = plan_nuevo['num_cuotas']
    plazo_actual = plan_actual['num_cuotas']

    # Calcular nueva cuota mensual (si cambió)
    nueva_cuota = None
    if tipo_abono == 'CAPITAL' and plan_nuevo['num_cuotas'] > 0:
        nueva_cuota = Decimal(str(plan_nuevo['cuotas'][0]['cuota']))

    resultado = {
        'requiere_reestructuracion': requiere_reestructuracion,
        'plan_actual': plan_actual,
        'plan_nuevo': plan_nuevo,
        'ahorro_intereses': float(ahorro),
        'tipo_abono_calculado': tipo_abono,
        'plazo_actual': plazo_actual,
        'nuevo_plazo': nuevo_plazo,
        'cuota_actual': float(cuota_normal),
        'nueva_cuota': float(nueva_cuota) if nueva_cuota else float(cuota_normal),
        'advertencia': None
    }

    if requiere_reestructuracion:
        if tipo_abono == 'CAPITAL':
            resultado['advertencia'] = (
                'Este abono a capital reducirá significativamente sus intereses, '
                'pero su cuota mensual cambiará. El plan de pagos será reestructurado.'
            )
        else:
            resultado['advertencia'] = (
                f'Este abono de ${monto_abono:,.0f} cubre más de 2 cuotas. '
                f'Su plan de pagos será reestructurado, ahorrará ${ahorro:,.0f} en intereses '
                f'y su nuevo plazo será de {nuevo_plazo} cuotas.'
            )

    return resultado


@transaction.atomic
def aplicar_abono_credito(credito, monto_abono, tipo_abono, usuario, referencia_pago):
    """
    Aplica un abono al crédito, crea el registro de reestructuración si es necesario,
    y actualiza la tabla de amortización.

    Args:
        credito: Instancia del modelo Credito
        monto_abono (Decimal): Monto del abono
        tipo_abono (str): 'NORMAL', 'CAPITAL', o 'MAYOR'
        usuario: Usuario que aprueba el abono
        referencia_pago (str): Referencia del pago que generó el abono

    Returns:
        tuple: (HistorialPago, ReestructuracionCredito o None)
    """
    from .models import ReestructuracionCredito

    # Analizar el abono
    analisis = analizar_abono_credito(credito, monto_abono, tipo_abono)

    # Crear el registro del pago
    pago = HistorialPago.objects.create(
        credito=credito,
        monto=monto_abono,
        referencia_pago=referencia_pago,
        estado=HistorialPago.EstadoPago.EXITOSO,
        notas=f"Abono tipo: {tipo_abono}. Ahorro en intereses: ${analisis['ahorro_intereses']:,.0f}"
    )

    # Guardar estado anterior del crédito
    saldo_anterior = credito.saldo_pendiente or Decimal('0.00')
    capital_anterior = credito.capital_pendiente or Decimal('0.00')
    plazo_anterior = calcular_cuotas_restantes(credito)

    # Si requiere reestructuración, crear el registro
    reestructuracion = None
    if analisis['requiere_reestructuracion']:
        reestructuracion = ReestructuracionCredito.objects.create(
            credito=credito,
            monto_abonado=monto_abono,
            tipo_abono=tipo_abono,
            plan_anterior=analisis['plan_actual'],
            plan_nuevo=analisis['plan_nuevo'],
            saldo_pendiente_anterior=saldo_anterior,
            capital_pendiente_anterior=capital_anterior,
            plazo_restante_anterior=plazo_anterior,
            saldo_pendiente_nuevo=Decimal(str(analisis['plan_nuevo']['total_pagar'])),
            capital_pendiente_nuevo=Decimal(str(analisis['plan_nuevo']['total_capital'])),
            plazo_restante_nuevo=analisis['nuevo_plazo'],
            ahorro_intereses=Decimal(str(analisis['ahorro_intereses'])),
            cuota_mensual_nueva=Decimal(str(analisis['nueva_cuota'])) if tipo_abono == 'CAPITAL' else None,
            aprobado_por=usuario,
            pago_relacionado=pago,
            observaciones=analisis['advertencia'] or ''
        )

    # Actualizar tabla de amortización
    if tipo_abono == 'CAPITAL':
        # Abono a capital: recalcular todas las cuotas pendientes
        registrar_detalle_contable_abono_capital(
            pago=pago,
            credito=credito,
            monto_aplicado=monto_abono,
        )
        _recalcular_amortizacion_por_capital(credito, analisis['plan_nuevo'])
    else:
        # Abono normal/mayor: marcar cuotas pagadas
        _marcar_cuotas_pagadas(credito, monto_abono, pago)

    # Actualizar campos del crédito
    credito.saldo_pendiente = Decimal(str(analisis['plan_nuevo']['total_pagar']))
    credito.capital_pendiente = Decimal(str(analisis['plan_nuevo']['total_capital']))

    if tipo_abono == 'CAPITAL' and analisis['nueva_cuota']:
        credito.valor_cuota = Decimal(str(analisis['nueva_cuota']))

    # Si se pagó todo el crédito, cambiar estado
    if credito.saldo_pendiente <= Decimal('0.01'):
        credito.estado = Credito.EstadoCredito.PAGADO
        credito.saldo_pendiente = Decimal('0.00')
        credito.capital_pendiente = Decimal('0.00')

    credito.save()

    logger.info(
        f"Abono aplicado al crédito {credito.numero_credito}. "
        f"Monto: ${monto_abono}, Tipo: {tipo_abono}, "
        f"Ahorro: ${analisis['ahorro_intereses']:,.0f}"
    )

    return pago, reestructuracion


def _recalcular_amortizacion_por_capital(credito, plan_nuevo):
    """
    Recalcula la tabla de amortización cuando se hace un abono a capital.
    Elimina las cuotas pendientes y crea nuevas con los valores recalculados.

    Args:
        credito: Instancia del modelo Credito
        plan_nuevo (dict): Nuevo plan de pagos
    """
    # Eliminar cuotas pendientes
    credito.tabla_amortizacion.filter(pagada=False).delete()

    # Crear nuevas cuotas
    for cuota_data in plan_nuevo['cuotas']:
        CuotaAmortizacion.objects.create(
            credito=credito,
            numero_cuota=cuota_data['numero'],
            fecha_vencimiento=datetime.fromisoformat(cuota_data['fecha_vencimiento']).date(),
            capital_a_pagar=Decimal(str(cuota_data['capital'])),
            interes_a_pagar=Decimal(str(cuota_data['interes'])),
            valor_cuota=Decimal(str(cuota_data['cuota'])),
            saldo_capital_pendiente=Decimal(str(cuota_data['saldo_pendiente'])),
            pagada=False
        )

    logger.info(f"Tabla de amortización recalculada para crédito {credito.numero_credito}")


def _marcar_cuotas_pagadas(credito, monto_abono, pago):
    """
    Marca cuotas como pagadas cuando se hace un abono normal o mayor.

    Args:
        credito: Instancia del modelo Credito
        monto_abono (Decimal): Monto del abono
        pago: Instancia de HistorialPago
    """
    monto_restante = monto_abono
    cuotas_pendientes = credito.tabla_amortizacion.filter(pagada=False).order_by('numero_cuota')
    aplicaciones_contables = []

    for cuota in cuotas_pendientes:
        if monto_restante >= cuota.valor_cuota:
            # Marcar cuota como pagada
            cuota.pagada = True
            cuota.fecha_pago = timezone.now()
            cuota.monto_pagado = cuota.valor_cuota
            cuota.save()

            monto_restante -= cuota.valor_cuota
            aplicaciones_contables.append({
                'credito': credito,
                'cuota': cuota,
                'monto_aplicado': cuota.valor_cuota,
            })

            # Actualizar el desglose del pago
            if pago.capital_abonado is None:
                pago.capital_abonado = Decimal('0.00')
                pago.intereses_pagados = Decimal('0.00')

            pago.capital_abonado += cuota.capital_a_pagar
            pago.intereses_pagados += cuota.interes_a_pagar
        else:
            break

    if aplicaciones_contables:
        registrar_detalle_contable_pago(pago=pago, aplicaciones=aplicaciones_contables)
    else:
        pago.save()
    logger.info(f"Cuotas marcadas como pagadas para crédito {credito.numero_credito}")
