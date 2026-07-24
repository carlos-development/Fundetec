from .common import *
from .common import _build_capacidad_descuento_context
from gestion_creditos.models import DetalleContablePago
from gestion_creditos.services.advisors import filter_creditos_by_asesor
from libranza.services.special_case_audit import create_special_case_audit
from libranza.services.special_case_originator import SpecialCaseOriginationError, originate_special_case_libranza
from libranza.services.special_cases import SpecialCaseSimulationInput, SpecialCaseSimulationError, simulate_special_case_libranza
from risk.services.portfolio_takeover import evaluate_portfolio_takeover
from risk.services.second_credit import SecondCreditService


def _admin_empresas_choices():
    return sorted(
        set(
            credit_services.dashboard_metrics
            ._base_admin_queryset()
            .exclude(empresa_nombre='SIN EMPRESA')
            .values_list('empresa_nombre', flat=True)
        )
    )


def _request_ip_address(request):
    forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if forwarded_for:
        return forwarded_for.split(',')[0].strip()
    return request.META.get('REMOTE_ADDR')


def _resolve_risk_customer_by_document(document_number):
    document_number = (document_number or '').strip()
    if not document_number:
        return None

    credito = (
        Credito.objects
        .select_related('usuario')
        .filter(
            Q(detalle_libranza__cedula=document_number)
            | Q(usuario__username=document_number)
            | Q(usuario__email=document_number)
        )
        .order_by('-fecha_solicitud', '-id')
        .first()
    )
    if credito:
        return credito.usuario

    return (
        User.objects
        .filter(Q(username=document_number) | Q(email=document_number))
        .order_by('id')
        .first()
    )


def _build_risk_diagnostic_payload(*, form, customer):
    scenario = form.cleaned_data['scenario']
    if not customer:
        return {
            'eligible': False,
            'reason': 'cliente_no_encontrado',
            'scenario': scenario,
            'document_number': form.cleaned_data['document_number'],
        }

    if scenario == RiskDiagnosticForm.SCENARIO_SECOND_CREDIT:
        result = SecondCreditService().evaluate(
            customer_id=customer.id,
            product_type=Credito.LineaCredito.LIBRANZA,
            monthly_income=form.cleaned_data.get('monthly_income'),
            projected_installment=form.cleaned_data.get('projected_monthly_payment'),
        )
    else:
        result = evaluate_portfolio_takeover(
            customer_id=customer.id,
            requested_amount=form.cleaned_data['requested_amount'],
            product_type=Credito.LineaCredito.LIBRANZA,
        )

    result = dict(result)
    result.update({
        'scenario': scenario,
        'document_number': form.cleaned_data['document_number'],
        'requested_amount': form.cleaned_data['requested_amount'],
        'projected_monthly_payment': form.cleaned_data['projected_monthly_payment'],
        'monthly_income': form.cleaned_data.get('monthly_income'),
        'customer_id': customer.id,
    })
    return result


def _build_admin_solicitudes_queryset(request, forced_linea=None):
    estado_filter = request.GET.get('estado', '')
    if estado_filter:
        solicitudes_base = Credito.objects.all()
    else:
        solicitudes_base = Credito.objects.exclude(estado__in=['ACTIVO', 'PAGADO', 'EN_MORA'])

    solicitudes_filtradas = credit_services.filtrar_creditos(request, solicitudes_base)
    if forced_linea:
        solicitudes_filtradas = solicitudes_filtradas.filter(linea=forced_linea)
    else:
        solicitudes_filtradas = solicitudes_filtradas.exclude(linea=Credito.LineaCredito.ADELANTO_NOMINA)

    return solicitudes_filtradas.select_related(
        'usuario',
        'detalle_libranza',
        'detalle_emprendimiento',
        'detalle_adelanto_nomina__vinculo_laboral__empresa',
    ).annotate(
        nombre_solicitante=Case(
            When(
                linea='LIBRANZA',
                then=Trim(
                    Concat(
                        Coalesce('detalle_libranza__nombres', Value('')),
                        Value(' '),
                        Coalesce('detalle_libranza__apellidos', Value(''))
                    )
                )
            ),
            When(
                linea='EMPRENDIMIENTO',
                then=Trim(Coalesce('detalle_emprendimiento__nombre', Value('')))
            ),
            When(
                linea='ADELANTO_NOMINA',
                then=Trim(Coalesce('detalle_adelanto_nomina__vinculo_laboral__nombre_empleado', Value('')))
            ),
            default=Trim(
                Concat(
                    Coalesce('usuario__first_name', Value('')),
                    Value(' '),
                    Coalesce('usuario__last_name', Value(''))
                )
            ),
            output_field=CharField()
        ),
        documento_solicitante=Case(
            When(linea='LIBRANZA', then=Coalesce('detalle_libranza__cedula', Value(''))),
            When(linea='EMPRENDIMIENTO', then=Coalesce('detalle_emprendimiento__numero_cedula', Value(''))),
            When(linea='ADELANTO_NOMINA', then=Coalesce('detalle_adelanto_nomina__vinculo_laboral__documento_empleado', Value(''))),
            default=Value(''),
            output_field=CharField()
        )
    ).order_by('-fecha_solicitud')


@staff_member_required
def admin_dashboard_view(request):
    """
    Muestra el dashboard principal administrativo.

    Delega la recolecci?n y procesamiento de todos los datos de contexto
    a la función `get_admin_dashboard_context` en el módulo de servicios para
    mantener la vista limpia y centrada en la renderizaci?n.
    """
    context = credit_services.dashboard_metrics.get_admin_dashboard_context(request.user, request=request)
    return render(request, 'gestion_creditos/admin_dashboard.html', context)


@staff_member_required
def admin_dashboard_export_view(request):
    context = credit_services.dashboard_metrics.get_admin_dashboard_context(request.user, request=request)
    empresa_filter = context.get('empresa_filter') or ''
    selected_asesor = context.get('selected_asesor')
    creditos_operativos = (
        credit_services.dashboard_metrics
        ._base_admin_queryset()
        .filter(estado__in=[Credito.EstadoCredito.ACTIVO, Credito.EstadoCredito.EN_MORA])
    )
    creditos_operativos = filter_creditos_by_asesor(creditos_operativos, selected_asesor)
    if empresa_filter:
        creditos_operativos = creditos_operativos.filter(empresa_nombre=empresa_filter)
    creditos_contables = credit_services.dashboard_metrics.get_platform_disbursed_creditos_queryset(
        credit_services.dashboard_metrics._base_admin_queryset()
    )
    creditos_contables = filter_creditos_by_asesor(creditos_contables, selected_asesor)
    if empresa_filter:
        creditos_contables = creditos_contables.filter(empresa_nombre=empresa_filter)
    detalle_contable_qs = (
        DetalleContablePago.objects.filter(credito__in=creditos_contables)
        .select_related('credito', 'cuota', 'pago')
        .order_by('credito__numero_credito', 'fecha_aplicacion', 'secuencia_aplicacion')
    )

    from openpyxl import Workbook
    from openpyxl.styles import Font

    workbook = Workbook()
    resumen = workbook.active
    resumen.title = 'Resumen ejecutivo'
    resumen.append(['Concepto', 'Valor'])
    resumen.append(['Empresa filtrada', empresa_filter or 'Todas'])
    resumen.append(['Asesor filtrado', selected_asesor.nombre if selected_asesor else 'Todos'])
    resumen.append(['Saldo total de cartera', context['saldo_cartera_total']])
    resumen.append(['Total en mora', context['monto_total_en_mora']])
    resumen.append(['Total de créditos operativos', context['total_creditos']])
    resumen.append(['Próximos a vencer (15 días)', context['proximos_vencer']])
    resumen.append(['Total recaudado', context['total_recaudado']])
    resumen.append(['Capital recuperado', context['capital_recuperado']])
    resumen.append(['Interes recuperado', context['interes_recuperado']])
    resumen.append(['Comision recuperada', context['comision_recuperada']])
    resumen.append(['IVA recuperado', context['iva_recuperado']])
    resumen.append(['Creditos con trazabilidad contable', context['creditos_con_trazabilidad_contable']])
    resumen.append(['Pagos con trazabilidad contable', context['pagos_con_trazabilidad_contable']])
    resumen.append(['Fecha de corte', timezone.now().strftime('%d/%m/%Y %H:%M')])

    cartera_linea = workbook.create_sheet('Cartera por linea')
    cartera_linea.append(['Linea', 'Creditos activos', 'Saldo total'])
    for item in context['creditos_por_linea']:
        cartera_linea.append([item['linea_label'], item['count'], item['saldo_total']])

    estados = workbook.create_sheet('Creditos por estado')
    estados.append(['Estado', 'Cantidad', 'Porcentaje'])
    for item in context['creditos_por_estado']:
        estados.append([item['estado'], item['count'], round(item['porcentaje'], 2)])

    empresas = workbook.create_sheet('Distribucion empresas')
    empresas.append(['Empresa', 'Creditos'])
    for item in context['creditos_por_empresa']:
        empresas.append([item['empresa_nombre'], item['count']])

    recaudo = workbook.create_sheet('Recaudo contable')
    recaudo.append([
        'Numero credito', 'Empresa', 'Linea', 'Total recaudado',
        'Capital recuperado', 'Interes recuperado', 'Comision recuperada', 'IVA recuperado',
    ])
    empresa_por_credito = {credito.id: credito.empresa_nombre for credito in creditos_contables}
    creditos_contables_agregados = (
        detalle_contable_qs.values(
            'credito_id',
            'credito__numero_credito',
            'credito__linea',
        )
        .annotate(
            total_recaudado=Coalesce(Sum('monto_total_aplicado'), Decimal('0.00')),
            capital_recuperado=Coalesce(Sum('capital_principal_aplicado'), Decimal('0.00')),
            interes_recuperado=Coalesce(Sum('interes_aplicado'), Decimal('0.00')),
            comision_recuperada=Coalesce(Sum('comision_aplicada'), Decimal('0.00')),
            iva_recuperado=Coalesce(Sum('iva_aplicado'), Decimal('0.00')),
        )
        .order_by('credito__numero_credito')
    )
    for item in creditos_contables_agregados:
        recaudo.append([
            item['credito__numero_credito'],
            empresa_por_credito.get(item['credito_id'], 'SIN EMPRESA'),
            item['credito__linea'],
            item['total_recaudado'],
            item['capital_recuperado'],
            item['interes_recuperado'],
            item['comision_recuperada'],
            item['iva_recuperado'],
        ])

    detalle_contable = workbook.create_sheet('Detalle contable')
    detalle_contable.append([
        'Numero credito', 'Referencia pago', 'Fecha aplicacion', 'Cuota', 'Secuencia',
        'Monto total', 'Capital recuperado', 'Interes recuperado', 'Comision recuperada', 'IVA recuperado',
    ])
    for detalle_pago in detalle_contable_qs:
        detalle_contable.append([
            detalle_pago.credito.numero_credito,
            detalle_pago.pago.referencia_pago,
            timezone.localtime(detalle_pago.fecha_aplicacion).strftime('%d/%m/%Y %H:%M'),
            detalle_pago.cuota.numero_cuota if detalle_pago.cuota_id else '',
            detalle_pago.secuencia_aplicacion,
            detalle_pago.monto_total_aplicado,
            detalle_pago.capital_principal_aplicado,
            detalle_pago.interes_aplicado,
            detalle_pago.comision_aplicada,
            detalle_pago.iva_aplicado,
        ])

    detalle = workbook.create_sheet('Detalle operativo')
    detalle.append([
        'Numero credito', 'Empresa', 'Linea', 'Estado', 'Cliente',
        'Fecha desembolso', 'Proximo pago', 'Saldo pendiente', 'Valor cuota',
    ])
    for credito in creditos_operativos.order_by('empresa_nombre', 'fecha_proximo_pago', 'numero_credito'):
        detalle.append([
            credito.numero_credito,
            credito.empresa_nombre,
            credito.get_linea_display(),
            credito.get_estado_display(),
            credito.nombre_cliente,
            credito.fecha_desembolso.strftime('%d/%m/%Y') if credito.fecha_desembolso else '',
            credito.fecha_proximo_pago.strftime('%d/%m/%Y') if credito.fecha_proximo_pago else '',
            credito.saldo_pendiente or Decimal('0.00'),
            credito.valor_cuota or Decimal('0.00'),
        ])

    for sheet in workbook.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        for column_cells in sheet.columns:
            length = max(len(str(cell.value or '')) for cell in column_cells[:50])
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 12), 32)

    filename_suffix = empresa_filter.replace(' ', '_') if empresa_filter else 'todas'
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="reporte_dashboard_admin_{filename_suffix}_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    )
    workbook.save(response)
    return response


@staff_member_required
def admin_solicitudes_view(request):
    """Vista para gestionar solicitudes pendientes"""

    solicitudes = _build_admin_solicitudes_queryset(request)
    paginator = Paginator(solicitudes, 20)
    page_number = request.GET.get('page')
    solicitudes_page = paginator.get_page(page_number)

    context = {
        'solicitudes': solicitudes_page,
        'estado_filter': request.GET.get('estado', ''),
        'linea_filter': request.GET.get('linea', ''),
        'empresa_filter': request.GET.get('empresa', ''),
        'search': request.GET.get('search', ''),
        'estados_choices': Credito.EstadoCredito.choices,
        'lineas_choices': [
            choice for choice in Credito.LineaCredito.choices
            if choice[0] != Credito.LineaCredito.ADELANTO_NOMINA
        ],
        'empresas_choices': _admin_empresas_choices(),
        'dashboard_title': 'Solicitudes de credito',
        'dashboard_subtitle': 'Solicitudes pendientes de revision de todas las lineas.',
        'solo_adelantos': False,
    }

    return render(request, 'gestion_creditos/admin_solicitudes.html', context)


@staff_member_required
def admin_adelantos_nomina_view(request):
    solicitudes = _build_admin_solicitudes_queryset(
        request,
        forced_linea=Credito.LineaCredito.ADELANTO_NOMINA,
    )
    paginator = Paginator(solicitudes, 20)
    page_number = request.GET.get('page')
    solicitudes_page = paginator.get_page(page_number)

    context = {
        'solicitudes': solicitudes_page,
        'estado_filter': request.GET.get('estado', ''),
        'linea_filter': Credito.LineaCredito.ADELANTO_NOMINA,
        'empresa_filter': request.GET.get('empresa', ''),
        'search': request.GET.get('search', ''),
        'estados_choices': Credito.EstadoCredito.choices,
        'lineas_choices': Credito.LineaCredito.choices,
        'empresas_choices': _admin_empresas_choices(),
        'dashboard_title': 'Solicitudes de adelanto de nomina',
        'dashboard_subtitle': 'Vista separada para evaluar adelantos sin mezclarlos con creditos tradicionales.',
        'solo_adelantos': True,
    }

    return render(request, 'gestion_creditos/admin_solicitudes.html', context)


@staff_member_required
def admin_creditos_activos_view(request):
    """Vista para gestionar créditos activos"""

    creditos_base = Credito.objects.filter(estado='ACTIVO')
    creditos_filtrados = credit_services.filtrar_creditos(request, creditos_base)

    stats_activos = creditos_filtrados.aggregate(
        total_creditos=Count('id'),
        valor_total=Sum('saldo_pendiente'),
        valor_promedio=Avg('monto_aprobado')
    )

    desembolsos_hoy = creditos_filtrados.filter(
        fecha_desembolso__date=timezone.now().date()
    ).count()

    creditos = creditos_filtrados.select_related(
        'usuario',
        'detalle_libranza__empresa',
        'detalle_emprendimiento',
        'detalle_adelanto_nomina__vinculo_laboral__empresa',
    ).order_by('-fecha_solicitud')

    paginator = Paginator(creditos, 20)
    page_number = request.GET.get('page')
    creditos_page = paginator.get_page(page_number)

    context = {
        'creditos': creditos_page,
        'total_creditos_activos': stats_activos.get('total_creditos') or 0,
        'valor_total_cartera_activa': stats_activos.get('valor_total') or 0,
        'valor_promedio_credito_activo': stats_activos.get('valor_promedio') or 0,
        'desembolsos_hoy': desembolsos_hoy,
        'linea_filter': request.GET.get('linea', ''),
        'empresa_filter': request.GET.get('empresa', ''),
        'search': request.GET.get('search', ''),
        'lineas_choices': Credito.LineaCredito.choices,
        'empresas_choices': _admin_empresas_choices(),
    }

    return render(request, 'gestion_creditos/admin_creditos_activos.html', context)


@staff_member_required
def admin_cartera_view(request):
    """
    Vista para la gestión de cartera, mostrando créditos en mora.
    """

    #? Base de créditos en mora
    creditos_en_mora = Credito.objects.filter(estado=Credito.EstadoCredito.EN_MORA)
    if not getattr(settings, 'LIBRANZA_AUTO_MARK_MORA_ENABLED', True):
        creditos_en_mora = creditos_en_mora.exclude(linea=Credito.LineaCredito.LIBRANZA)

    #? Aplicar filtros de búsqueda, línea de crédito y empresa
    creditos_filtrados = credit_services.filtrar_creditos(request, creditos_en_mora)
    
    #? Anotaciones y ordenamiento
    # Se calcula la diferencia entre hoy y la fecha de vencimiento para poder ordenar por ella.
    today = timezone.now().date()
    creditos_con_dias_mora = creditos_filtrados.annotate(
        dias_en_mora_db=ExpressionWrapper(
            Value(today) - F('fecha_proximo_pago'),
            output_field=DurationField()
        )
    )

    creditos = creditos_con_dias_mora.select_related(
        'usuario',
        'detalle_libranza__empresa',
        'detalle_emprendimiento',
        'detalle_adelanto_nomina__vinculo_laboral__empresa',
    ).order_by('-dias_en_mora_db') # Ordenar por el campo anotado

    #? Paginación
    paginator = Paginator(creditos, 20)
    page_number = request.GET.get('page')
    creditos_page = paginator.get_page(page_number)

    #? Estadísticas de la cartera en mora
    stats_cartera_mora = creditos_filtrados.aggregate(
        total_creditos=Count('id'),
        saldo_pendiente_total=Sum('saldo_pendiente'),
        monto_original_en_mora=Sum('total_a_pagar')
    )
    monto_total_en_mora = credit_services.calcular_total_en_mora(creditos_filtrados)
    stats_cartera_mora['monto_total_en_mora'] = monto_total_en_mora

    monto_original = stats_cartera_mora.get('monto_original_en_mora') or 0
    monto_pendiente = stats_cartera_mora.get('saldo_pendiente_total') or 0
    
    monto_pagado = monto_original - monto_pendiente

    #? calculo de tasa de recuperación (puede mejorar mas adelante)
    tasa_recuperacion = (monto_pagado / monto_original) * 100 if monto_original > 0 else 0

    context = {
        'creditos': creditos_page,
        'stats': stats_cartera_mora,
        'tasa_recuperacion': round(tasa_recuperacion, 2),
        'linea_filter': request.GET.get('linea', ''),
        'empresa_filter': request.GET.get('empresa', ''),
        'search': request.GET.get('search', ''),
        'lineas_choices': Credito.LineaCredito.choices,
        'empresas_choices': _admin_empresas_choices(),
    }
    
    return render(request, 'gestion_creditos/admin_cartera.html', context)


@staff_member_required
@permission_required('gestion_creditos.can_originate_special_libranza', raise_exception=True)
def admin_libranza_special_case_simulator_view(request):
    result = None
    audit = None

    if request.method == 'POST':
        form = SpecialCaseLibranzaSimulationForm(request.POST)
        if form.is_valid():
            try:
                result = simulate_special_case_libranza(
                    SpecialCaseSimulationInput(
                        amount=form.cleaned_data['amount'],
                        term_months=form.cleaned_data['term_months'],
                        monthly_rate=form.cleaned_data['monthly_rate'],
                        commission_rate=form.cleaned_data.get('commission_rate'),
                        commission_amount=form.cleaned_data.get('commission_amount'),
                        vat_rate=form.cleaned_data['vat_rate'],
                    )
                )
                audit = create_special_case_audit(
                    simulation_result=result,
                    created_by=request.user,
                    business_reason=form.cleaned_data['business_reason'],
                    ip_address=_request_ip_address(request),
                    user_agent=request.META.get('HTTP_USER_AGENT', ''),
                )
                messages.success(request, f'Simulacion auditada con ID {audit.id}.')
            except SpecialCaseSimulationError as exc:
                form.add_error(None, str(exc))
    else:
        form = SpecialCaseLibranzaSimulationForm()

    return render(
        request,
        'gestion_creditos/admin_libranza_special_case_simulator.html',
        {
            'form': form,
            'result': result,
            'audit': audit,
        },
    )


@staff_member_required
@permission_required('gestion_creditos.can_originate_special_libranza', raise_exception=True)
def admin_libranza_special_case_originate_view(request, audit_id):
    audit = get_object_or_404(CreditoReglaEspecialAudit, pk=audit_id)
    if audit.credito_id:
        messages.error(request, 'Esta simulacion ya fue originada.')
        return render(
            request,
            'gestion_creditos/admin_libranza_special_case_originate.html',
            {'audit': audit, 'form': None, 'already_originated': True},
            status=400,
        )

    if request.method == 'POST':
        form = SpecialCaseLibranzaOriginationForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                result = originate_special_case_libranza(
                    audit_id=audit.id,
                    applicant_data=form.cleaned_data,
                    files=request.FILES,
                    originated_by=request.user,
                )
                messages.success(
                    request,
                    f'Credito especial {result.credito.numero_credito} creado en revision.',
                )
                return redirect('gestion:credito_detalle', credito_id=result.credito.id)
            except SpecialCaseOriginationError as exc:
                form.add_error(None, str(exc))
    else:
        form = SpecialCaseLibranzaOriginationForm()

    return render(
        request,
        'gestion_creditos/admin_libranza_special_case_originate.html',
        {'audit': audit, 'form': form, 'already_originated': False},
    )


@staff_member_required
@permission_required('gestion_creditos.can_run_risk_diagnostic', raise_exception=True)
def admin_risk_diagnostic_view(request):
    result = None

    if request.method == 'POST':
        form = RiskDiagnosticForm(request.POST)
        if form.is_valid():
            customer = _resolve_risk_customer_by_document(form.cleaned_data['document_number'])
            result = _build_risk_diagnostic_payload(form=form, customer=customer)
    else:
        form = RiskDiagnosticForm()

    return render(
        request,
        'gestion_creditos/admin_risk_diagnostic.html',
        {
            'form': form,
            'result': result,
        },
    )


@staff_member_required
def procesar_solicitud_view(request, credito_id):
    """Aprobar o rechazar una solicitud usando el servicio centralizado."""
    if request.method != 'POST':
        messages.error(request, "Método no permitido.")
        return redirect('gestion:credito_detalle', credito_id=credito_id)

    credito = get_object_or_404(
        Credito,
        id=credito_id,
        estado__in=[Credito.EstadoCredito.SOLICITUD, Credito.EstadoCredito.EN_REVISION, Credito.EstadoCredito.APROBADO]
    )
    action = request.POST.get('action')
    pagador_decision = None
    pagador_aprobado = False

    if credito.linea == Credito.LineaCredito.LIBRANZA:
        pagador_decision = _obtener_decision_pagador(credito)
        pagador_aprobado = (
            pagador_decision is not None
            and pagador_decision.estado_nuevo == Credito.EstadoCredito.APROBADO_PAGADOR
        )
    
    nuevo_estado = None
    if action == 'approve':
        if credito.linea == Credito.LineaCredito.LIBRANZA:
            messages.error(
                request,
                "Las solicitudes de libranza ya no se aprueban desde administracion. "
                "Cuando el pagador aprueba, el sistema las envia directo a firma."
            )
            return redirect('gestion:credito_detalle', credito_id=credito_id)
        nuevo_estado = Credito.EstadoCredito.APROBADO
        if credito.linea == Credito.LineaCredito.LIBRANZA and not pagador_aprobado:
            if pagador_decision and pagador_decision.estado_nuevo == Credito.EstadoCredito.RECHAZADO:
                messages.error(request, "El pagador ya rechaz? la solicitud. Revisa la observación antes de continuar.")
            else:
                messages.error(request, "El pagador aún no ha aprobado la solicitud.")
            return redirect('gestion:credito_detalle', credito_id=credito_id)
        try:
            # El frontend ya envía el número en formato "1400000.50".
            # No se necesita limpieza manual, solo convertir a Decimal.
            monto_aprobado_str = request.POST.get('monto_aprobado', '0')
            plazo_str = request.POST.get('plazo_aprobado', '')

            if not monto_aprobado_str or not plazo_str:
                messages.error(request, "Para aprobar, el monto y el plazo son obligatorios.")
                return redirect('gestion:credito_detalle', credito_id=credito_id)

            credito.monto_aprobado = Decimal(monto_aprobado_str)
            credito.plazo = int(plazo_str)
            credito.save(update_fields=['monto_aprobado', 'plazo'])
        
        except (ValueError, TypeError, decimal.InvalidOperation) as e:
            messages.error(request, f"Monto o plazo inválido: {e}")
            return redirect('gestion:credito_detalle', credito_id=credito_id)

    elif action == 'reject':
        nuevo_estado = Credito.EstadoCredito.RECHAZADO
        messages.warning(request, f'Crédito {credito.numero_credito} rechazado.')
    else:
        messages.error(request, "Acción no válida.")
        return redirect('gestion:credito_detalle', credito_id=credito_id)

    motivo = request.POST.get('observations', 'Decisión inicial de la solicitud.')
    
    try:
        # Primer paso: Cambiar a APROBADO o RECHAZADO
        credit_services.gestionar_cambio_estado_credito(
            credito=credito,
            nuevo_estado=nuevo_estado,
            usuario_modificacion=request.user,
            motivo=motivo
        )

        # Segundo paso (solo si se aprueba): Iniciar la preparaci?n para la firma
        if nuevo_estado == Credito.EstadoCredito.APROBADO:
            credit_services.preparar_documento_para_firma(
                credito=credito,
                usuario_modificacion=request.user
            )
            messages.success(request, f'Crédito {credito.numero_credito} aprobado y pasado a PENDIENTE DE FIRMA.')

    except Exception as e:
        messages.error(request, f"Ocurrió un error inesperado durante el procesamiento: {e}")
        logger.error(f"Error al procesar solicitud del crédito {credito.id}: {e}", exc_info=True)


    return redirect('gestion:credito_detalle', credito_id=credito_id)


@staff_member_required
def detalle_credito_view(request, credito_id):
    """Ver detalles completos de un crédito"""
    
    credito = get_object_or_404(Credito.objects.select_related('detalle_libranza', 'detalle_emprendimiento'), id=credito_id)
    
    historial_pagos = HistorialPago.objects.filter(credito=credito, estado='EXITOSO').order_by('-fecha_aplicacion', '-fecha_pago')
    historial_estados = HistorialEstado.objects.filter(credito=credito).order_by('-fecha')
    resumen_pagos = credit_services.obtener_resumen_pagos_credito(credito, historial_pagos=historial_pagos)
    monto_total_pagado = resumen_pagos['total_pagado']

    #! Unificar el acceso a los detalles del crédito (NO SE EST? USANDO)
    detalle_credito = credito.detalle

    #! Los c?lculos ahora se manejan en el modelo o en servicios,
    #! la vista solo se encarga de mostrar la información.

    #? Nuevos c?lculos para la vista de detalle
    cuotas_pagadas = resumen_pagos['cuotas_pagadas']
    cuotas_restantes = resumen_pagos['cuotas_restantes']

    # Tabla de amortización (si existe)
    tabla_amortizacion = credito.tabla_amortizacion.all().order_by('numero_cuota')

    pagador_decision = None
    pagador_aprobado = False
    pagador_rechazado = False

    if credito.linea == Credito.LineaCredito.LIBRANZA:
        pagador_decision = HistorialEstado.objects.filter(
            credito=credito,
            usuario_modificacion__perfil_pagador__isnull=False,
            estado_nuevo__in=[Credito.EstadoCredito.APROBADO_PAGADOR, Credito.EstadoCredito.RECHAZADO]
        ).order_by('-fecha').first()
        if pagador_decision:
            pagador_aprobado = pagador_decision.estado_nuevo == Credito.EstadoCredito.APROBADO_PAGADOR
            pagador_rechazado = pagador_decision.estado_nuevo == Credito.EstadoCredito.RECHAZADO

    # Determinar si el crédito puede ser procesado (aprobado/rechazado)
    puede_procesar = credito.estado in [Credito.EstadoCredito.SOLICITUD, Credito.EstadoCredito.EN_REVISION]
    if credito.linea == Credito.LineaCredito.LIBRANZA:
        puede_procesar = False

    context = {
        'credito': credito,

        # âœ… Campos ahora vienen del modelo Credito
        'monto_solicitado': credito.monto_solicitado,
        'plazo_solicitado': credito.plazo_solicitado,
        'monto_aprobado': credito.monto_aprobado,
        'plazo': credito.plazo,
        'tasa_interes': credito.tasa_interes,
        'saldo_pendiente': resumen_pagos['saldo_pendiente'],
        'valor_cuota': credito.valor_cuota,
        'fecha_proximo_pago': resumen_pagos['fecha_proximo_pago'],
        'total_a_pagar': credito.total_a_pagar,
        'comision': credito.comision,
        'iva_comision': credito.iva_comision,

        # Campos que Sí vienen del detalle
        'detalle': credito.detalle,  # Usa la property
        'historial_pagos': historial_pagos,
        'historial_estados': historial_estados,
        'puede_procesar': puede_procesar,
        'cuotas_pagadas': cuotas_pagadas,
        'cuotas_restantes': cuotas_restantes,
        'monto_total_pagado': monto_total_pagado,
        'tabla_amortizacion': tabla_amortizacion,  #  NUEVA: Tabla de amortización
        'pagador_decision': pagador_decision,
        'pagador_aprobado': pagador_aprobado,
        'pagador_rechazado': pagador_rechazado,
        'capacidad_descuento': _build_capacidad_descuento_context(credito),
        'libranza_tasa_mensual': obtener_tasa_credito(Credito.LineaCredito.LIBRANZA),
    }
    
    return render(request, 'gestion_creditos/admin_detalle_credito.html', context)


@staff_member_required
@require_POST
def confirmar_desembolso_view(request, credito_id):
    """
    Vista dedicada para que finanzas confirme el desembolso y active el crédito.
    """
    credito = get_object_or_404(Credito, id=credito_id)
    comprobante = request.FILES.get('comprobante_pago')

    # 1. Validar estado actual
    if credito.estado != Credito.EstadoCredito.PENDIENTE_TRANSFERENCIA:
        messages.error(request, f"El crédito no está en estado 'Pendiente de Transferencia'. Estado actual: {credito.get_estado_display()}.")
        return redirect('gestion:credito_detalle', credito_id=credito.id)

    # 2. Validar que se haya subido el comprobante
    if not comprobante:
        messages.error(request, "Es obligatorio adjuntar el comprobante de desembolso.")
        return redirect('gestion:credito_detalle', credito_id=credito.id)

    # 3. Ejecutar el cambio de estado a ACTIVO
    try:
        credit_services.gestionar_cambio_estado_credito(
            credito=credito,
            nuevo_estado=Credito.EstadoCredito.ACTIVO,
            motivo="Desembolso confirmado y comprobante adjuntado por el equipo de finanzas.",
            comprobante=comprobante,
            usuario_modificacion=request.user
        )
        messages.success(request, f"Crédito {credito.numero_credito} activado exitosamente.")
    except Exception as e:
        messages.error(request, f"Ocurrió un error inesperado al activar el crédito: {e}")
        logger.error(f"Error al activar crédito {credito.id} vía confirmación de desembolso: {e}", exc_info=True)

    return redirect('gestion:credito_detalle', credito_id=credito.id)


@staff_member_required
@require_POST
def agregar_pago_manual_view(request, credito_id):
    credito = get_object_or_404(Credito, id=credito_id)
    monto = request.POST.get('monto')
    auth_key = request.POST.get('auth_key')

    if not monto or not auth_key:
        messages.error(request, "Monto y clave de autorización son requeridos.")
        return redirect('gestion:credito_detalle', credito_id=credito.id)

    if auth_key != getattr(settings, 'MANUAL_PAYMENT_AUTH_KEY', None):
        messages.error(request, "Clave de autorización no válida.")
        return redirect('gestion:credito_detalle', credito_id=credito.id)

    try:
        monto_decimal = Decimal(monto)
        if monto_decimal <= 0:
            raise ValueError("El monto debe ser positivo.")

        with transaction.atomic():
            pago = HistorialPago.objects.create(
                credito=credito,
                monto=monto_decimal,
                referencia_pago=f"MANUAL-{credito.id}-{timezone.now().strftime('%Y%m%d%H%M%S%f')}",
                estado=HistorialPago.EstadoPago.EXITOSO
            )

            detalle = credito.detalle

            if detalle:
                #? Actualizar saldo y estado usando el helper
                credit_services.actualizar_saldo_tras_pago(credito, monto_decimal, pago=pago)
            
            messages.success(request, f"Abono de ${monto_decimal:,.2f} registrado exitosamente.")

    except (ValueError, TypeError) as e:
        messages.error(request, f"Error en el monto: {e}")
    except Exception as e:
        messages.error(request, f"Ocurrió un error inesperado: {e}")

    return redirect('gestion:credito_detalle', credito_id=credito.id)


@staff_member_required
@require_POST
def saldar_credito_formalmente_view(request, credito_id):
    credito = get_object_or_404(Credito, id=credito_id)
    auth_key = request.POST.get('auth_key')
    motivo = (request.POST.get('motivo_saldo') or '').strip()

    if credito.estado == Credito.EstadoCredito.PAGADO:
        messages.info(request, f"El crÃ©dito {credito.numero_credito} ya estÃ¡ saldado.")
        return redirect('gestion:credito_detalle', credito_id=credito.id)

    if credito.estado not in [Credito.EstadoCredito.ACTIVO, Credito.EstadoCredito.EN_MORA]:
        messages.error(request, 'Solo se pueden saldar crÃ©ditos activos o en mora.')
        return redirect('gestion:credito_detalle', credito_id=credito.id)

    if not auth_key:
        messages.error(request, 'La clave de autorizaciÃ³n es requerida.')
        return redirect('gestion:credito_detalle', credito_id=credito.id)

    if auth_key != getattr(settings, 'MANUAL_PAYMENT_AUTH_KEY', None):
        messages.error(request, 'Clave de autorizaciÃ³n no vÃ¡lida.')
        return redirect('gestion:credito_detalle', credito_id=credito.id)

    try:
        from gestion_creditos.services.credit_lifecycle import saldar_credito_formalmente

        saldar_credito_formalmente(
            credito,
            actor=request.user,
            motivo=motivo or 'CrÃ©dito saldado por cierre administrativo controlado.',
            fecha_operacion=timezone.now(),
        )
        messages.success(request, f"CrÃ©dito {credito.numero_credito} saldado correctamente.")
    except Exception as exc:
        logger.exception("Error al saldar formalmente el crÃ©dito %s", credito.numero_credito)
        messages.error(request, f"No fue posible saldar el crÃ©dito: {exc}")

    return redirect('gestion:credito_detalle', credito_id=credito.id)


@staff_member_required
def descargar_documentos_view(request, credito_id):
    credito = get_object_or_404(Credito, id=credito_id) #! Obtener el crédito
    buffer = io.BytesIO() #! Crear un buffer en memoria para el ZIP

    with zipfile.ZipFile(buffer, 'w') as zip_file:
        detalle = credito.detalle
        
        document_map = {
            Credito.LineaCredito.LIBRANZA: [
                'cedula_frontal', 'cedula_trasera', 'certificado_laboral', 
                'desprendible_nomina', 'certificado_bancario'
            ],
            Credito.LineaCredito.EMPRENDIMIENTO: ['foto_negocio']
        }
        
        document_fields = document_map.get(credito.linea, [])

        if detalle:
            for field_name in document_fields:
                file_field = getattr(detalle, field_name, None)
                if file_field and hasattr(file_field, 'path'):
                    try:
                        zip_file.write(file_field.path, file_field.name)
                    except FileNotFoundError:
                        logger.warning(f"Archivo no encontrado para el crédito {credito.id}: {file_field.path}")

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename="documentos_credito_{credito.id}.zip"'
    return response


@staff_member_required
def documentacion_credito_view(request, credito_id):
    credito = get_object_or_404(
        Credito.objects.select_related('detalle_libranza', 'detalle_emprendimiento', 'usuario'),
        id=credito_id
    )

    documentos = []

    def infer_kind(url_value, filename=None):
        candidates = []
        if filename:
            candidates.append(filename)
        if url_value:
            candidates.append(url_value)
        for value in candidates:
            clean = value.split('?', 1)[0].lower()
            if clean.endswith('.pdf'):
                return 'pdf'
            if clean.endswith(('.png', '.jpg', '.jpeg', '.webp', '.gif')):
                return 'image'
        if url_value and url_value.startswith('http'):
            return 'link'
        return 'file'

    def build_url(file_field):
        if not file_field or not getattr(file_field, 'name', None):
            return ''
        try:
            return request.build_absolute_uri(file_field.url)
        except Exception:
            return file_field.name

    def build_preview_url(file_field):
        if not file_field or not getattr(file_field, 'name', None):
            return ''
        preview_path = reverse('gestion:documento_preview')
        return request.build_absolute_uri(f"{preview_path}?path={quote(file_field.name)}")

    def add_doc(title, file_field=None, url=None, source='', status='', created_at=None, signed_at=None, description=''):
        doc_url = ''
        filename = None
        if file_field:
            filename = getattr(file_field, 'name', None)
            doc_url = build_preview_url(file_field) or build_url(file_field)
        if not doc_url and url:
            doc_url = url
        if not doc_url:
            return
        if not created_at and not signed_at:
            created_at = credito.fecha_solicitud
        documentos.append({
            'title': title,
            'url': doc_url,
            'kind': infer_kind(doc_url, filename=filename),
            'source': source,
            'status': status,
            'created_at': created_at,
            'signed_at': signed_at,
            'description': description,
        })

    # Documentos de solicitud
    if credito.linea == Credito.LineaCredito.LIBRANZA and credito.detalle_libranza:
        detalle = credito.detalle_libranza
        add_doc('Cédula (frontal)', file_field=detalle.cedula_frontal, source='Solicitud')
        add_doc('Cédula (trasera)', file_field=detalle.cedula_trasera, source='Solicitud')
        add_doc('Certificado laboral', file_field=detalle.certificado_laboral, source='Solicitud')
        add_doc('Desprendible de nómina', file_field=detalle.desprendible_nomina, source='Solicitud')
        add_doc('Certificado bancario', file_field=detalle.certificado_bancario, source='Solicitud')
    elif credito.linea == Credito.LineaCredito.EMPRENDIMIENTO and credito.detalle_emprendimiento:
        detalle = credito.detalle_emprendimiento
        add_doc('Fotos del negocio (PDF)', file_field=detalle.foto_negocio, source='Solicitud')
        for imagen in detalle.imagenes_negocio.all():
            add_doc(
                f"Imagen negocio ({imagen.get_tipo_imagen_display()})",
                file_field=imagen.imagen,
                source='Im?genes',
                created_at=imagen.fecha_subida,
                description=imagen.descripcion or ''
            )

    # Pagaré y firma
    pagare = None
    try:
        pagare = credito.pagare
    except Pagare.DoesNotExist:
        pagare = None

    if pagare:
        add_doc(
            'Pagaré generado',
            file_field=pagare.archivo_pdf,
            source='ZapSign',
            status=pagare.get_estado_display(),
            created_at=pagare.fecha_creacion
        )
        signed_url = build_url(pagare.archivo_pdf_firmado) or (pagare.zapsign_signed_file_url or '')
        add_doc(
            'Pagaré firmado',
            url=signed_url,
            source='ZapSign',
            status=pagare.get_estado_display(),
            created_at=pagare.fecha_firma,
            signed_at=pagare.fecha_firma
        )
        if pagare.zapsign_sign_url:
            add_doc(
                'Enlace de firma (ZapSign)',
                url=pagare.zapsign_sign_url,
                source='ZapSign',
                status=pagare.get_estado_display(),
                created_at=pagare.fecha_envio
            )

    # Comprobantes de desembolso u otros archivos en historial de estado
    estados_con_comprobante = HistorialEstado.objects.filter(
        credito=credito,
        comprobante_pago__isnull=False
    ).order_by('-fecha')
    for estado in estados_con_comprobante:
        add_doc(
            'Comprobante de desembolso',
            file_field=estado.comprobante_pago,
            source='Desembolso',
            status=estado.get_estado_nuevo_display(),
            created_at=estado.fecha,
            description=estado.motivo or ''
        )

    def sort_key(doc):
        return doc.get('created_at') or doc.get('signed_at') or credito.fecha_solicitud

    documentos = sorted(documentos, key=sort_key, reverse=True)

    context = {
        'credito': credito,
        'documentos': documentos,
        'total_documentos': len(documentos),
    }
    return render(request, 'gestion_creditos/admin_documentos_credito.html', context)


@staff_member_required
@xframe_options_exempt
def documento_preview_view(request):
    path = (request.GET.get('path') or '').strip()
    if not path:
        raise Http404("Documento no encontrado.")

    try:
        full_path = safe_join(settings.MEDIA_ROOT, path)
    except SuspiciousFileOperation:
        raise Http404("Documento no encontrado.")

    if not os.path.exists(full_path):
        raise Http404("Documento no encontrado.")

    content_type, _ = mimetypes.guess_type(full_path)
    response = FileResponse(open(full_path, 'rb'), content_type=content_type or 'application/octet-stream')
    response['Content-Disposition'] = f'inline; filename="{os.path.basename(full_path)}"'
    return response
