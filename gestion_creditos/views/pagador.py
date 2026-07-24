from .common import *
from .common import _build_capacidad_descuento_context, _obtener_decision_pagador
from ..models import CuotaAmortizacion
from ..services.pagador_collaborators_service import build_pagador_collaborators_context


@login_required(login_url='/pagador/login/')
@pagador_required
@require_http_methods(["GET"])
def descargar_plantilla_empleados_view(request):
    response = HttpResponse(
        plantilla_empleados_xlsx(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="plantilla_empleados_aprobado.xlsx"'
    return response


@login_required(login_url='/pagador/login/')
@pagador_required
@require_http_methods(["GET", "POST"])
def pagador_carga_empleados_view(request):
    resultados = None
    if request.method == 'POST':
        form = EmployeeBulkUploadForm(request.POST, request.FILES)
        if form.is_valid():
            resultados = procesar_carga_empleados(
                form.cleaned_data['archivo'],
                empresa=request.empresa,
                actor=request.user,
            )
            creados = sum(1 for row in resultados if row['estado'] in {'creado', 'actualizado'})
            errores = sum(1 for row in resultados if row['estado'] == 'error')
            if errores:
                messages.warning(request, f'La carga finalizo con {creados} filas procesadas y {errores} errores.')
            else:
                messages.success(request, f'Se procesaron {creados} solicitantes correctamente.')
    else:
        form = EmployeeBulkUploadForm()

    return render(request, 'pagador/carga_empleados.html', {
        'form': form,
        'resultados': resultados,
        'empresa': request.empresa,
        **build_pagador_collaborators_context(
            request.empresa,
            search=request.GET.get('empleado_search', '').strip(),
            estado_filter=request.GET.get('empleado_estado', '').strip(),
        ),
        'dashboard_title': 'Gestión de solicitantes',
        'section_description': 'Consulta solicitantes detectados desde convenios, solicitudes y créditos de la institución.',
        'legacy_flow_title': 'Completar datos de usuarios existentes',
        'legacy_flow_description': (
            'Usa esta opción solo cuando ya cargaste la base operativa y quieres completar '
            'correo o nombre de usuarios antiguos que ya estaban relacionados con la institución.'
        ),
    })


@login_required(login_url='/pagador/login/')
@pagador_required
@require_http_methods(["POST"])
def pagador_reconciliar_empleados_view(request):
    resultados = reconciliar_usuarios_empleados_legacy(empresa=request.empresa)
    messages.success(
        request,
        f'Se revisaron {len(resultados)} usuarios ya vinculados a {request.empresa.nombre} para completar sus datos básicos.'
    )
    return redirect('pagador:carga_empleados')


@login_required(login_url='/pagador/login/')
@pagador_required
@require_http_methods(["POST"])
def pagador_actualizar_empleado_view(request, vinculo_id):
    vinculo = get_object_or_404(
        VinculoLaboralEmpresa.objects.select_related('usuario', 'empresa'),
        pk=vinculo_id,
        empresa=request.empresa,
    )
    form = EmployeeDirectUpdateForm(request.POST)
    if not form.is_valid():
        messages.error(request, 'No se pudo actualizar el solicitante. Revisa los campos obligatorios.')
        return redirect('pagador:carga_empleados')

    data = form.cleaned_data
    vinculo.nombre_empleado = data['nombre_empleado'].strip().upper()
    vinculo.documento_empleado = data['documento_empleado']
    vinculo.correo_empleado = (data.get('correo_empleado') or '').strip().lower()
    vinculo.telefono_empleado = ''.join(ch for ch in (data.get('telefono_empleado') or '') if ch.isdigit())
    vinculo.fecha_alta_aprobado = data['fecha_alta_aprobado']
    vinculo.salario_base_mensual = data.get('salario_base_mensual') or Decimal('0.00')
    vinculo.auxilio_transporte_mensual = data.get('auxilio_transporte_mensual') or Decimal('0.00')
    vinculo.descuentos_fijos_mensuales = data.get('descuentos_fijos_mensuales') or Decimal('0.00')
    vinculo.estado_vinculo = data['estado_vinculo']
    vinculo.validado_por_pagador = data.get('validado_por_pagador', False)
    vinculo.observaciones = 'Actualizado desde gestión directa del convenio.'
    vinculo.cargado_por = request.user
    vinculo.save(update_fields=[
        'nombre_empleado',
        'documento_empleado',
        'correo_empleado',
        'telefono_empleado',
        'fecha_alta_aprobado',
        'salario_base_mensual',
        'auxilio_transporte_mensual',
        'descuentos_fijos_mensuales',
        'estado_vinculo',
        'validado_por_pagador',
        'observaciones',
        'cargado_por',
        'actualizado_en',
    ])

    user_updates = []
    if vinculo.correo_empleado and vinculo.usuario.email != vinculo.correo_empleado:
        vinculo.usuario.email = vinculo.correo_empleado
        user_updates.append('email')
    name_parts = vinculo.nombre_empleado.split()
    first_name = name_parts[0] if name_parts else ''
    last_name = ' '.join(name_parts[1:]) if len(name_parts) > 1 else ''
    if first_name and vinculo.usuario.first_name != first_name[:150]:
        vinculo.usuario.first_name = first_name[:150]
        user_updates.append('first_name')
    if vinculo.usuario.last_name != last_name[:150]:
        vinculo.usuario.last_name = last_name[:150]
        user_updates.append('last_name')
    if user_updates:
        vinculo.usuario.save(update_fields=user_updates)

    messages.success(request, f'Solicitante {vinculo.nombre_empleado} actualizado.')
    return redirect('pagador:carga_empleados')


PAGADOR_PER_PAGE_CHOICES = (10, 20, 50)
PAGADOR_ROUNDING_RESIDUAL_THRESHOLD = Decimal('2.00')
def _pagador_allowed_document_fields():
    return (
        ('cedula_frontal', 'Cedula frontal'),
        ('cedula_trasera', 'Cedula reverso'),
        ('certificado_laboral', 'Contrato / soporte laboral'),
        ('certificado_bancario', 'Certificado bancario'),
    )


def _infer_pagador_document_kind(file_field):
    filename = (getattr(file_field, 'name', '') or '').split('?', 1)[0].lower()
    if filename.endswith('.pdf'):
        return 'pdf'
    if filename.endswith(('.png', '.jpg', '.jpeg', '.webp', '.gif')):
        return 'image'
    return 'file'


def _build_pagador_credit_documents(request, credito):
    allowed_fields = (
        _pagador_allowed_document_fields()
    )
    documents = []
    detalle = getattr(credito, 'detalle_libranza', None)
    if not detalle:
        return documents

    preview_path = reverse('pagador:documento_preview')
    for field_name, label in allowed_fields:
        file_field = getattr(detalle, field_name, None)
        if not file_field or not getattr(file_field, 'name', ''):
            continue
        estado = 'Cargado'
        if field_name == 'certificado_bancario':
            estado = dict([
                ('pendiente', 'Pendiente extraccion'),
                ('completo', 'Validado'),
                ('error', 'Requiere revision'),
            ]).get(detalle.certificado_bancario_estado_extraccion, 'Cargado')
        documents.append({
            'title': label,
            'url': request.build_absolute_uri(f"{preview_path}?path={quote(file_field.name)}"),
            'kind': _infer_pagador_document_kind(file_field),
            'source': 'Solicitud',
            'status': estado,
            'created_at': credito.fecha_solicitud,
            'signed_at': None,
            'description': f"{detalle.nombre_completo} - {detalle.cedula}",
        })
    return documents


def _pagador_can_access_document_path(empresa, path):
    if not path:
        return False
    detalles = CreditoLibranza.objects.filter(empresa=empresa)
    for detalle in detalles:
        for field_name, _label in _pagador_allowed_document_fields():
            file_field = getattr(detalle, field_name, None)
            if file_field and getattr(file_field, 'name', '') == path:
                return True
    return False


def _get_pagador_per_page(request):
    try:
        per_page = int(request.GET.get('per_page', 10))
    except (TypeError, ValueError):
        per_page = 10
    return per_page if per_page in PAGADOR_PER_PAGE_CHOICES else 10


def _pagador_estado_priority_expression():
    return Case(
        When(estado=Credito.EstadoCredito.EN_REVISION, then=Value(10)),
        When(estado=Credito.EstadoCredito.PENDIENTE_FIRMA, then=Value(20)),
        When(estado=Credito.EstadoCredito.PENDIENTE_TRANSFERENCIA, then=Value(30)),
        When(estado=Credito.EstadoCredito.ACTIVO, then=Value(40)),
        When(estado=Credito.EstadoCredito.EN_MORA, then=Value(50)),
        When(estado=Credito.EstadoCredito.PAGADO, then=Value(60)),
        default=Value(90),
    )


def _build_pagador_creditos_queryset(empresa, search_query='', estado_filter='', sort_by='estado_operativo', forced_linea=None):
    total_pagado_subquery = HistorialPago.objects.filter(
        credito_id=F('pk'),
        estado=HistorialPago.EstadoPago.EXITOSO
    ).values('credito_id').annotate(total=Sum('monto')).values('total')

    creditos_empresa = Credito.objects.filter(
        Q(
            linea=Credito.LineaCredito.LIBRANZA,
            detalle_libranza__empresa=empresa
        ) |
        Q(
            linea=Credito.LineaCredito.ADELANTO_NOMINA,
            detalle_adelanto_nomina__vinculo_laboral__empresa=empresa
        )
    ).exclude(
        estado__in=[Credito.EstadoCredito.RECHAZADO, Credito.EstadoCredito.SOLICITUD]
    ).select_related(
        'detalle_libranza',
        'detalle_adelanto_nomina__vinculo_laboral__empresa',
        'usuario',
    )

    if forced_linea:
        creditos_empresa = creditos_empresa.filter(linea=forced_linea)

    decision_pagador_qs = HistorialEstado.objects.filter(
        credito_id=OuterRef('pk'),
        usuario_modificacion__perfil_pagador__isnull=False,
        estado_nuevo__in=[Credito.EstadoCredito.APROBADO_PAGADOR, Credito.EstadoCredito.RECHAZADO]
    ).order_by('-fecha')

    creditos_empresa = creditos_empresa.annotate(
        total_pagado=Coalesce(Subquery(total_pagado_subquery, output_field=DecimalField()), Value(Decimal(0)))
    ).annotate(
        pagador_decision_estado=Subquery(decision_pagador_qs.values('estado_nuevo')[:1]),
        pagador_decision_motivo=Subquery(decision_pagador_qs.values('motivo')[:1]),
    )

    creditos_empresa = creditos_empresa.annotate(
        cliente_nombre_busqueda=Case(
            When(
                linea=Credito.LineaCredito.LIBRANZA,
                then=Trim(
                    Concat(
                        Coalesce('detalle_libranza__nombres', Value('')),
                        Value(' '),
                        Coalesce('detalle_libranza__apellidos', Value(''))
                    )
                )
            ),
            When(
                linea=Credito.LineaCredito.ADELANTO_NOMINA,
                then=Coalesce('detalle_adelanto_nomina__vinculo_laboral__nombre_empleado', Value(''))
            ),
            default=Value(''),
            output_field=CharField(),
        ),
        cliente_documento_busqueda=Case(
            When(linea=Credito.LineaCredito.LIBRANZA, then=Coalesce('detalle_libranza__cedula', Value(''))),
            When(linea=Credito.LineaCredito.ADELANTO_NOMINA, then=Coalesce('detalle_adelanto_nomina__vinculo_laboral__documento_empleado', Value(''))),
            default=Value(''),
            output_field=CharField(),
        ),
        estado_priority=_pagador_estado_priority_expression(),
    )

    if search_query:
        creditos_empresa = creditos_empresa.filter(
            Q(cliente_nombre_busqueda__icontains=search_query) |
            Q(cliente_documento_busqueda__icontains=search_query)
        )

    if estado_filter:
        creditos_empresa = creditos_empresa.filter(estado=estado_filter)

    valid_sort_fields = [
        'cliente_nombre_busqueda', '-cliente_nombre_busqueda',
        'cliente_documento_busqueda', '-cliente_documento_busqueda',
        'fecha_solicitud', '-fecha_solicitud',
        'fecha_proximo_pago', '-fecha_proximo_pago',
        'monto_aprobado', '-monto_aprobado',
        'saldo_pendiente', '-saldo_pendiente',
        'estado', '-estado',
        'estado_priority', '-estado_priority'
    ]
    sort_map = {
        'cliente_nombre': 'cliente_nombre_busqueda',
        '-cliente_nombre': '-cliente_nombre_busqueda',
        'cliente_documento': 'cliente_documento_busqueda',
        '-cliente_documento': '-cliente_documento_busqueda',
    }
    sort_by_real = sort_map.get(sort_by, sort_by)
    if sort_by_real == 'estado_operativo':
        creditos_empresa = creditos_empresa.order_by('estado_priority', 'fecha_proximo_pago', '-fecha_solicitud')
    elif sort_by_real == '-estado_operativo':
        creditos_empresa = creditos_empresa.order_by('-estado_priority', '-fecha_solicitud')
    elif sort_by_real in valid_sort_fields:
        creditos_empresa = creditos_empresa.order_by(sort_by_real)
    else:
        creditos_empresa = creditos_empresa.order_by('estado_priority', 'fecha_proximo_pago', '-fecha_solicitud')

    return creditos_empresa.distinct()


def _attach_pagador_payment_context(creditos_page):
    creditos = list(creditos_page.object_list)
    cuotas = (
        CuotaAmortizacion.objects.filter(credito__in=creditos, pagada=False)
        .order_by('credito_id', 'numero_cuota')
    )

    cuotas_por_credito = {}
    for cuota in cuotas:
        cuotas_por_credito.setdefault(cuota.credito_id, []).append(cuota)

    total_visible = Decimal('0.00')
    creditos_con_pago_directo = 0
    for credito in creditos:
        cuotas_pendientes = cuotas_por_credito.get(credito.id, [])
        cuota = cuotas_pendientes[0] if cuotas_pendientes else None
        saldo_pendiente = credito.saldo_pendiente or Decimal('0.00')
        monto_sugerido = Decimal('0.00')
        if cuota:
            restante_cuota = (cuota.valor_cuota or Decimal('0.00')) - (cuota.monto_pagado or Decimal('0.00'))
            if restante_cuota < Decimal('0.00'):
                restante_cuota = Decimal('0.00')
            if (
                len(cuotas_pendientes) > 1
                and restante_cuota > Decimal('0.00')
                and restante_cuota <= PAGADOR_ROUNDING_RESIDUAL_THRESHOLD
            ):
                cuota = cuotas_pendientes[1]
                monto_sugerido = (cuota.valor_cuota or Decimal('0.00')) - (cuota.monto_pagado or Decimal('0.00'))
                if monto_sugerido < Decimal('0.00'):
                    monto_sugerido = Decimal('0.00')
            else:
                monto_sugerido = restante_cuota
        elif saldo_pendiente > Decimal('0.00'):
            if credito.valor_cuota and credito.valor_cuota > Decimal('0.00'):
                monto_sugerido = min(credito.valor_cuota, saldo_pendiente)
            else:
                monto_sugerido = saldo_pendiente
        if monto_sugerido < Decimal('0.00'):
            monto_sugerido = Decimal('0.00')

        credito.siguiente_cuota_pagador = cuota
        credito.monto_obligacion_actual = monto_sugerido.quantize(Decimal('0.01'))
        credito.monto_sugerido_pagador = credito.monto_obligacion_actual
        credito.monto_sugerido_pagador_raw = format(credito.monto_sugerido_pagador, '.2f')
        credito.obligacion_actual_label = (
            f"Cuota {cuota.numero_cuota}" if cuota else "Saldo pendiente"
        )
        credito.obligacion_actual_fecha = getattr(cuota, 'fecha_vencimiento', None)
        credito.puede_pagar_obligacion = (
            credito.estado in [Credito.EstadoCredito.ACTIVO, Credito.EstadoCredito.EN_MORA]
            and credito.monto_sugerido_pagador > Decimal('0.00')
        )
        if credito.puede_pagar_obligacion:
            total_visible += credito.monto_sugerido_pagador
            creditos_con_pago_directo += 1
    return creditos_page, total_visible, creditos_con_pago_directo


def _build_pagador_estado_resumen(base_queryset):
    estado_labels = dict(Credito.EstadoCredito.choices)
    order = [
        Credito.EstadoCredito.EN_REVISION,
        Credito.EstadoCredito.PENDIENTE_FIRMA,
        Credito.EstadoCredito.PENDIENTE_TRANSFERENCIA,
        Credito.EstadoCredito.ACTIVO,
        Credito.EstadoCredito.EN_MORA,
        Credito.EstadoCredito.PAGADO,
    ]
    counts = {}
    total = 0
    for _credito_id, estado in base_queryset.values_list('id', 'estado').distinct():
        total += 1
        counts[estado] = counts.get(estado, 0) + 1
    resumen = [{'value': '', 'label': 'Todos', 'count': total}]
    for estado in order:
        count = counts.get(estado)
        if count:
            resumen.append({
                'value': estado,
                'label': estado_labels.get(estado, estado),
                'count': count,
            })
    return resumen


def _build_pagador_dashboard_context(request, *, forced_linea=None):
    empresa = request.empresa
    search_query = request.GET.get('search', '').strip()
    estado_filter = request.GET.get('estado', '').strip()
    sort_by = request.GET.get('sort_by', 'estado_operativo').strip() or 'estado_operativo'
    per_page = _get_pagador_per_page(request)

    creditos_base = _build_pagador_creditos_queryset(
        empresa=empresa,
        search_query=search_query,
        sort_by=sort_by,
        forced_linea=forced_linea,
    )
    estado_resumen = _build_pagador_estado_resumen(creditos_base)
    creditos_filtrados = creditos_base.filter(estado=estado_filter) if estado_filter else creditos_base
    total_registros = creditos_filtrados.values('id').distinct().count()

    paginator = Paginator(creditos_filtrados, per_page)
    creditos_page = paginator.get_page(request.GET.get('page'))
    creditos_page, total_visible_pagable, creditos_con_pago_directo = _attach_pagador_payment_context(creditos_page)

    query_params = request.GET.copy()
    query_params.pop('page', None)
    errores_pago_masivo = request.session.pop('errores_pago_masivo', None)
    solicitudes_pendientes = creditos_base.filter(estado=Credito.EstadoCredito.EN_REVISION)
    pagadores_activos = (
        PerfilPagador.objects
        .select_related('usuario')
        .filter(empresa=empresa, es_pagador=True)
        .order_by('usuario__email', 'usuario__username')
    )

    context = {
        'empresa': empresa,
        'creditos': creditos_page,
        'errores_pago_masivo': errores_pago_masivo,
        'pago_masivo_form': PagoMasivoEmpresaUploadForm(),
        'pago_obligaciones_form': PagoObligacionesSeleccionadasForm(initial={
            'nota': 'Pago agrupado registrado por la empresa para las obligaciones seleccionadas.',
        }),
        'solicitudes_pendientes_count': solicitudes_pendientes.values('id').distinct().count(),
        'search_query': search_query,
        'estado_filter': estado_filter,
        'sort_by': sort_by,
        'per_page': per_page,
        'per_page_choices': PAGADOR_PER_PAGE_CHOICES,
        'estados_choices': [choice for choice in Credito.EstadoCredito.choices if choice[0] not in ['RECHAZADO', 'SOLICITUD']],
        'estado_resumen': estado_resumen,
        'dashboard_title': 'Apoyos educativos' if forced_linea == Credito.LineaCredito.ADELANTO_NOMINA else 'Créditos y solicitudes',
        'section_description': (
            'Consulta y aplica pagos de adelantos vigentes desde un solo listado.'
            if forced_linea == Credito.LineaCredito.ADELANTO_NOMINA
            else 'Gestiona solicitudes, créditos vigentes y pagos directos sin depender del Excel como flujo principal.'
        ),
        'total_registros': total_registros,
        'querystring_without_page': query_params.urlencode(),
        'solo_adelantos': forced_linea == Credito.LineaCredito.ADELANTO_NOMINA,
        'next_url': request.get_full_path(),
        'creditos_con_pago_directo': creditos_con_pago_directo,
        'total_visible_pagable': total_visible_pagable,
        'show_excel_fallback': forced_linea != Credito.LineaCredito.ADELANTO_NOMINA,
        'pagadores_activos': pagadores_activos,
        'pagadores_activos_count': pagadores_activos.count(),
    }
    return context


def _pagador_redirect_target(request):
    next_url = (request.POST.get('next') or '').strip()
    if next_url.startswith('/'):
        return next_url
    return (
        reverse('pagador:adelantos_dashboard')
        if request.POST.get('origen') == 'adelantos'
        else reverse('pagador:dashboard')
    )


@login_required(login_url='/pagador/login/')
@pagador_required
def pagador_dashboard_view(request):
    """
    Dashboard para el usuario pagador de una empresa.
    Muestra todos los créditos de libranza de los empleados de su empresa, con filtros y ordenamiento.
    """
    empresa = request.empresa

    #? --- Filtros y Búsqueda ---
    search_query = request.GET.get('search', '')
    estado_filter = request.GET.get('estado', '')
    sort_by = request.GET.get('sort_by', '-monto_aprobado') # Ordenar por monto de crédito descendente por defecto

    creditos_empresa = _build_pagador_creditos_queryset(
        empresa=empresa,
        search_query=search_query,
        estado_filter=estado_filter,
        sort_by=sort_by,
    )
    total_registros = creditos_empresa.count()
    paginator = Paginator(creditos_empresa, 20)
    creditos_page = paginator.get_page(request.GET.get('page'))
    query_params = request.GET.copy()
    query_params.pop('page', None)
    errores_pago_masivo = request.session.pop('errores_pago_masivo', None)
    obligaciones_pendientes, total_obligaciones_sugerido = _build_pagador_obligaciones_context(empresa)

    # Solicitudes pendientes de validaci?n por parte del pagador.
    solicitudes_pendientes = creditos_empresa.filter(
        estado__in=[Credito.EstadoCredito.SOLICITUD, Credito.EstadoCredito.EN_REVISION]
    )

    context = {
        'empresa': empresa,
        'creditos': creditos_page,
        'errores_pago_masivo': errores_pago_masivo,
        'pago_masivo_form': PagoMasivoEmpresaUploadForm(),
        'solicitudes_pendientes_count': solicitudes_pendientes.count(),
        'search_query': search_query,
        'estado_filter': estado_filter,
        'sort_by': sort_by,
        'estados_choices': [choice for choice in Credito.EstadoCredito.choices if choice[0] not in ['RECHAZADO', 'SOLICITUD']],
        'dashboard_title': 'Panel del pagador',
        'total_registros': total_registros,
        'querystring_without_page': query_params.urlencode(),
        'solo_adelantos': False,
        'obligaciones_pendientes': obligaciones_pendientes,
        'total_obligaciones_sugerido': total_obligaciones_sugerido,
        'pago_obligaciones_form': PagoObligacionesSeleccionadasForm(initial={
            'nota': 'Pago agrupado registrado por la empresa para las obligaciones seleccionadas.',
        }),
    }
    
    return render(request, 'pagador/pagador_dashboard.html', context)


@login_required(login_url='/pagador/login/')
@pagador_required
def pagador_adelantos_dashboard_view(request):
    empresa = request.empresa
    search_query = request.GET.get('search', '')
    estado_filter = request.GET.get('estado', '')
    sort_by = request.GET.get('sort_by', '-fecha_solicitud')

    creditos_empresa = _build_pagador_creditos_queryset(
        empresa=empresa,
        search_query=search_query,
        estado_filter=estado_filter,
        sort_by=sort_by,
        forced_linea=Credito.LineaCredito.ADELANTO_NOMINA,
    )
    total_registros = creditos_empresa.count()
    paginator = Paginator(creditos_empresa, 20)
    creditos_page = paginator.get_page(request.GET.get('page'))
    query_params = request.GET.copy()
    query_params.pop('page', None)
    errores_pago_masivo = request.session.pop('errores_pago_masivo', None)
    obligaciones_pendientes, total_obligaciones_sugerido = _build_pagador_obligaciones_context(empresa)
    solicitudes_pendientes = creditos_empresa.filter(
        estado__in=[Credito.EstadoCredito.SOLICITUD, Credito.EstadoCredito.EN_REVISION]
    )

    context = {
        'empresa': empresa,
        'creditos': creditos_page,
        'errores_pago_masivo': errores_pago_masivo,
        'pago_masivo_form': PagoMasivoEmpresaUploadForm(),
        'solicitudes_pendientes_count': solicitudes_pendientes.count(),
        'search_query': search_query,
        'estado_filter': estado_filter,
        'sort_by': sort_by,
        'estados_choices': [choice for choice in Credito.EstadoCredito.choices if choice[0] not in ['RECHAZADO', 'SOLICITUD']],
        'dashboard_title': 'Apoyos educativos',
        'total_registros': total_registros,
        'querystring_without_page': query_params.urlencode(),
        'solo_adelantos': True,
        'obligaciones_pendientes': obligaciones_pendientes,
        'total_obligaciones_sugerido': total_obligaciones_sugerido,
        'pago_obligaciones_form': PagoObligacionesSeleccionadasForm(initial={
            'nota': 'Pago agrupado registrado por la empresa para las obligaciones seleccionadas.',
        }),
    }

    return render(request, 'pagador/pagador_dashboard.html', context)


@login_required(login_url='/pagador/login/')
@pagador_required
def pagador_dashboard_view(request):
    return render(request, 'pagador/pagador_dashboard.html', _build_pagador_dashboard_context(request))


@login_required(login_url='/pagador/login/')
@pagador_required
def pagador_adelantos_dashboard_view(request):
    return render(
        request,
        'pagador/pagador_dashboard.html',
        _build_pagador_dashboard_context(request, forced_linea=Credito.LineaCredito.ADELANTO_NOMINA),
    )


def _get_pagador_credito_or_404(request, credito_id):
    credito = get_object_or_404(
        Credito.objects.select_related('detalle_libranza', 'detalle_adelanto_nomina__vinculo_laboral__empresa'),
        id=credito_id,
        linea__in=[Credito.LineaCredito.LIBRANZA, Credito.LineaCredito.ADELANTO_NOMINA],
    )
    if credito.empresa_relacionada != request.empresa:
        raise Http404("Credito no encontrado.")
    return credito


@login_required(login_url='/pagador/login/')
@pagador_required
def pagador_documentacion_credito_view(request, credito_id):
    credito = _get_pagador_credito_or_404(request, credito_id)
    documentos = _build_pagador_credit_documents(request, credito)
    return render(request, 'pagador/pagador_documentos_credito.html', {
        'credito': credito,
        'documentos': documentos,
        'total_documentos': len(documentos),
        'empresa_credito': credito.empresa_relacionada,
    })


@login_required(login_url='/pagador/login/')
@pagador_required
@xframe_options_exempt
def pagador_documento_preview_view(request):
    path = (request.GET.get('path') or '').strip()
    if not _pagador_can_access_document_path(request.empresa, path):
        raise Http404("Documento no encontrado.")
    try:
        full_path = safe_join(settings.MEDIA_ROOT, path)
    except SuspiciousFileOperation:
        raise Http404("Documento no encontrado.")
    if not os.path.exists(full_path):
        raise Http404("Documento no encontrado.")
    content_type, _ = mimetypes.guess_type(full_path)
    response = FileResponse(open(full_path, 'rb'), content_type=content_type or 'application/octet-stream')
    response['Content-Disposition'] = f'inline; filename="{os.path.basename(full_path)}"'
    return response


@login_required(login_url='/pagador/login/')
@pagador_required
def pagador_detalle_credito_view(request, credito_id):
    """
    Muestra el detalle de un crédito específico para el pagador.
    """
    empresa = request.empresa

    credito = get_object_or_404(
        Credito.objects.select_related('detalle_libranza', 'detalle_adelanto_nomina__vinculo_laboral__empresa'),
        id=credito_id,
        linea__in=[Credito.LineaCredito.LIBRANZA, Credito.LineaCredito.ADELANTO_NOMINA],
    )

    #? Verificar que el crédito pertenece a la empresa del pagador
    if credito.empresa_relacionada != empresa:
        messages.error(request, "No tiene permiso para ver este crédito.")
        return redirect('pagador:dashboard')

    historial_pagos = HistorialPago.objects.filter(credito=credito, estado=HistorialPago.EstadoPago.EXITOSO).order_by('-fecha_aplicacion', '-fecha_pago')
    resumen_pagos = credit_services.obtener_resumen_pagos_credito(credito, historial_pagos=historial_pagos)
    total_pagado = resumen_pagos['total_pagado']
    siguiente_cuota = credito.tabla_amortizacion.filter(pagada=False).order_by('numero_cuota').first()
    monto_sugerido = (
        (siguiente_cuota.valor_cuota - (siguiente_cuota.monto_pagado or Decimal('0.00')))
        if siguiente_cuota else
        (credito.valor_cuota or Decimal('0.00'))
    )
    if monto_sugerido < Decimal('0.00'):
        monto_sugerido = Decimal('0.00')
    
    #? Usar el saldo pendiente del modelo que ya se actualiza correctamente
    saldo_pendiente = resumen_pagos['saldo_pendiente']
    pagador_decision = _obtener_decision_pagador(credito)
    capacidad_descuento = _build_capacidad_descuento_context(credito)

    context = {
        'credito': credito,
        'historial_pagos': historial_pagos,
        'total_pagado': total_pagado,
        'saldo_pendiente': saldo_pendiente,
        'fecha_proximo_pago': resumen_pagos['fecha_proximo_pago'],
        'cuotas_pagadas': resumen_pagos['cuotas_pagadas'],
        'cuotas_restantes': resumen_pagos['cuotas_restantes'],
        'pagador_decision': pagador_decision,
        'pagador_aprobado': bool(pagador_decision and pagador_decision.estado_nuevo == Credito.EstadoCredito.APROBADO_PAGADOR),
        'pagador_rechazado': bool(pagador_decision and pagador_decision.estado_nuevo == Credito.EstadoCredito.RECHAZADO),
        'empresa_credito': credito.empresa_relacionada,
        'capacidad_descuento': capacidad_descuento,
        'pago_offline_form': PagoCreditoOfflineForm(initial={
            'monto': monto_sugerido,
            'nota': 'Pago registrado por la empresa para esta cuota.',
        }),
        'monto_sugerido_pago': monto_sugerido,
    }
    
    return render(request, 'pagador/pagador_detalle_credito.html', context)


@login_required(login_url='/pagador/login/')
@require_POST
@pagador_required
def pagador_registrar_pago_offline_view(request, credito_id):
    empresa = request.empresa
    credito = get_object_or_404(
        Credito.objects.select_related('detalle_libranza', 'detalle_adelanto_nomina__vinculo_laboral__empresa'),
        id=credito_id,
        linea__in=[Credito.LineaCredito.LIBRANZA, Credito.LineaCredito.ADELANTO_NOMINA],
    )
    if credito.empresa_relacionada != empresa:
        messages.error(request, "No tiene permiso para registrar pagos sobre este credito.")
        return redirect('pagador:dashboard')

    form = PagoCreditoOfflineForm(request.POST, request.FILES)
    if not form.is_valid():
        for field_errors in form.errors.values():
            for error in field_errors:
                messages.error(request, error)
        return redirect('pagador:credito_detalle', credito_id=credito.id)

    referencia = (form.cleaned_data.get('referencia_pago') or '').strip().upper()
    if not referencia:
        referencia = f"OFFLINE-{credito.id}-{timezone.now().strftime('%Y%m%d%H%M%S%f')}"

    try:
        pago, created = credit_services.registrar_pago_credito(
            credito=credito,
            monto=form.cleaned_data['monto'],
            referencia_pago=referencia,
            metodo_pago=form.cleaned_data['metodo_pago'],
            origen_registro=HistorialPago.OrigenRegistro.REGISTRO_MANUAL_PAGADOR,
            usuario=request.user,
            empresa=empresa,
            comprobante=form.cleaned_data.get('comprobante'),
            notas=form.cleaned_data.get('nota') or 'Pago offline registrado por pagador.',
        )
        if created:
            messages.success(request, f"Pago de ${pago.monto:,.2f} aplicado correctamente.")
        else:
            messages.warning(request, f"La referencia {referencia} ya existia y no se aplico de nuevo.")
    except Exception as exc:
        logger.exception("Error al registrar pago offline de pagador para %s", credito.numero_credito)
        messages.error(request, f"No pudimos registrar el pago: {exc}")

    return redirect('pagador:credito_detalle', credito_id=credito.id)


@login_required(login_url='/pagador/login/')
@require_POST
@pagador_required
def pagador_pagar_obligaciones_seleccionadas_view(request):
    empresa = request.empresa
    form = PagoObligacionesSeleccionadasForm(request.POST, request.FILES)
    if not form.is_valid():
        for field_errors in form.errors.values():
            for error in field_errors:
                messages.error(request, error)
        return redirect('pagador:dashboard')

    selected_ids = request.POST.getlist('obligaciones')
    if not selected_ids:
        messages.error(request, 'Debes seleccionar al menos una obligaciÃ³n para aplicar el pago.')
        return redirect('pagador:dashboard')

    base_qs = Credito.objects.filter(
        Q(linea=Credito.LineaCredito.LIBRANZA, detalle_libranza__empresa=empresa)
        | Q(linea=Credito.LineaCredito.ADELANTO_NOMINA, detalle_adelanto_nomina__vinculo_laboral__empresa=empresa),
        estado__in=[Credito.EstadoCredito.ACTIVO, Credito.EstadoCredito.EN_MORA],
        id__in=selected_ids,
    ).select_related('detalle_libranza', 'detalle_adelanto_nomina__vinculo_laboral__empresa', 'usuario')

    creditos_map = {str(credito.id): credito for credito in base_qs}
    obligaciones = []
    for credito_id in selected_ids:
        credito = creditos_map.get(str(credito_id))
        if not credito:
            continue
        raw_amount = request.POST.get(f'monto_{credito.id}', '').strip()
        try:
            monto = Decimal(raw_amount)
        except Exception:
            messages.error(request, f'El monto de la obligaciÃ³n {credito.numero_credito} no es vÃ¡lido.')
            return redirect('pagador:dashboard')
        if monto <= Decimal('0.00'):
            messages.error(request, f'El monto de la obligaciÃ³n {credito.numero_credito} debe ser mayor a cero.')
            return redirect('pagador:dashboard')
        obligaciones.append({'credito': credito, 'monto': monto})

    if not obligaciones:
        messages.error(request, 'No encontramos obligaciones vÃ¡lidas para procesar.')
        return redirect('pagador:dashboard')

    try:
        from gestion_creditos.services.payments import aplicar_pago_obligaciones_seleccionadas

        pagos_aplicados = aplicar_pago_obligaciones_seleccionadas(
            empresa=empresa,
            actor=request.user,
            obligaciones=obligaciones,
            metodo_pago=form.cleaned_data['metodo_pago'],
            nota=form.cleaned_data['nota'],
            comprobante=form.cleaned_data.get('comprobante'),
        )
        total_aplicado = sum((pago.monto for pago in pagos_aplicados), Decimal('0.00'))
        messages.success(
            request,
            f'Se aplicaron {len(pagos_aplicados)} obligaciones por un total de ${total_aplicado:,.2f}.'
        )
    except Exception as exc:
        logger.exception('Error al aplicar obligaciones seleccionadas para empresa %s', empresa.nombre)
        messages.error(request, f'No fue posible aplicar las obligaciones seleccionadas: {exc}')

    return redirect('pagador:dashboard')


@login_required(login_url='/pagador/login/')
@require_POST
@pagador_required
def pagador_pagar_obligaciones_seleccionadas_view(request):
    redirect_target = _pagador_redirect_target(request)
    empresa = request.empresa
    form = PagoObligacionesSeleccionadasForm(request.POST, request.FILES)
    if not form.is_valid():
        for field_errors in form.errors.values():
            for error in field_errors:
                messages.error(request, error)
        return redirect(redirect_target)

    selected_ids = request.POST.getlist('obligaciones')
    if not selected_ids:
        messages.error(request, 'Debes seleccionar al menos una obligación para aplicar el pago.')
        return redirect(redirect_target)

    base_qs = Credito.objects.filter(
        Q(linea=Credito.LineaCredito.LIBRANZA, detalle_libranza__empresa=empresa)
        | Q(linea=Credito.LineaCredito.ADELANTO_NOMINA, detalle_adelanto_nomina__vinculo_laboral__empresa=empresa),
        estado__in=[Credito.EstadoCredito.ACTIVO, Credito.EstadoCredito.EN_MORA],
        id__in=selected_ids,
    ).select_related('detalle_libranza', 'detalle_adelanto_nomina__vinculo_laboral__empresa', 'usuario')

    creditos_map = {str(credito.id): credito for credito in base_qs}
    obligaciones = []
    for credito_id in selected_ids:
        credito = creditos_map.get(str(credito_id))
        if not credito:
            continue
        raw_amount = request.POST.get(f'monto_{credito.id}', '').strip()
        try:
            monto = Decimal(raw_amount)
        except Exception:
            messages.error(request, f'El monto del crédito {credito.numero_credito} no es válido.')
            return redirect(redirect_target)
        if monto <= Decimal('0.00'):
            messages.error(request, f'El monto del crédito {credito.numero_credito} debe ser mayor a cero.')
            return redirect(redirect_target)
        obligaciones.append({'credito': credito, 'monto': monto})

    if not obligaciones:
        messages.error(request, 'No encontramos obligaciones válidas para procesar.')
        return redirect(redirect_target)

    try:
        from gestion_creditos.services.payments import aplicar_pago_obligaciones_seleccionadas

        pagos_aplicados = aplicar_pago_obligaciones_seleccionadas(
            empresa=empresa,
            actor=request.user,
            obligaciones=obligaciones,
            metodo_pago=form.cleaned_data['metodo_pago'],
            nota=form.cleaned_data['nota'],
            comprobante=form.cleaned_data.get('comprobante'),
        )
        total_aplicado = sum((pago.monto for pago in pagos_aplicados), Decimal('0.00'))
        messages.success(
            request,
            f'Se aplicaron {len(pagos_aplicados)} obligaciones por un total de ${total_aplicado:,.2f}.'
        )
    except Exception as exc:
        logger.exception('Error al aplicar obligaciones seleccionadas para empresa %s', empresa.nombre)
        messages.error(request, f'No fue posible aplicar las obligaciones seleccionadas: {exc}')

    return redirect(redirect_target)


@login_required(login_url='/pagador/login/')
@require_POST
@pagador_required
def pagador_decidir_solicitud_view(request, credito_id):
    """
    Permite al pagador aprobar o rechazar solicitudes de libranza de su empresa.
    """
    empresa = request.empresa
    credito = get_object_or_404(
        Credito.objects.select_related('detalle_libranza', 'detalle_adelanto_nomina__vinculo_laboral__empresa'),
        id=credito_id,
        linea__in=[Credito.LineaCredito.LIBRANZA, Credito.LineaCredito.ADELANTO_NOMINA],
    )

    if credito.estado not in [Credito.EstadoCredito.SOLICITUD, Credito.EstadoCredito.EN_REVISION]:
        messages.info(request, "Esta solicitud ya no admite decisiones del convenio.")
        return redirect('pagador:dashboard')

    if credito.empresa_relacionada != empresa:
        messages.error(request, "No tiene permiso para gestionar este credito.")
        return redirect('pagador:dashboard')

    action = request.POST.get('action')
    motivo = (request.POST.get('motivo') or '').strip()

    if action not in ['approve', 'reject']:
        messages.error(request, "Accion no valida.")
        return redirect('pagador:dashboard')

    try:
        with transaction.atomic():
            credito = (
                Credito.objects
                .select_for_update()
                .get(id=credito.id)
            )

            decision_existente = _obtener_decision_pagador(credito)
            if decision_existente:
                estado_actual = (
                    "aprobada"
                    if decision_existente.estado_nuevo == Credito.EstadoCredito.APROBADO_PAGADOR
                    else "rechazada"
                )
                messages.info(request, f"La decisión del convenio ya fue registrada como {estado_actual}.")
                return redirect('pagador:dashboard')

            if credito.estado not in [Credito.EstadoCredito.SOLICITUD, Credito.EstadoCredito.EN_REVISION]:
                messages.info(request, "Esta solicitud ya cambio de estado y no admite una nueva decision.")
                return redirect('pagador:dashboard')

            if action == 'approve':
                if credito.monto_aprobado is None:
                    credito.monto_aprobado = credito.monto_solicitado
                if credito.plazo is None:
                    credito.plazo = credito.plazo_solicitado
                credito.save(update_fields=['monto_aprobado', 'plazo'])

                motivo_final = motivo or "Aceptado por convenio y enviado directamente a firma."
                credit_services.gestionar_cambio_estado_credito(
                    credito=credito,
                    nuevo_estado=Credito.EstadoCredito.APROBADO_PAGADOR,
                    usuario_modificacion=request.user,
                    motivo=motivo_final
                )
                credit_services.preparar_documento_para_firma(
                    credito=credito,
                    usuario_modificacion=request.user
                )
                messages.success(
                    request,
                    f"Solicitud {credito.numero_credito} aprobada por convenio y enviada a firma."
                )
            else:
                motivo_final = motivo or "Rechazado por convenio."
                credit_services.gestionar_cambio_estado_credito(
                    credito=credito,
                    nuevo_estado=Credito.EstadoCredito.RECHAZADO,
                    usuario_modificacion=request.user,
                    motivo=motivo_final
                )
                messages.warning(request, f"Credito {credito.numero_credito} rechazado.")
    except Exception as e:
        messages.error(request, f"Ocurrio un error al procesar la solicitud: {e}")
        logger.error(f"Error al decidir solicitud {credito.id} por pagador: {e}", exc_info=True)

    return redirect('pagador:dashboard')


@login_required(login_url='/pagador/login/')
@require_POST
@pagador_required
def pagador_procesar_pagos_view(request):
    """
    Recibe la plantilla oficial en Excel y crea un lote borrador para confirmación offline.
    """
    empresa = request.empresa
    form = PagoMasivoEmpresaUploadForm(request.POST, request.FILES)

    if not form.is_valid():
        errores = []
        for field_errors in form.errors.values():
            errores.extend(field_errors)
        request.session['errores_pago_masivo'] = errores
        messages.error(request, "No pudimos validar el archivo Excel de la carga de pagos.")
        return redirect('pagador:dashboard')

    archivo = form.cleaned_data['archivo']
    _, errores, lote = credit_services.crear_borrador_pagos_masivos_archivo(
        archivo,
        empresa,
        usuario=request.user,
    )

    if errores:
        request.session['errores_pago_masivo'] = errores
        messages.error(request, "No pudimos preparar la carga de pagos.")
        return redirect('pagador:dashboard')

    return redirect('pagador:pagos_masivos_confirmar', lote_id=lote.id)


def _build_pagador_pagos_masivos_preview(lote):
    pagos_validos, errores, _ = credit_services.validar_archivo_pagos_masivos(lote.archivo, lote.empresa)
    monto_total = sum(p['monto'] for p in pagos_validos)
    return pagos_validos, errores, monto_total


@login_required(login_url='/pagador/login/')
@pagador_required
@require_http_methods(["GET", "POST"])
def pagador_confirmar_pagos_masivos_view(request, lote_id):
    lote = get_object_or_404(LotePagoEmpresa.objects.select_related('empresa', 'creado_por'), pk=lote_id, empresa=request.empresa)
    pagos_validos, errores_preview, monto_total = _build_pagador_pagos_masivos_preview(lote)

    if lote.estado != LotePagoEmpresa.EstadoLote.CARGADO:
        messages.warning(request, f'La carga de pagos #{lote.id} ya fue procesada o no está disponible para confirmación.')
        return redirect('pagador:dashboard')

    if request.method == 'POST':
        form = PagoMasivoEmpresaConfirmForm(request.POST, request.FILES, instance=lote)
        if form.is_valid():
            pagos_exitosos, errores = credit_services.procesar_lote_pago_empresa(
                lote,
                usuario=request.user,
                comprobante=form.cleaned_data.get('comprobante'),
                notas=form.cleaned_data.get('notas') or '',
            )
            if errores:
                context = {
                    'empresa': request.empresa,
                    'lote': lote,
                    'pagos_validos': pagos_validos,
                    'errores_pago_masivo': errores,
                    'monto_total': monto_total,
                    'cantidad_pagos': len(pagos_validos),
                    'form': form,
                }
                return render(request, 'pagador/confirmacion_pago_masivo.html', context)

            try:
                from ..email_service import enviar_resumen_pago_masivo_pagador
                enviar_resumen_pago_masivo_pagador(
                    lote=lote,
                    pagos_aplicados=pagos_exitosos,
                    monto_total=monto_total,
                    pagador_email=request.user.email,
                    pagador_nombre=request.user.get_full_name() or request.user.username,
                )
            except Exception:
                logger.exception(
                    "No pudimos enviar el resumen de la carga de pagos al pagador para lote %s",
                    lote.id,
                )

            messages.success(
                request,
                f'Se aplicaron {pagos_exitosos} pagos correctamente. La carga de pagos #{lote.id} quedó registrada.'
            )
            return redirect('pagador:dashboard')
    else:
        form = PagoMasivoEmpresaConfirmForm(instance=lote)

    context = {
        'empresa': request.empresa,
        'lote': lote,
        'pagos_validos': pagos_validos,
        'errores_pago_masivo': errores_preview,
        'monto_total': monto_total,
        'cantidad_pagos': len(pagos_validos),
        'form': form,
    }
    return render(request, 'pagador/confirmacion_pago_masivo.html', context)


def _pagador_report_rows(request, creditos, empresa, report_type):
    def _fmt_dt(value):
        if not value:
            return ''
        try:
            return value.strftime('%d/%m/%Y %H:%M')
        except AttributeError:
            return str(value)

    def _fmt_decimal(value):
        return f'{value}' if value is not None else ''

    def _file_url(file_field):
        if not file_field:
            return ''
        try:
            return request.build_absolute_uri(file_field.url)
        except (ValueError, AttributeError):
            return file_field.name

    headers_completo = [
        'Empresa', 'Numero credito', 'Estado', 'Linea', 'Fecha solicitud', 'Fecha actualizacion',
        'Fecha desembolso', 'Monto solicitado', 'Plazo solicitado', 'Monto aprobado', 'Plazo aprobado',
        'Tasa interes mensual', 'Comision', 'IVA comision', 'Total a pagar', 'Saldo pendiente',
        'Capital pendiente', 'Valor cuota', 'Fecha proximo pago', 'Total pagado', 'Fecha ultimo pago',
        'Documento enviado', 'Usuario', 'Email usuario', 'Nombre completo', 'Nombres', 'Apellidos',
        'Cedula', 'Correo', 'Telefono', 'Direccion', 'Ingresos mensuales', 'Cedula frontal',
        'Cedula trasera', 'Certificado laboral', 'Desprendible nomina', 'Certificado bancario',
        'Pagare numero', 'Pagare estado', 'Pagare estado codigo', 'Pagare fecha envio',
        'Pagare fecha firma',
    ]
    headers_reducido = [
        'Empresa', 'Numero credito', 'Estado', 'Linea', 'Fecha solicitud', 'Monto solicitado',
        'Plazo solicitado', 'Monto aprobado', 'Plazo aprobado', 'Saldo pendiente', 'Valor cuota',
        'Fecha proximo pago', 'Total pagado', 'Nombre completo', 'Cedula', 'Correo',
        'Telefono', 'Pagare estado', 'Pagare fecha firma',
    ]

    rows = []
    for credito in creditos:
        detalle = credito.detalle
        usuario = credito.usuario
        try:
            pagare = credito.pagare
        except Pagare.DoesNotExist:
            pagare = None

        if credito.linea == Credito.LineaCredito.LIBRANZA:
            nombre_completo = detalle.nombre_completo if detalle else ''
            nombres = detalle.nombres if detalle else ''
            apellidos = detalle.apellidos if detalle else ''
            cedula = detalle.cedula if detalle else ''
            correo = detalle.correo_electronico if detalle else ''
            telefono = detalle.telefono if detalle else ''
            direccion = detalle.direccion if detalle else ''
            ingresos = _fmt_decimal(getattr(detalle, 'ingresos_mensuales', None)) if detalle else ''
            cedula_frontal = _file_url(detalle.cedula_frontal) if detalle else ''
            cedula_trasera = _file_url(detalle.cedula_trasera) if detalle else ''
            certificado_laboral = _file_url(detalle.certificado_laboral) if detalle else ''
            desprendible_nomina = _file_url(detalle.desprendible_nomina) if detalle else ''
            certificado_bancario = _file_url(detalle.certificado_bancario) if detalle else ''
        else:
            vinculo = detalle.vinculo_laboral if detalle else None
            nombre_completo = vinculo.nombre_empleado if vinculo else ''
            nombres = vinculo.nombre_empleado if vinculo else ''
            apellidos = ''
            cedula = vinculo.documento_empleado if vinculo else ''
            correo = vinculo.correo_empleado if vinculo else ''
            telefono = vinculo.telefono_empleado if vinculo else ''
            direccion = ''
            ingresos = _fmt_decimal(getattr(vinculo, 'salario_base_mensual', None)) if vinculo else ''
            cedula_frontal = ''
            cedula_trasera = ''
            certificado_laboral = ''
            desprendible_nomina = ''
            certificado_bancario = ''

        row_completo = [
            empresa.nombre,
            credito.numero_credito,
            credito.get_estado_display(),
            credito.get_linea_display(),
            _fmt_dt(credito.fecha_solicitud),
            _fmt_dt(credito.fecha_actualizacion),
            _fmt_dt(credito.fecha_desembolso),
            _fmt_decimal(credito.monto_solicitado),
            credito.plazo_solicitado or '',
            _fmt_decimal(credito.monto_aprobado),
            credito.plazo or '',
            _fmt_decimal(credito.tasa_interes),
            _fmt_decimal(credito.comision),
            _fmt_decimal(credito.iva_comision),
            _fmt_decimal(credito.total_a_pagar),
            _fmt_decimal(credito.saldo_pendiente),
            _fmt_decimal(credito.capital_pendiente),
            _fmt_decimal(credito.valor_cuota),
            _fmt_dt(credito.fecha_proximo_pago),
            _fmt_decimal(getattr(credito, 'total_pagado', None)),
            _fmt_dt(getattr(credito, 'ultimo_pago', None)),
            'Si' if credito.documento_enviado else 'No',
            usuario.username if usuario else '',
            usuario.email if usuario else '',
            nombre_completo,
            nombres,
            apellidos,
            cedula,
            correo,
            telefono,
            direccion,
            ingresos,
            cedula_frontal,
            cedula_trasera,
            certificado_laboral,
            desprendible_nomina,
            certificado_bancario,
            pagare.numero_pagare if pagare else '',
            pagare.get_estado_display() if pagare else '',
            pagare.estado if pagare else '',
            _fmt_dt(pagare.fecha_envio) if pagare else '',
            _fmt_dt(pagare.fecha_firma) if pagare else '',
        ]
        row_reducido = [
            empresa.nombre,
            credito.numero_credito,
            credito.get_estado_display(),
            credito.get_linea_display(),
            _fmt_dt(credito.fecha_solicitud),
            _fmt_decimal(credito.monto_solicitado),
            credito.plazo_solicitado or '',
            _fmt_decimal(credito.monto_aprobado),
            credito.plazo or '',
            _fmt_decimal(credito.saldo_pendiente),
            _fmt_decimal(credito.valor_cuota),
            _fmt_dt(credito.fecha_proximo_pago),
            _fmt_decimal(getattr(credito, 'total_pagado', None)),
            nombre_completo,
            cedula,
            correo,
            telefono,
            pagare.get_estado_display() if pagare else '',
            _fmt_dt(pagare.fecha_firma) if pagare else '',
        ]
        rows.append(row_reducido if report_type == 'reducido' else row_completo)

    return (headers_reducido if report_type == 'reducido' else headers_completo), rows


@login_required(login_url='/pagador/login/')
@pagador_required
def descargar_csv_cuotas_pendientes_view(request):
    """
    Genera la plantilla oficial en Excel para carga offline de cuotas.
    """
    empresa = request.empresa

    creditos = Credito.objects.filter(
        linea=Credito.LineaCredito.LIBRANZA,
        detalle_libranza__empresa=empresa,
        estado__in=[Credito.EstadoCredito.ACTIVO, Credito.EstadoCredito.EN_MORA]
    ).select_related('detalle_libranza').order_by('detalle_libranza__cedula')

    from decimal import Decimal, ROUND_CEILING
    rows = []
    for credito in creditos:
        cedula = str(credito.detalle_libranza.cedula)

        cuota = credito.tabla_amortizacion.filter(pagada=False).order_by('numero_cuota').first()
        if cuota:
            valor = cuota.valor_cuota - (cuota.monto_pagado or Decimal('0.00'))
        else:
            valor = credito.valor_cuota or Decimal('0.00')

        if valor < 0:
            valor = Decimal('0.00')

        monto = int(valor.to_integral_value(rounding=ROUND_CEILING))

        rows.append([cedula, monto, '', '', ''])

    filename = f'cuotas_pendientes_{empresa.nombre}_{timezone.now().strftime("%Y%m%d")}'
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Pagos a cargar'

    headers = ['cedula', 'monto_a_pagar', 'referencia_pago', 'fecha_pago', 'nota']
    sheet.append(headers)
    for row in rows:
        sheet.append(row)

    header_fill = PatternFill(fill_type='solid', fgColor='0B5ED7')
    header_font = Font(color='FFFFFF', bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    widths = {'A': 18, 'B': 18, 'C': 26, 'D': 18, 'E': 42}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    sheet.freeze_panes = 'A2'
    sheet['D1'].comment = Comment(
        'Usa el formato DD/MM/AAAA. Ejemplo: 30/03/2026. '
        'Si no diligencias la fecha, el sistema tomará la fecha actual.',
        'Aprobado',
    )
    sheet['E1'].comment = Comment(
        'Describe brevemente el contexto del pago. Ejemplo: Nómina marzo pagada por transferencia.',
        'Aprobado',
    )

    for row_idx in range(2, len(rows) + 2):
        sheet[f'D{row_idx}'].number_format = 'DD/MM/YYYY'
        sheet[f'E{row_idx}'].alignment = Alignment(wrap_text=True)

    date_validation = DataValidation(
        type='date',
        operator='between',
        formula1='DATE(2020,1,1)',
        formula2='DATE(2100,12,31)',
        allow_blank=True,
    )
    date_validation.prompt = 'Usa una fecha con formato DD/MM/AAAA. Ejemplo: 30/03/2026.'
    date_validation.promptTitle = 'Fecha de pago'
    date_validation.error = 'La fecha debe estar en formato DD/MM/AAAA. Ejemplo: 30/03/2026.'
    date_validation.errorTitle = 'Fecha inválida'
    sheet.add_data_validation(date_validation)
    date_validation.add(f'D2:D{max(len(rows) + 1, 500)}')

    instrucciones = workbook.create_sheet(title='Instrucciones')
    instrucciones['A1'] = 'Cómo usar esta plantilla'
    instrucciones['A1'].font = Font(bold=True, size=13)
    instrucciones['A3'] = '1. No cambies los nombres de las columnas.'
    instrucciones['A4'] = '2. Si ya aparece el valor sugerido de la cuota, solo ajusta si la empresa pagó un monto diferente.'
    instrucciones['A5'] = '3. En fecha_pago usa siempre DD/MM/AAAA. Ejemplo: 30/03/2026.'
    instrucciones['A6'] = '4. La nota debe explicar el contexto del pago. Ejemplo: Nómina marzo pagada por transferencia.'
    instrucciones['A7'] = '5. Después de subir el archivo, podrás revisar la carga y adjuntar el comprobante antes de confirmar.'
    instrucciones.column_dimensions['A'].width = 110
    workbook.save(response)
    return response


@login_required(login_url='/pagador/login/')
@pagador_required
def descargar_reporte_pagador_view(request):
    """
    Genera y descarga en Excel el reporte operativo de los creditos de libranza de la empresa.
    """
    empresa = request.empresa
    search_query = request.GET.get('search', '').strip()
    estado_filter = request.GET.get('estado', '').strip()
    sort_by = request.GET.get('sort_by', 'cliente_documento')

    creditos = _build_pagador_creditos_queryset(empresa, search_query, estado_filter, sort_by)

    report_type = request.GET.get('tipo', 'completo').strip().lower()
    if report_type not in ['completo', 'reducido']:
        report_type = 'completo'

    headers, rows = _pagador_report_rows(request, creditos, empresa, report_type)
    filename = f'reporte_pagador_{report_type}_{empresa.nombre}_{timezone.now().strftime("%Y%m%d")}'
    from openpyxl import Workbook

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet(title='Reporte pagador')
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    workbook.save(response)
    return response


@login_required(login_url='/pagador/login/')
@pagador_required
def iniciar_pago_view(request, credito_id):
    """Inicia el flujo de pago para una cuota de un crédito de libranza."""
    credito = get_object_or_404(Credito, id=credito_id, linea=Credito.LineaCredito.LIBRANZA)
    
    #? Asegurarse de que el pagador solo pueda pagar créditos de su empresa
    if credito.detalle_libranza.empresa != request.empresa:
        messages.error(request, "No tiene permisos para pagar este crédito.")
        return redirect('pagador:dashboard')

    valor_cuota = credito.valor_cuota
    if not valor_cuota or valor_cuota <= 0:
        messages.error(request, "El crédito no tiene un valor de cuota válido para pagar.")
        return redirect('pagador:dashboard')

    context = {
        'credito': credito,
        'valor_cuota': valor_cuota,
        'referencia_pago': f"ONLINE-{credito.id}-{timezone.now().strftime('%Y%m%d%H%M%S%f')}"
    }
    return render(request, 'pagador/simulacion_pago.html', context)


@login_required
@require_POST
def procesar_pago_callback_view(request):
    """Procesa el callback de la pasarela de pagos simulada."""
    status = request.POST.get('status')
    credito_id = request.POST.get('credito_id')
    monto = request.POST.get('monto')
    print("monto: ",monto)
    referencia = request.POST.get('referencia')

    credito = get_object_or_404(Credito, id=credito_id)
    if status == 'success':
        try:
            if not monto:
                raise ValueError("El monto recibido está vacío.")

            with transaction.atomic():
                #? Eliminamos los puntos y reemplazamos la coma por punto para Decimal
                monto_limpio = monto.replace('.', '').replace(',', '.')
                #? Convertir a Decimal (Ej: "1,234.56" -> Decimal("1234.56")
                monto_decimal = Decimal(monto_limpio)
                
                credit_services.registrar_pago_credito(
                    credito=credito,
                    monto=monto_decimal,
                    referencia_pago=referencia,
                    metodo_pago=HistorialPago.MetodoPago.OFFLINE_MANUAL,
                    origen_registro=HistorialPago.OrigenRegistro.REGISTRO_MANUAL_PAGADOR,
                    usuario=request.user if request.user.is_authenticated else None,
                    empresa=credito.empresa_relacionada,
                    notas='Pago registrado por callback legacy de simulacion.',
                )
                
                messages.success(request, f"Pago de ${monto_decimal:,.2f} para el crédito #{credito.id} procesado exitosamente.")

        except (ValueError, TypeError, decimal.ConversionSyntax) as e:
            messages.error(request, f"Ocurrió un error al procesar el pago: {e}")
        except Exception as e:
            messages.error(request, f"Ocurrió un error inesperado al procesar el pago: {e}")
    else:
        messages.error(request, f"El pago para el crédito #{credito.id} fue fallido o cancelado.")

    return redirect('gestion_creditos:pagador_dashboard')


def _parse_wompi_datetime(value):
    if not value:
        return None
    try:
        return timezone.localtime(datetime.fromisoformat(value.replace('Z', '+00:00')))
    except (ValueError, TypeError):
        return None


def _extract_credito_id_from_reference(reference):
    if not reference or '-' not in reference:
        return None
    parts = reference.split('-')
    if len(parts) < 2:
        return None
    try:
        return int(parts[1])
    except (TypeError, ValueError):
        return None


def _get_metodo_pago_wompi(transaction_data):
    payment_method = transaction_data.get('payment_method') or {}
    method_type = transaction_data.get('payment_method_type') or payment_method.get('type')
    if not method_type:
        return None, None

    extra = payment_method.get('extra') or {}
    banco = extra.get('financial_institution_name') or extra.get('financial_institution_code')
    detalle = None
    last_four = extra.get('last_four') or extra.get('card_last_four')
    if last_four:
        detalle = f"**** {last_four}"
    elif payment_method.get('phone_number'):
        detalle = payment_method.get('phone_number')
    return method_type, banco or detalle


def _enviar_resumen_pago_pagador(request, credito, transaction_data):
    try:
        from gestion_creditos.email_service import enviar_confirmacion_pago
    except Exception:
        return

    pagador_email = request.user.email
    if not pagador_email:
        return

    transaction_id = transaction_data.get('id')
    reference = transaction_data.get('reference')
    cache_key = f"wompi:pagador:email:{reference or transaction_id}:{pagador_email}"
    if not cache.add(cache_key, True, timeout=86400):
        return

    monto_pagado = Decimal(transaction_data.get('amount_in_cents', 0)) / 100
    metodo_pago, banco = _get_metodo_pago_wompi(transaction_data)
    fecha_pago = _parse_wompi_datetime(
        transaction_data.get('finalized_at') or transaction_data.get('created_at')
    )
    cta_url = request.build_absolute_uri(
        reverse('pagador:pago_wompi_resumen', kwargs={'transaction_id': transaction_id})
    ) if transaction_id else request.build_absolute_uri(reverse('pagador:dashboard'))

    enviar_confirmacion_pago(
        credito,
        monto_pagado,
        credito.saldo_pendiente or Decimal('0.00'),
        destinatario=pagador_email,
        nombre_destinatario=request.user.get_full_name() or request.user.username,
        referencia=reference,
        metodo_pago=metodo_pago,
        banco=banco,
        fecha_pago=fecha_pago,
        cta_url=cta_url,
        cta_label='Ver comprobante'
    )


def _get_credito_pagador_from_reference(request, reference):
    credito_id = request.session.get('credito_id') or _extract_credito_id_from_reference(reference)
    if not credito_id:
        return None
    return Credito.objects.filter(
        id=credito_id,
        linea=Credito.LineaCredito.LIBRANZA,
        detalle_libranza__empresa=request.empresa
    ).select_related('detalle_libranza').first()


@login_required(login_url='/pagador/login/')
@pagador_required
def iniciar_pago_wompi_view(request, credito_id):
    """
    Muestra el formulario de selección de m?todo de pago con WOMPI
    """
    from ..services.wompi_client import WompiClient, WompiAPIException

    credito = get_object_or_404(Credito, id=credito_id, linea=Credito.LineaCredito.LIBRANZA)

    # Verificar que el pagador tenga permisos
    if credito.detalle_libranza.empresa != request.empresa:
        messages.error(request, "No tiene permisos para pagar este crédito.")
        return redirect('pagador:dashboard')

    valor_cuota = credito.valor_cuota
    if not valor_cuota or valor_cuota <= 0:
        messages.error(request, "El crédito no tiene un valor de cuota válido para pagar.")
        return redirect('pagador:dashboard')

    cuota_pendiente = credito.tabla_amortizacion.filter(pagada=False).order_by('numero_cuota').first()
    if not cuota_pendiente:
        messages.error(request, "Este credito no tiene cuotas pendientes por pagar.")
        return redirect('pagador:dashboard')

    # Obtener acceptance token de WOMPI
    client = WompiClient()
    try:
        acceptance_response = client.get_acceptance_token()
        acceptance_token = acceptance_response['data']['presigned_acceptance']['acceptance_token']

        # Obtener lista de bancos PSE
        bancos_pse = client.get_pse_financial_institutions()
    except WompiAPIException as e:
        logger.error(f"Error al obtener datos de WOMPI: {str(e)}")
        messages.error(request, "Error al conectar con la pasarela de pagos. Por favor intenta más tarde.")
        return redirect('pagador:dashboard')

    context = {
        'credito': credito,
        'valor_cuota': int(valor_cuota),
        'valor_cuota_centavos': int(valor_cuota * 100),  # Convertir a centavos
        'referencia_pago': f"CUOTA-{credito.id}-{cuota_pendiente.numero_cuota}",
        'acceptance_token': acceptance_token,
        'bancos_pse': bancos_pse,
        'customer_email': credito.detalle_libranza.correo_electronico,
        'customer_name': credito.detalle_libranza.nombre_completo,
        'customer_phone': credito.detalle_libranza.telefono,
        'wompi_public_key': settings.WOMPI_PUBLIC_KEY,
    }

    return render(request, 'pagador/pago_wompi.html', context)


@login_required
@pagador_required
@require_POST
def procesar_pago_wompi_view(request):
    """
    Procesa el pago con WOMPI segun el metodo seleccionado.
    """
    from ..services.wompi_client import WompiClient, WompiAPIException

    intent = None
    wants_json = 'application/json' in (request.content_type or '')

    try:
        payload = request.POST
        if wants_json:
            try:
                payload = json.loads(request.body.decode('utf-8'))
            except json.JSONDecodeError:
                return JsonResponse({'error': 'Invalid JSON'}, status=400)

        payment_method_type = payload.get('payment_method')
        credito_id = payload.get('credito_id')
        amount_in_cents_raw = payload.get('amount_in_cents')
        reference = payload.get('reference')
        customer_email = payload.get('customer_email')
        acceptance_token = payload.get('acceptance_token')
        tipo_pago = (payload.get('tipo_pago') or '').upper()

        if not amount_in_cents_raw or not reference or not credito_id:
            if wants_json:
                return JsonResponse({'error': 'Missing payment data'}, status=400)
            messages.error(request, 'Datos de pago incompletos.')
            return redirect('pagador:dashboard')

        try:
            amount_in_cents = int(amount_in_cents_raw)
        except (TypeError, ValueError):
            if wants_json:
                return JsonResponse({'error': 'Invalid amount'}, status=400)
            messages.error(request, 'Monto invalido.')
            return redirect('pagador:dashboard')

        credito = get_object_or_404(
            Credito,
            id=credito_id,
            linea=Credito.LineaCredito.LIBRANZA
        )

        if credito.detalle_libranza.empresa != request.empresa:
            if wants_json:
                return JsonResponse({'error': 'Forbidden'}, status=403)
            messages.error(request, 'No tiene permisos para pagar este credito.')
            return redirect('pagador:dashboard')

        monto_decimal = Decimal(amount_in_cents) / 100
        client_ip = (request.META.get('HTTP_X_FORWARDED_FOR') or '').split(',')[0].strip() or request.META.get('REMOTE_ADDR')
        user_label = request.user.username if request.user.is_authenticated else 'anonymous'
        user_agent = (request.META.get('HTTP_USER_AGENT') or '')[:255]
        referer = (request.META.get('HTTP_REFERER') or '')[:255]
        request_id = (request.META.get('HTTP_X_REQUEST_ID') or '')[:64]
        logger.info(
            'Wompi intento pago: view=pagador credito=%s user=%s ip=%s ref=%s method=%s amount=%s req=%s',
            credito.id,
            user_label,
            client_ip,
            reference,
            payment_method_type,
            amount_in_cents,
            request_id
        )

        if payment_method_type not in ['CARD', 'PSE', 'NEQUI', 'BANCOLOMBIA_TRANSFER']:
            if wants_json:
                return JsonResponse({'error': 'Invalid payment method'}, status=400)
            messages.error(request, 'Metodo de pago no valido.')
            return redirect('pagador:dashboard')

        if not acceptance_token:
            if wants_json:
                return JsonResponse({'error': 'Missing acceptance token'}, status=400)
            messages.error(request, 'Falta acceptance token.')
            return redirect('pagador:dashboard')

        if reference and reference.startswith('CUOTA-'):
            parts = reference.split('-')
            if len(parts) < 3 or parts[1] != str(credito.id):
                if wants_json:
                    return JsonResponse({'error': 'Invalid reference'}, status=400)
                messages.error(request, 'Referencia de pago invalida.')
                return redirect('pagador:dashboard')
            try:
                cuota_num = int(parts[2])
            except (TypeError, ValueError):
                if wants_json:
                    return JsonResponse({'error': 'Invalid reference'}, status=400)
                messages.error(request, 'Referencia de pago invalida.')
                return redirect('pagador:dashboard')
            cuota = credito.tabla_amortizacion.filter(numero_cuota=cuota_num).first()
            if not cuota:
                if wants_json:
                    return JsonResponse({'error': 'Cuota not found'}, status=404)
                messages.error(request, 'La cuota indicada no existe.')
                return redirect('pagador:dashboard')
            if cuota.pagada:
                if wants_json:
                    return JsonResponse({'error': 'Cuota already paid'}, status=409)
                messages.warning(request, 'Esta cuota ya esta pagada.')
                return redirect('pagador:dashboard')

        rate_limit = getattr(settings, 'WOMPI_RATE_LIMIT_ATTEMPTS', 3)
        rate_window = getattr(settings, 'WOMPI_RATE_LIMIT_WINDOW_SECONDS', 60)
        attempt_key = f'wompi:attempts:pagador:{credito.id}:{client_ip}'
        attempts = cache.get(attempt_key, 0)
        if attempts >= rate_limit:
            if wants_json:
                return JsonResponse({'error': 'Rate limit'}, status=429)
            messages.warning(request, 'Demasiados intentos. Espera un momento y vuelve a intentar.')
            return redirect('pagador:credito_detalle', credito_id=credito.id)
        cache.set(attempt_key, attempts + 1, timeout=rate_window)

        cooldown_seconds = getattr(settings, 'WOMPI_DUPLICATE_COOLDOWN_SECONDS', 300)
        window_minutes = getattr(settings, 'WOMPI_DUPLICATE_WINDOW_MINUTES', 10)
        lock_key = f'wompi:lock:pagador:{credito.id}:{reference}:{amount_in_cents}'
        if not cache.add(lock_key, True, timeout=cooldown_seconds):
            logger.warning(
                'Pago duplicado bloqueado por lock: credito=%s user=%s ip=%s ref=%s',
                credito.id,
                user_label,
                client_ip,
                reference
            )
            if wants_json:
                return JsonResponse({'error': 'Duplicate payment'}, status=409)
            messages.warning(request, 'Ya hay un pago en proceso para este credito. Espera unos minutos y verifica el estado.')
            return redirect('pagador:credito_detalle', credito_id=credito.id)

        reciente = timezone.now() - timedelta(minutes=window_minutes)
        active_intent = WompiIntent.objects.filter(
            credito=credito,
            referencia=reference,
            status__in=[WompiIntent.Estado.CREATED, WompiIntent.Estado.PENDING]
        ).order_by('-created_at').first()
        if active_intent and active_intent.created_at >= reciente:
            if wants_json:
                return JsonResponse({'error': 'Payment already pending'}, status=409)
            messages.warning(request, 'Ya hay un pago en proceso para esta cuota.')
            return redirect('pagador:credito_detalle', credito_id=credito.id)
        if WompiIntent.objects.filter(
            credito=credito,
            referencia=reference,
            status=WompiIntent.Estado.APPROVED
        ).exists():
            if wants_json:
                return JsonResponse({'error': 'Payment already approved'}, status=409)
            messages.warning(request, 'Esta cuota ya fue pagada.')
            return redirect('pagador:credito_detalle', credito_id=credito.id)

        if HistorialPago.objects.filter(referencia_pago=reference).exists():
            if wants_json:
                return JsonResponse({'error': 'Payment already recorded'}, status=409)
            messages.warning(request, 'Ya registramos un pago para esta referencia.')
            return redirect('pagador:credito_detalle', credito_id=credito.id)

        if not customer_email:
            customer_email = getattr(credito.detalle_libranza, 'correo_electronico', '') or request.user.email

        client = WompiClient()
        redirect_url = request.build_absolute_uri(reverse('pagador:pago_wompi_callback'))

        intent = WompiIntent.objects.create(
            credito=credito,
            referencia=reference,
            amount_in_cents=amount_in_cents,
            payment_method=payment_method_type,
            status=WompiIntent.Estado.CREATED,
            usuario=request.user,
            ip_address=client_ip,
            user_agent=user_agent,
            referer=referer
        )

        if payment_method_type == 'CARD':
            card_token_response = client.tokenize_card(
                card_number=payload.get('card_number', '').replace(' ', ''),
                cvc=payload.get('cvc'),
                exp_month=payload.get('exp_month'),
                exp_year=payload.get('exp_year'),
                card_holder=payload.get('card_holder')
            )
            card_token = card_token_response['data']['id']

            payment_method = WompiClient.build_card_payment_method(
                token=card_token,
                installments=int(payload.get('installments', 1))
            )
            customer_data = None

        elif payment_method_type == 'PSE':
            payment_method = WompiClient.build_pse_payment_method(
                financial_institution_code=payload.get('financial_institution_code'),
                user_type=int(payload.get('user_type')),
                user_legal_id_type=payload.get('user_legal_id_type'),
                user_legal_id=payload.get('user_legal_id'),
                payment_description=f'Pago cuota {reference}'
            )
            customer_data = WompiClient.build_customer_data(
                phone_number=f"57{payload.get('phone_number')}",
                full_name=payload.get('full_name')
            )

        elif payment_method_type == 'NEQUI':
            payment_method = WompiClient.build_nequi_payment_method(
                phone_number=payload.get('nequi_phone')
            )
            customer_data = None

        elif payment_method_type == 'BANCOLOMBIA_TRANSFER':
            payment_method = WompiClient.build_bancolombia_transfer_payment_method(
                payment_description=f'Pago cuota {reference}'
            )
            customer_data = None
        else:
            if wants_json:
                return JsonResponse({'error': 'Invalid payment method'}, status=400)
            messages.error(request, 'Metodo de pago no valido.')
            return redirect('pagador:dashboard')

        transaction = client.create_transaction(
            amount_in_cents=amount_in_cents,
            currency='COP',
            customer_email=customer_email,
            payment_method=payment_method,
            reference=reference,
            acceptance_token=acceptance_token,
            redirect_url=redirect_url,
            customer_data=customer_data
        )

        transaction_data = transaction.get('data', {})
        transaction_id = transaction_data.get('id')
        transaction_status = transaction_data.get('status')
        if intent:
            intent.status = _map_wompi_status_to_intent(transaction_status)
            if transaction_id:
                intent.wompi_transaction_id = transaction_id
            intent.save(update_fields=['status', 'wompi_transaction_id', 'updated_at'])

        request.session['wompi_transaction_id'] = transaction_data.get('id')
        request.session['credito_id'] = credito_id
        request.session['reference'] = reference

        logger.info(f'Wompi transaction response: {transaction}')

        if payment_method_type in ['PSE', 'NEQUI', 'BANCOLOMBIA_TRANSFER']:
            payment_method_data = transaction_data.get('payment_method', {})
            extra_data = payment_method_data.get('extra', {})
            async_url = extra_data.get('async_payment_url')

            if not async_url:
                logger.warning(f'No async_payment_url en respuesta de Wompi. Payment method data: {payment_method_data}')
                wait_url = f"{reverse('pagador:pago_wompi_callback')}?wait=1&id={transaction_data.get('id')}"
                if wants_json:
                    return JsonResponse({'status': transaction_status, 'wait_url': wait_url})
                return redirect(wait_url)

            if wants_json:
                return JsonResponse({'status': transaction_status, 'async_url': async_url, 'transaction_id': transaction_id})
            return redirect(async_url)

        status = transaction_status
        if status == 'APPROVED':
            pago, created = credit_services.registrar_pago_credito(
                credito=credito,
                monto=monto_decimal,
                referencia_pago=reference,
                metodo_pago=HistorialPago.MetodoPago.WOMPI,
                origen_registro=HistorialPago.OrigenRegistro.PASARELA_WOMPI,
                usuario=request.user if request.user.is_authenticated else None,
                empresa=credito.empresa_relacionada,
                wompi_intento=intent,
                notas='Pago de cuota procesado por Wompi para pagador.',
            )
            if tipo_pago != 'MASIVO':
                credito.refresh_from_db()
                _enviar_resumen_pago_pagador(request, credito, transaction_data)
            if wants_json:
                receipt_url = reverse('pagador:pago_wompi_resumen', kwargs={'transaction_id': transaction_id})
                return JsonResponse({'status': 'APPROVED', 'transaction_id': transaction_id, 'receipt_url': receipt_url})
            messages.success(request, f'Pago de ${monto_decimal:,.2f} procesado exitosamente.')
            return redirect('pagador:pago_wompi_resumen', transaction_id=transaction_id)
        if status == 'DECLINED':
            if intent:
                intent.status = WompiIntent.Estado.DECLINED
                intent.save(update_fields=['status', 'updated_at'])
            if wants_json:
                return JsonResponse({'status': 'DECLINED', 'transaction_id': transaction_id})
            messages.error(request, 'El pago fue rechazado. Por favor intenta con otro metodo.')
            return redirect('pagador:credito_detalle', credito_id=credito.id)

        if wants_json:
            return JsonResponse({'status': status or 'PENDING', 'transaction_id': transaction_id})
        messages.warning(request, 'El pago esta pendiente de confirmacion.')
        return redirect('pagador:credito_detalle', credito_id=credito.id)

    except WompiAPIException as e:
        logger.error(f'Error en WOMPI: {str(e)}')
        if intent:
            intent.status = WompiIntent.Estado.ERROR
            intent.save(update_fields=['status', 'updated_at'])
        if wants_json:
            return JsonResponse({'error': str(e)}, status=400)
        messages.error(request, f'Error al procesar el pago: {str(e)}')
        return redirect('pagador:dashboard')
    except Exception as e:
        logger.error(f'Error inesperado: {str(e)}')
        if intent:
            intent.status = WompiIntent.Estado.ERROR
            intent.save(update_fields=['status', 'updated_at'])
        if wants_json:
            return JsonResponse({'error': 'Internal server error'}, status=500)
        messages.error(request, 'Ocurrio un error inesperado. Por favor intenta de nuevo.')
        return redirect('pagador:dashboard')


@require_http_methods(["GET"])
def pago_wompi_callback_view(request):
    """
    Callback despues de que el usuario completa el pago en WOMPI (PSE, Nequi, Bancolombia)
    """
    from ..services.wompi_client import WompiClient, WompiAPIException

    transaction_id = request.GET.get('id') or request.session.get('wompi_transaction_id')

    if not transaction_id:
        messages.error(request, 'No se encontro informacion de la transaccion.')
        return redirect('pagador:dashboard')

    client = WompiClient()

    try:
        wompi_transaction = client.get_transaction(transaction_id)
        transaction_data = wompi_transaction.get('data', {})
        status = transaction_data.get('status')

        if transaction_id:
            WompiIntent.objects.filter(wompi_transaction_id=transaction_id).update(
                status=_map_wompi_status_to_intent(status)
            )

        try:
            attempt = int(request.GET.get('attempt', 0))
        except (TypeError, ValueError):
            attempt = 0
        max_attempts = 12

        if status not in ['APPROVED', 'DECLINED']:
            if attempt < max_attempts:
                refresh_url = f"{reverse('pagador:pago_wompi_callback')}?wait=1&attempt={attempt + 1}&id={transaction_id}"
                return render(request, 'pagador/pago_wompi_espera.html', {
                    'refresh_url': refresh_url,
                    'attempts_left': max_attempts - attempt,
                })
            messages.warning(request, f'El pago esta en estado: {status}')
            request.session.pop('pagos_csv_pendientes', None)
            request.session.pop('wompi_transaction_id', None)
            request.session.pop('reference', None)
            request.session.pop('credito_id', None)
            return redirect('pagador:dashboard')

        pagos_csv_pendientes = request.session.get('pagos_csv_pendientes')

        if pagos_csv_pendientes:
            reference = request.session.get('reference', f"CSV-MASIVO-{timezone.now().strftime('%Y%m%d%H%M%S')}")

            if status == 'APPROVED':
                monto_total = Decimal(transaction_data.get('amount_in_cents', 0)) / 100
                pagos_exitosos = 0

                with transaction.atomic():
                    for pago_info in pagos_csv_pendientes:
                        credito = Credito.objects.filter(id=pago_info['credito_id']).first()
                        if credito:
                            monto_pago = Decimal(pago_info['monto'])

                            pago, created = credit_services.registrar_pago_credito(
                                credito=credito,
                                monto=monto_pago,
                                referencia_pago=f"{reference}-{credito.id}",
                                metodo_pago=HistorialPago.MetodoPago.WOMPI,
                                origen_registro=HistorialPago.OrigenRegistro.PASARELA_WOMPI,
                                usuario=request.user if request.user.is_authenticated else None,
                                empresa=credito.empresa_relacionada,
                                notas='Pago masivo procesado por Wompi.',
                            )

                            if created:
                                pagos_exitosos += 1

                messages.success(
                    request,
                    f'Pago masivo procesado exitosamente. Se aplicaron {pagos_exitosos} pagos por un total de ${monto_total:,.2f}'
                )
            elif status == 'DECLINED':
                messages.error(request, 'El pago fue rechazado. No se aplicaron los pagos del CSV.')
            else:
                messages.warning(request, f'El pago esta en estado: {status}')

            request.session.pop('pagos_csv_pendientes', None)
            request.session.pop('wompi_transaction_id', None)
            request.session.pop('reference', None)

            return redirect('pagador:dashboard')

        credito_id = request.session.get('credito_id')
        reference = request.session.get('reference') or transaction_data.get('reference')

        if not credito_id and reference and '-' in reference:
            parts = reference.split('-')
            if len(parts) >= 2:
                credito_id = parts[1]

        if not credito_id:
            messages.error(request, 'Sesion expirada. Por favor intenta de nuevo.')
            return redirect('pagador:dashboard')

        credito = get_object_or_404(Credito, id=credito_id)

        if status == 'APPROVED':
            monto_decimal = Decimal(transaction_data.get('amount_in_cents', 0)) / 100
            pago, created = credit_services.registrar_pago_credito(
                credito=credito,
                monto=monto_decimal,
                referencia_pago=reference,
                metodo_pago=HistorialPago.MetodoPago.WOMPI,
                origen_registro=HistorialPago.OrigenRegistro.PASARELA_WOMPI,
                usuario=request.user if request.user.is_authenticated else None,
                empresa=credito.empresa_relacionada,
                notas='Pago individual procesado por Wompi.',
            )
            credito.refresh_from_db()
            _enviar_resumen_pago_pagador(request, credito, transaction_data)
            messages.success(request, f'Pago de ${monto_decimal:,.2f} procesado exitosamente.')
        elif status == 'DECLINED':
            messages.error(request, 'El pago fue rechazado.')
        else:
            messages.warning(request, f'El pago esta en estado: {status}')

        request.session.pop('wompi_transaction_id', None)
        request.session.pop('credito_id', None)
        request.session.pop('reference', None)

        if status == 'APPROVED':
            return redirect('pagador:pago_wompi_resumen', transaction_id=transaction_id)
        return redirect('pagador:credito_detalle', credito_id=credito.id)

    except WompiAPIException as e:
        logger.error(f'Error al consultar transaccion: {str(e)}')
        messages.error(request, 'Error al verificar el estado del pago.')
        return redirect('pagador:dashboard')


@login_required(login_url='/pagador/login/')
@pagador_required
def pagador_pago_resumen_wompi_view(request, transaction_id):
    """
    Muestra el resumen de pago para el pagador.
    """
    from ..services.wompi_client import WompiClient, WompiAPIException

    client = WompiClient()
    try:
        wompi_transaction = client.get_transaction(transaction_id)
        transaction_data = wompi_transaction.get('data', {})
    except WompiAPIException as e:
        logger.error(f"Error consultando transaccion WOMPI {transaction_id}: {e}")
        messages.error(request, "No pudimos obtener el resumen del pago.")
        return redirect('pagador:dashboard')

    status = transaction_data.get('status')
    if status != 'APPROVED':
        messages.warning(request, f"El pago esta en estado: {status}")
        return redirect('pagador:dashboard')

    reference = transaction_data.get('reference')
    credito = _get_credito_pagador_from_reference(request, reference)
    if not credito:
        messages.error(request, "No se encontro el credito asociado al pago.")
        return redirect('pagador:dashboard')

    monto_pagado = Decimal(transaction_data.get('amount_in_cents', 0)) / 100
    metodo_pago, banco = _get_metodo_pago_wompi(transaction_data)
    fecha_pago = _parse_wompi_datetime(
        transaction_data.get('finalized_at') or transaction_data.get('created_at')
    ) or timezone.now()

    context = {
        'credito': credito,
        'pagador_nombre': request.user.get_full_name() or request.user.username,
        'pagador_email': request.user.email,
        'referencia_pago': reference,
        'transaction_id': transaction_id,
        'monto_pagado': monto_pagado,
        'metodo_pago': metodo_pago,
        'banco': banco,
        'fecha_pago': fecha_pago,
        'estado_pago': status,
        'saldo_pendiente': credito.saldo_pendiente,
        'fecha_proximo_pago': credito.fecha_proximo_pago,
        'comprobante_url': reverse('pagador:pago_wompi_comprobante', kwargs={'transaction_id': transaction_id}),
    }

    return render(request, 'pagador/pago_wompi_confirmado.html', context)


@login_required(login_url='/pagador/login/')
@pagador_required
def pagador_pago_comprobante_wompi_view(request, transaction_id):
    """
    Genera el comprobante PDF del pago para el pagador.
    """
    from ..services.wompi_client import WompiClient, WompiAPIException
    from django.template.loader import render_to_string
    from django.templatetags.static import static
    from weasyprint import HTML

    client = WompiClient()
    try:
        wompi_transaction = client.get_transaction(transaction_id)
        transaction_data = wompi_transaction.get('data', {})
    except WompiAPIException as e:
        logger.error(f"Error consultando transaccion WOMPI {transaction_id}: {e}")
        messages.error(request, "No pudimos generar el comprobante.")
        return redirect('pagador:dashboard')

    status = transaction_data.get('status')
    if status != 'APPROVED':
        messages.warning(request, f"El pago esta en estado: {status}")
        return redirect('pagador:dashboard')

    reference = transaction_data.get('reference')
    credito = _get_credito_pagador_from_reference(request, reference)
    if not credito:
        messages.error(request, "No se encontro el credito asociado al pago.")
        return redirect('pagador:dashboard')

    monto_pagado = Decimal(transaction_data.get('amount_in_cents', 0)) / 100
    metodo_pago, banco = _get_metodo_pago_wompi(transaction_data)
    fecha_pago = _parse_wompi_datetime(
        transaction_data.get('finalized_at') or transaction_data.get('created_at')
    ) or timezone.now()

    context = {
        'logo_url': request.build_absolute_uri(static('images/logo-dark.png')),
        'credito': credito,
        'pagador_nombre': request.user.get_full_name() or request.user.username,
        'pagador_email': request.user.email,
        'referencia_pago': reference,
        'transaction_id': transaction_id,
        'monto_pagado': monto_pagado,
        'metodo_pago': metodo_pago,
        'banco': banco,
        'fecha_pago': fecha_pago,
        'estado_pago': status,
        'saldo_pendiente': credito.saldo_pendiente,
        'fecha_proximo_pago': credito.fecha_proximo_pago,
    }

    html = render_to_string('pagador/pago_wompi_comprobante.html', context)
    pdf_bytes = HTML(string=html).write_pdf()

    filename = f"comprobante_{reference or transaction_id}.pdf"
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=\"{filename}\"'
    return response


@login_required(login_url='/pagador/login/')
@pagador_required
def iniciar_pago_masivo_wompi_view(request):
    """
    Muestra el formulario de selección de m?todo de pago con WOMPI para pagos masivos
    """
    from ..services.wompi_client import WompiClient, WompiAPIException

    if request.method != 'POST':
        messages.error(request, "Método no permitido.")
        return redirect('pagador:dashboard')

    # Obtener IDs de créditos seleccionados
    creditos_ids_str = request.POST.get('creditos_ids', '')
    if not creditos_ids_str:
        messages.error(request, "No se seleccionaron créditos para pagar.")
        return redirect('pagador:dashboard')

    try:
        creditos_ids = [int(id.strip()) for id in creditos_ids_str.split(',') if id.strip()]
    except ValueError:
        messages.error(request, "IDs de créditos inválidos.")
        return redirect('pagador:dashboard')

    if not creditos_ids:
        messages.error(request, "No se seleccionaron créditos válidos.")
        return redirect('pagador:dashboard')

    # Obtener créditos
    creditos = Credito.objects.filter(
        id__in=creditos_ids,
        linea=Credito.LineaCredito.LIBRANZA,
        detalle_libranza__empresa=request.empresa,
        estado=Credito.EstadoCredito.ACTIVO
    ).select_related('detalle_libranza')

    if not creditos.exists():
        messages.error(request, "No se encontraron créditos válidos para pagar.")
        return redirect('pagador:dashboard')

    # Calcular monto total
    monto_total = sum(c.valor_cuota for c in creditos if c.valor_cuota)
    if monto_total <= 0:
        messages.error(request, "El monto total a pagar es inválido.")
        return redirect('pagador:dashboard')

    # Obtener acceptance token de WOMPI
    client = WompiClient()
    try:
        acceptance_response = client.get_acceptance_token()
        acceptance_token = acceptance_response['data']['presigned_acceptance']['acceptance_token']

        # Obtener lista de bancos PSE
        bancos_pse = client.get_pse_financial_institutions()
    except WompiAPIException as e:
        logger.error(f"Error al obtener datos de WOMPI: {str(e)}")
        messages.error(request, "Error al conectar con la pasarela de pagos. Por favor intenta más tarde.")
        return redirect('pagador:dashboard')

    # Guardar en sesión
    request.session['creditos_ids_pago_masivo'] = creditos_ids
    request.session['monto_total_pago_masivo'] = str(monto_total)

    context = {
        'creditos': creditos,
        'cantidad_creditos': creditos.count(),
        'monto_total': int(monto_total),
        'monto_total_centavos': int(monto_total * 100),  # Convertir a centavos
        'referencia_pago': f"MASIVO-{timezone.now().strftime('%Y%m%d%H%M%S')}",
        'acceptance_token': acceptance_token,
        'bancos_pse': bancos_pse,
        'customer_email': request.empresa.correo_contacto if hasattr(request.empresa, 'correo_contacto') else request.user.email,
        'customer_name': request.empresa.nombre,
        'customer_phone': request.empresa.telefono if hasattr(request.empresa, 'telefono') else '',
        'wompi_public_key': settings.WOMPI_PUBLIC_KEY,
    }

    return render(request, 'pagador/pago_masivo_wompi.html', context)


@require_http_methods(["GET"])
def get_pse_banks_view(request):
    """
    API endpoint para obtener la lista de bancos PSE
    """
    from ..services.wompi_client import WompiClient, WompiAPIException

    client = WompiClient()
    try:
        banks = client.get_pse_financial_institutions()
        return JsonResponse(banks, safe=False)
    except WompiAPIException as e:
        return JsonResponse({'error': str(e)}, status=500)
